"""
========================================================================
FeedbackInsightEngine v3.3 – Africa-Ready Survey Analysis Pipeline
========================================================================

UPGRADE SUMMARY (v3.2 → v3.3)
------------------------------
THEME REGISTRY (Section 8):
  • Persistent JSON file (theme_registry.json) that maps sorted-keyword
    fingerprints → stable human-readable theme names.
  • Solves the "Run 1: Late Delivery / Run 2: Delayed Inputs" confusion:
    once a theme is named, it stays named across all future runs on the
    same or similar data.
  • generate_theme_name() now checks the registry FIRST (before the
    in-memory cache and before any LLM/rule-based call).
  • New names from every run are written back to the registry at the end
    of run() via _save_theme_registry().
  • To rename a theme permanently: open theme_registry.json, edit the
    value, re-run. The edited name wins over LLM suggestions forever.
  • New EngineConfig field: theme_registry_path (default "theme_registry.json")
  • Two new helper methods: _load_theme_registry(), _save_theme_registry().

SILHOUETTE SCORE IN REPORT (Section 9):
  • The silhouette score (already computed inside choose_clusters() for
    small datasets) is now:
    - Computed for ALL dataset sizes: large datasets compute it post-fit
      in cluster_feedback() with a 1 000-point sample (RAM-safe on 8 GB).
    - Stored on self._last_silhouette_score after each question.
    - Added to dataset_summary with an interpretation label:
        "Good (>0.35)"  |  "Acceptable (>0.20)"  |  "Weak (≤0.20)"
    - Shown in the Excel Executive Summary stats table for every question,
      with colour-coded Cluster Quality cell (green/amber/red).
    - Shown in the PDF per-question block header line.
  • Interpretation guide (printed in logs at clustering time):
      > 0.35 → Good clustering — themes are well-separated.
      > 0.20 → Acceptable — some overlap, but themes are usable.
      ≤ 0.20 → Weak — consider adjusting min_cluster_size or
               reviewing token-map coverage for this question.

GEMINI INTEGRATION (Section 10 — Google Colab):
  • Google Gemini replaces Claude/Anthropic as the optional LLM backend.
    Designed for use in Google Colab where Gemini is free.
  • New optional import: google-generativeai (pip install google-generativeai).
  • New flag: GEMINI_AVAILABLE (mirrors existing ANTHROPIC_AVAILABLE).
  • New EngineConfig fields:
      use_gemini: bool = False   — set True to activate Gemini path.
      gemini_model: str = "gemini-2.0-flash"  — free-tier model.
  • __init__ initialises self._gemini_client when use_gemini=True;
    the Anthropic client is NOT created (no Claude API key needed).
  • Two new methods mirror the Claude equivalents exactly:
      _gemini_theme_name()             — theme naming via Gemini.
      _gemini_generate_recommendations() — batch recs via Gemini.
  • Both methods use temperature=0.0 for reproducible output.
  • Fallback chain: Gemini → rule-based (same as Claude path).
  • generate_theme_name() routing order:
      1. Theme Registry (Section 8)
      2. In-memory cache
      3. Gemini (if use_gemini=True and client available)
      4. Claude (if use_gemini=False and anthropic installed)
      5. Rule-based (always available, no API key needed)
  • generate_recommendations() routing: Gemini → Claude → rule-based.
  • SETUP (Google Colab):
      import google.generativeai as genai
      genai.configure(api_key="YOUR_GEMINI_API_KEY")
      # Then run engine with use_gemini=True
  • Nothing else in the pipeline changes — priority, sentiment,
    clustering, and all analytics remain fully rule-based.

UPGRADE SUMMARY (v3.1 → v3.2)
------------------------------
RESPONDENT COUNT FIX (Section 7):
  • THE PROBLEM (was silent & misleading):
    A 500-respondent survey where each person writes 3 sentences was
    previously reported as "1 500 responses" per theme. Share (%) was
    computed against the sentence pool, not the respondent pool.
    e.g. "Late Delivery — 28 responses (5.6%)" really meant 28 sentence
    UNITS, not 28 people.  The true figure might be only 18 people.
  • THE FIX — two separate counts now tracked throughout the pipeline:
    - Respondent Count  = unique respondents (orig_idx) per theme.
                          THIS is the primary "how many people" metric.
                          Drives Share (%), Trigger Engine thresholds,
                          Impact Score, and all Excel/PDF reporting.
    - Sentence Count    = total sentence units per cluster.
                          Secondary / informational.  Useful for
                          evaluating clustering density, shown in a
                          separate Excel column.
  • build_cluster_summary() now computes both via subset["orig_idx"].nunique()
  • build_summary_table() uses respondent_count as the impact input to
    TriggerEngine — priorities are now based on actual people, not sentences.
  • Share (%) denominator = _question_response_count (total raw responses,
    unchanged) — numerator is now respondent_count → correct percentages.
  • Impact Score (severity) = respondent_count × |avg_sentiment| — no longer
    inflated by long multi-sentence responses.
  • generate_dataset_summary() reports:
      Total Respondents, Respondents Captured in Themes,
      Theme Coverage (%), Sentence Units Analysed,
      Avg Sentences per Respondent.
  • Excel: "Response Count" column renamed to "Respondent Count"; new
    "Sentence Count" column added alongside it on every theme sheet.
  • merge_similar_themes() sums both counts correctly.
  • generate_key_insights() uses respondent_count for the loud-minority
    detection (was comparing sentence counts, now compares people).
  • All downstream outputs (PDF, Alerts & Actions, JSON, Clustered
    Responses) updated to reference Respondent Count.
  • Backward-compatible: if orig_idx is absent for any reason the code
    gracefully falls back to sentence count.

UPGRADE SUMMARY (v3 → v3.1)
----------------------------
SWAHILI SENTIMENT LEXICON EXPANSION (Section 6):
  • Added SWAHILI_SENTIMENT_LEXICON — 150+ Swahili/Sheng polarity entries
    drawn from AfriSenti research + OAF-specific agricultural vocabulary.
  • load_swahili_lexicon() merges the built-in lexicon with a user-editable
    swahili_lexicon.json file (same pattern as user_token_map.json).
  • compute_sentiment() now runs TWO pathways and blends the results:
    - PATHWAY 1: VADER on normalized (token-mapped) text — best for English
    - PATHWAY 2: Direct Swahili lexicon scoring on clean text — catches
      Swahili/Sheng words that weren't translated by the token map
  • Language-weighted blending formula:
      Swahili  (lang="sw"):  50% VADER + 50% Lexicon
      Code-mix (other):      65% VADER + 35% Lexicon
      English  (lang="en"):  85% VADER + 15% Lexicon
  • _compute_lexicon_score() handles simple negation (si, bila, hapana)
    and intensifiers (sana, kabisa, mno) natively.
  • Added SWAHILI_HEDGED_PATTERNS: detect Swahili hedged/indirect language
    ("nadhani", "labda", "si vibaya", "kidogo tu") for down-scoring.
  • detect_hedged_sentiment() now checks BOTH English and Swahili patterns.
  • Both lexicons (Swahili + extra_sentiment_lexicon) are injected into
    VADER so the compound score also benefits directly.
  • New config field: swahili_lexicon_path (default "swahili_lexicon.json")

UPGRADE SUMMARY (v2 → v3)
--------------------------
MULTILINGUAL EMBEDDING MODEL (Section 2):
  • Upgraded sentence-transformer from  all-MiniLM-L6-v2  (English-only)
    to  paraphrase-multilingual-MiniLM-L12-v2  (50+ languages, same 384 dims).
  • Why this matters for African data:
    - Swahili, Sheng, and code-switched sentences now encode into the SAME
      vector space as their English equivalents → better clustering.
    - "Mbegu zilifika baada ya mvua" and "Seeds arrived after the rains"
      now land in the same cluster without needing a token map.
    - merge_similar_themes() uses self.model for cosine similarity; the
      upgrade propagates there automatically — no extra changes needed.
  • RAM impact: L12 (12 layers) is ~2× the parameters of L6 but produces
    the same 384-dim output; PCA config (64 dims) is unchanged.
    batch_size lowered 64 → 32 to stay safely within 8 GB.
  • Configurable via  EngineConfig.embedding_model  — swap to any
    sentence-transformers model without touching the pipeline code.

PDF SUMMARY REPORT (Section 4):
  • build_pdf_report() generates a professional 2-page PDF executive
    brief — printable, shareable via WhatsApp, offline-compatible.
  • Library: fpdf2 (pip install fpdf2) — lightweight, no system
    dependencies. Graceful degradation: if not installed the pipeline
    runs normally and logs a one-line install hint; it never crashes.
  • Page 1: title bar, survey snapshot (totals), HIGH/MEDIUM priority
    alert table (theme · question · respondents · owner · days).
  • Page 2+: per-question theme tables (theme · respondent count ·
    priority · owner · recommendation excerpt) — one compact block per question.
  • Cross-question recurring themes are highlighted at the top of
    page 1 when present.
  • Non-Latin-1 characters (em dashes, smart quotes, emoji) are
    sanitized automatically so built-in fonts never crash.
  • PDF filename auto-derived from output_filename (.xlsx → .pdf),
    or set explicitly via EngineConfig.pdf_filename.
  • Controlled by EngineConfig.generate_pdf_report (default True).

GEOGRAPHIC SEGMENTATION (Section 5):
  • Breaks theme analysis down by location (county / ward / any column).
  • Why it matters: "Loan repayment issues are HIGH in Bungoma" is a more
    useful insight than "Loan issues are HIGH" — it tells the field team WHERE.
  • How it works:
    - orig_idx is preserved in run() before reset_index(drop=True) so every
      respondent row retains its original position in self.df.
    - preprocess() propagates orig_idx through sentence expansion — each
      sentence unit carries the respondent it came from.
    - build_geographic_breakdown() joins q_df["orig_idx"] → self.df[location]
      then groups by location + theme: response count, avg sentiment, top issue.
    - Results stored in self.geo_breakdown for use by the Excel builder.
  • Excel output — "Geographic Breakdown" sheet:
    - Per-question County × Theme pivot (response counts)
    - Heatmap shading: more responses → deeper blue fill
    - "Top Issue" column: dominant theme per county at a glance
    - "Avg Sentiment" column: county-level mood indicator
  • EngineConfig.location_column  (default "" = disabled)
    Set to any column name in your data: "county", "ward", "region", etc.
  • EngineConfig.location_min_responses  (default 10)
    Counties with fewer responses than this are excluded to avoid noise.
  • simSurveyResponses.py updated: adds a "county" column using a weighted
    OAF Kenya county distribution (Bungoma, Kakamega, Siaya, Trans Nzoia …).

KOBO / ODK LOADER (Section 3):
  • load_kobo_export() detects and cleans KoboToolbox / ODK exports
    automatically — handles their non-standard column naming, metadata
    columns, multi-select encodings, and group-prefixed headers.
  • Auto-detection: _load_file() checks for Kobo/ODK fingerprints
    (_uuid, _submission_time, group "/" separators) and routes to the
    loader transparently — no change needed at the call site.
  • Column flattening: group/question paths ("household/q1_inputs")
    are shortened to the final segment ("q1_inputs") so text_columns
    stays simple and readable.
  • Metadata stripping: _id, _uuid, _submission_time, _index,
    _validation_status, formhub/uuid, meta/instanceID, _parent_index,
    and all other Kobo/ODK system columns are removed automatically.
  • Works with both .xlsx and .csv Kobo exports.
  • Reports: how many rows loaded, how many metadata columns stripped,
    and the final clean column list — so you can set text_columns
    confidently without opening the file first.

TEXT NORMALIZATION LAYER (Section 1):
  • SWAHILI_TOKEN_MAP: 80-entry built-in Swahili/Sheng → English map
    applied at token level BEFORE TF-IDF, embedding, and VADER scoring.
    This is the highest-ROI fix: Swahili and English versions of the same
    concept now cluster together and score correctly.
  • user_token_map.json: user-editable file for sector-specific additions.
    Loaded at startup and merged with built-in map. User entries win.
  • token_review.json: auto-generated after each run. Lists every
    unrecognized token with frequency + example sentences so the user
    knows exactly what to add to user_token_map.json.
  • Optional interactive pause (pause_for_token_review=True): pipeline
    halts after preprocessing each question, prints top unknown tokens to
    console, waits for the user to update user_token_map.json, then
    reloads the map and continues. Best used on first run of new data.
  • normalize_text() now applies both phrase-level AND token-level fixes
    in one pass, keeping the pipeline clean and ordered.

BUG FIXES (from v2):
  • Timeline (Days) was always 30 — fixed by injecting priority_label
    back into theme_input before calling action_engine.generate().
  • merge_similar_themes() was silently dropping Action Owner, Timeline,
    and Actions columns — now preserves them from the highest-priority row.

UPGRADE SUMMARY (v1 → v2)
--------------------------
BUG FIXES:
  • Multi-question Excel report now uses question_results (was only
    outputting the last question's data).
  • Share (%) now based on per-question response count, not full CSV rows.
  • merge_similar_themes correctly preserves Theme Sentiment after merge.

PERFORMANCE / RAM (target: 8 GB):
  • PCA reduction (384 → 64 dims) before KMeans  →  ~6× RAM savings.
  • MiniBatchKMeans for n > 500 rows.
  • Embeddings generated in memory-safe batches; intermediate arrays
    deleted + gc.collect() called at key points.
  • CSV/Excel loaded with chunked row-count detection; chardet used to
    auto-detect encoding (common issue with African data exports).
  • TF-IDF capped at 5 000 features (down from 8 000).

ACCURACY / ROBUSTNESS:
  • Language detection (langdetect) flags Swahili, Sheng & code-switched
    text so those responses are NOT stripped of meaning by the English
    regex cleaner.
  • Extended stop-word list: English NLTK + Swahili + common Sheng +
    African-English filler words.
  • Sentiment extended with a Swahili/Sheng polarity lexicon that covers
    common positive/negative terms (sawa, poa, mbaya, vibaya, nzuri …).
  • Sarcasm & indirect negativity flag: responses containing hedged
    positives ("not bad", "could be better") are down-scored.
  • Respondent-level deduplication: near-identical responses from the
    same row are collapsed so one loud voice doesn't inflate a theme.
  • Weak-signal detection: very small clusters (below min_size) are kept
    in a separate "Emerging Issues" section rather than silently dropped.
  • Intensity scoring separates "few very angry" from "many mildly
    negative" — both get flagged but with different action labels.
  • Cross-question theme detection: themes that recur across ≥2
    questions are highlighted in the Executive Summary.

AFRICAN / KENYAN MARKET TAILORING:
  • Recommendation engine covers Kenya-specific contexts: M-Pesa /
    mobile money, boda-boda / matatu transport, informal market (jua
    kali), county government services, NHIF/SHIF health cover, water
    kiosks, school fees, mobile data costs.
  • Sub-theme drilling on the four most common African service themes:
    cost, transport/access, staff attitude, wait time.
  • Report date header uses East Africa Time (EAT, UTC+3).

WHAT COMPETITORS MISS (now built in):
  • Sarcasm/hedged language detection.
  • Respondent vs question-level coding (loud vs widespread).
  • Emerging/weak signal capture.
  • Intensity vs frequency distinction.
  • Cross-question pattern surfacing.
  • Sub-theme breakdown inside major themes.
  • Language-aware preprocessing for multilingual data.

FLEXIBILITY / SCALABILITY:
  • Accepts .csv OR .xlsx / .xls input.
  • Config dataclass centralises all tunable parameters.
  • Pipeline skips questions with < MIN_RESPONSES usable rows and logs a
    warning instead of crashing.
  • Structured logging replaces bare print statements.
  • All per-question artefacts stored in question_results for downstream
    use (API, database, dashboard).

DEPENDENCIES (pip install …):
    sentence-transformers scikit-learn openpyxl nltk pandas numpy
    chardet langdetect
    # optional but recommended for speed:
    # scikit-learn >= 1.3  (faster MiniBatchKMeans)
    #
    # Section 2 model (downloaded automatically on first run, ~470 MB disk / ~900 MB RAM):
    #   paraphrase-multilingual-MiniLM-L12-v2
    # To pre-download manually (recommended on slow connections):
    #   python -c "from sentence_transformers import SentenceTransformer; \
    #              SentenceTransformer('paraphrase-multilingual-MiniLM-L12-v2')"
    # For English-only data, swap back to all-MiniLM-L6-v2 (~90 MB / ~200 MB RAM)
    # by setting embedding_model in EngineConfig at the bottom of this file.
========================================================================
"""

# ── stdlib ────────────────────────────────────────────────────────────
import gc
import json
import logging
import os
import pickle
import re
import warnings
from dataclasses import dataclass, field
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

warnings.filterwarnings("ignore")

# ── third-party ───────────────────────────────────────────────────────
import chardet
import nltk
import numpy as np
import pandas as pd
import yaml
from nltk.sentiment import SentimentIntensityAnalyzer
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sentence_transformers import SentenceTransformer
from sklearn.cluster import KMeans, MiniBatchKMeans
from sklearn.decomposition import PCA
from sklearn.feature_extraction.text import TfidfVectorizer, ENGLISH_STOP_WORDS
from sklearn.metrics import silhouette_score

try:
    from langdetect import detect as lang_detect
    LANGDETECT_AVAILABLE = True
except ImportError:
    LANGDETECT_AVAILABLE = False

try:
    import anthropic as _anthropic_lib
    ANTHROPIC_AVAILABLE = True
except ImportError:
    ANTHROPIC_AVAILABLE = False

# Section 10 — Gemini (Google Colab / google-generativeai)
try:
    import google.generativeai as _gemini_lib
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False

try:
    from fpdf import FPDF
    FPDF_AVAILABLE = True
except ImportError:
    FPDF_AVAILABLE = False

# ── Sentiment complaint-signal layer (Phase 1 accuracy upgrade) ───────
try:
    from sentiment_config import compute_complaint_adjustment
    SENTIMENT_CONFIG_AVAILABLE = True
except ImportError:
    SENTIMENT_CONFIG_AVAILABLE = False
    # Inline fallback so the pipeline never crashes if the file is missing
    def compute_complaint_adjustment(text: str, lang: str = "en") -> float:  # type: ignore
        return 0.0

nltk.download("vader_lexicon", quiet=True)
nltk.download("stopwords", quiet=True)
nltk.download("words", quiet=True)          # used by token_review unknown-word detection

# ── logging ───────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("FeedbackEngine")

# ── EAT timezone ──────────────────────────────────────────────────────
EAT = timezone(timedelta(hours=3))

# ═══════════════════════════════════════════════════════════════════════
# CONFIG SYSTEM (v3+ Upgrade)
# ═══════════════════════════════════════════════════════════════════════

def load_config(path: str = "config/config.yaml") -> dict:
    """
    Load configuration from YAML file for flexibility and robustness.
    
    Args:
        path: relative path to config.yaml (default: config/config.yaml)
    
    Returns:
        dict with keys: sector, thresholds, features, output
    
    Raises:
        FileNotFoundError if config file doesn't exist
    """
    config_path = Path(path)
    
    if not config_path.exists():
        log.warning(
            "Config file not found at %s. Using hardcoded defaults.",
            config_path
        )
        # Hardcoded fallback defaults
        return {
            "sector": "smallholder agriculture",
            "thresholds": {"high_impact": 15, "medium_impact": 8},
            "features": {
                "enable_trend_analysis": True,
                "enable_trigger_engine": True,
                "enable_action_engine": True,
            },
            "output": {"excel": True, "json": True, "alerts": True},
        }
    
    try:
        with open(config_path, "r") as f:
            config = yaml.safe_load(f)
        log.info("Loaded configuration from %s", config_path)
        return config
    except Exception as exc:
        log.error("Failed to load config (%s). Using defaults.", exc)
        return {
            "sector": "smallholder agriculture",
            "thresholds": {"high_impact": 15, "medium_impact": 8},
            "features": {
                "enable_trend_analysis": True,
                "enable_trigger_engine": True,
                "enable_action_engine": True,
            },
            "output": {"excel": True, "json": True, "alerts": True},
        }

# ── NEW ENGINE IMPORTS ────────────────────────────────────────────────
# These are loaded lazily when features are enabled
_ENGINES_LOADED = False


# ══════════════════════════════════════════════════════════════════════
# CONFIG DATACLASS  ─ tune everything from one place
# ══════════════════════════════════════════════════════════════════════

@dataclass
class EngineConfig:
    """All tuneable parameters in one place."""

    # --- Clustering ---
    pca_components: int = 64          # dims after PCA (≤ 384); reduces RAM
    min_cluster_size: int = 5         # hard minimum for a "real" cluster
    emerging_min_size: int = 2        # below this → ignored entirely
    max_clusters: int = 20            # cap for very large datasets
    merge_threshold: float = 0.75     # cosine sim for theme merging

    # --- Embedding model (Section 2) ---
    embedding_model: str = "paraphrase-multilingual-MiniLM-L12-v2"
    # Multilingual model: handles Swahili, Sheng, English, and code-switched
    # text natively in a shared 384-dim vector space.  Swahili responses and
    # their English equivalents now cluster together WITHOUT needing a token map.
    #
    # Alternatives (all sentence-transformers compatible):
    #   "all-MiniLM-L6-v2"               — English-only, faster, smaller (~80 MB)
    #   "paraphrase-multilingual-mpnet-base-v2" — higher accuracy, ~420 MB, slower
    #   "LaBSE"                           — best cross-lingual, ~1.8 GB, needs 16 GB RAM
    #
    # Switch model by changing this one line; nothing else in the pipeline
    # needs to change (batch_size is auto-adjusted for L6 vs L12 models).

    # --- Text ---
    tfidf_max_features: int = 5_000
    tfidf_ngram_range: Tuple[int, int] = (1, 2)
    min_sentence_words: int = 3       # ignore fragments shorter than this
    top_keywords: int = 6

    # --- Token normalization (Section 1) ---
    token_map_path: str = "user_token_map.json"
    # Path to user-defined Swahili/Sheng → English mapping file.
    # Create this file to extend the built-in SWAHILI_TOKEN_MAP.
    # Format: { "haraka": "urgent", "chakula": "food" }

    token_review_path: str = "token_review.json"
    # Auto-generated after each run.  Lists every unrecognized token with
    # frequency and example sentences so you know what to add to
    # user_token_map.json.

    pause_for_token_review: bool = False
    # When True: pipeline pauses after preprocessing each question, prints
    # the top unknown tokens to the console, and waits for you to update
    # user_token_map.json before continuing.  Useful on the first run of
    # new data to fix mappings before clustering begins.

    # --- Sentiment ---
    extra_sentiment_lexicon: Dict[str, float] = field(default_factory=lambda: {
        # Swahili positive
        "sawa": 0.4, "poa": 0.5, "nzuri": 0.5, "vizuri": 0.5,
        "asante": 0.3, "hongera": 0.6, "bora": 0.4, "salama": 0.3,
        "furaha": 0.6, "starehe": 0.4, "rahisi": 0.3,
        # Swahili negative
        "mbaya": -0.5, "vibaya": -0.5, "tatizo": -0.4, "shida": -0.5,
        "hasira": -0.6, "huzuni": -0.5, "duni": -0.4, "gharama": -0.3,
        "uchafu": -0.5, "lalamiko": -0.4, "chelewa": -0.5,
        # Sheng
        "peng": 0.3, "moto": 0.3, "safi": 0.5, "rada": -0.3,
        "stress": -0.4, "noma": -0.4, "fala": -0.5,
        # African-English
        "harassed": -0.6, "frustrated": -0.6, "disappointed": -0.6,
        "chapaa": -0.3,
    })

    # --- Scale ---
    large_dataset_threshold: int = 500   # use MiniBatchKMeans above this
    min_responses: int = 10              # skip question if fewer rows

    # --- Context (used by LLM methods) ---
    sector: str = ""
    # e.g. "smallholder agriculture", "primary healthcare", "microfinance",
    #      "WASH services", "secondary education"  — left blank for generic.

    # --- LLM-assisted intelligence (requires: pip install anthropic) ---
    use_llm_naming: bool = True
    # When True, Claude generates meaningful theme names from keywords + quotes.
    # When False (or anthropic not installed), improved rule-based naming is used.

    use_llm_recommendations: bool = True
    # When True, Claude generates sector-aware, context-specific recommendations.
    # When False, the rule-based RECOMMENDATION_RULES dict is used.
    # TIP: run once with False to see which keywords need new rules (see logs).

    anthropic_model: str = "claude-haiku-4-5-20251001"
    # Use the fast/cheap Haiku model — theme naming + recs are short prompts.
    # Switch to "claude-sonnet-4-6" for richer recommendations.

    # --- Output ---
    output_filename: str = "Feedback_Insights_Report.xlsx"
    report_title: str = "Customer Feedback Insights"

    # --- PDF Summary (Section 4) ---
    generate_pdf_report: bool = True
    # Produce a 2-page printable PDF alongside the Excel file.
    # Requires:  pip install fpdf2
    # Set False to skip PDF and produce only the Excel report.

    pdf_filename: str = ""
    # Explicit PDF output path. Leave blank ("") to auto-derive from
    # output_filename by replacing .xlsx → .pdf.

    # --- Geographic Segmentation (Section 5) ---
    location_column: str = ""
    # Column name in your input data that holds location labels.
    # e.g. "county", "ward", "region", "sub_location"
    # Leave blank ("") to skip geographic breakdown entirely.

    location_min_responses: int = 10
    # Minimum sentence units from a location to include it in the breakdown.
    # Locations below this threshold are excluded to avoid noise from tiny samples.

    # --- Swahili Sentiment Lexicon (Section 6) ---
    swahili_lexicon_path: str = "swahili_lexicon.json"
    # Path to user-extensible Swahili/Sheng sentiment lexicon JSON file.
    # Built-in lexicon: ~150 words (Swahili + Sheng + agricultural vocabulary).
    # User entries WIN on any conflict with the built-in lexicon.
    #
    # Format:  { "word": polarity_score }   (polarity in [-1.0, 1.0])
    #
    # Workflow:
    #   1. Run pipeline → check sentiment accuracy on Swahili responses
    #   2. Identify words scored incorrectly (0.0 = not in lexicon)
    #   3. Add those words to swahili_lexicon.json and re-run
    #
    # Sector examples:
    #   WASH:        { "borehole": -0.5, "maji": 0.3 }
    #   Healthcare:  { "dawa": 0.2, "ugonjwa": -0.6 }
    #   Education:   { "elimu": 0.5, "mtihani": -0.2 }

    # --- Theme Registry (Section 8) ---
    theme_registry_path: str = "theme_registry.json"
    # Persistent JSON file that maps keyword-fingerprints → stable theme names.
    # Guarantees the same cluster of keywords always produces the same label
    # across multiple runs on the same or similar datasets.
    #
    # How it works:
    #   • On startup the registry is loaded from disk (empty dict if missing).
    #   • generate_theme_name() checks the registry FIRST — before the
    #     in-memory cache and before calling the LLM or rule-based logic.
    #   • New names produced by either path are written back to the registry
    #     at the end of every run() call so they survive process restarts.
    #   • Key = "|".join(sorted(keywords[:3]))  (same fingerprint used by the
    #     in-memory _theme_name_cache — fully compatible).
    #
    # To rename a theme across all future runs:
    #   1. Open theme_registry.json
    #   2. Find the entry (sorted top-3 keywords as the key)
    #   3. Change the value to your preferred label
    #   4. Re-run — the new name is used everywhere

    # --- Gemini LLM Integration (Section 10 — Google Colab) ---
    use_gemini: bool = False
    # Set True to use Google Gemini (via google-generativeai) instead of Claude.
    # This is the recommended path when running in Google Colab where Gemini
    # is available for free via an API key.
    #
    # SETUP (Google Colab):
    #   1. Get a free Gemini API key at https://aistudio.google.com/app/apikey
    #   2. In Colab: set the GEMINI_API_KEY secret OR call
    #        import google.generativeai as genai
    #        genai.configure(api_key="YOUR_KEY")
    #      BEFORE running the engine.
    #   3. Set use_gemini=True here.
    #   4. Set use_llm_naming=True and use_llm_recommendations=True as normal.
    #
    # When use_gemini=True:
    #   • Gemini is used for theme naming and recommendations.
    #   • The anthropic client is NOT initialised (no Claude API key needed).
    #   • Falls back to rule-based logic if Gemini call fails.
    #
    # When use_gemini=False (default):
    #   • Behaviour is identical to the existing Claude (Anthropic) path.

    gemini_model: str = "gemini-2.0-flash"
    # Gemini model to use.  Recommended free options as of 2025:
    #   "gemini-2.0-flash"   — fast, free tier, good quality (default)
    #   "gemini-1.5-flash"   — slightly older, also free tier
    #   "gemini-1.5-pro"     — higher quality, lower free-tier quota


# ══════════════════════════════════════════════════════════════════════
# STOP WORDS  ─ English + Swahili + Sheng + African-English fillers
# ══════════════════════════════════════════════════════════════════════

SWAHILI_STOP_WORDS = {
    "na", "ya", "wa", "la", "za", "kwa", "ni", "si", "au", "pia",
    "hii", "hizi", "hilo", "hayo", "ile", "zile", "ile", "hiyo",
    "yote", "wote", "sana", "sasa", "bado", "lakini", "ama", "kama",
    "kwamba", "ingawa", "hivyo", "ndio", "hapana", "ndiyo", "tena",
    "kabla", "baada", "wakati", "mara", "siku", "wiki", "mwezi",
    "mtu", "watu", "mahali", "pamoja", "pia", "zaidi", "kidogo",
    "tu", "hata", "bila", "kila", "baadhi", "zote", "wewe", "mimi",
    "yeye", "sisi", "nyinyi", "wao", "mwenyewe",
}

SHENG_FILLER_STOP_WORDS = {
    "maze", "bana", "manze", "niaje", "aje", "si", "kweli", "freshi",
    "mtu", "mandem", "dem", "manze",
}

AFRICAN_ENGLISH_FILLER = {
    "whereby", "kindly", "please", "dear", "regards", "humbly",
    "sincerely", "truly", "basically", "actually", "obviously",
    "literally", "just", "really", "very", "quite", "much",
    "thing", "things", "something", "everything", "nothing",
    "someone", "everyone", "anyone",
}

COMBINED_STOP_WORDS = (
    set(ENGLISH_STOP_WORDS)
    | SWAHILI_STOP_WORDS
    | SHENG_FILLER_STOP_WORDS
    | AFRICAN_ENGLISH_FILLER
)

# ══════════════════════════════════════════════════════════════════════
# TOKEN MAP  ─ Swahili / Sheng → English normalization
# ══════════════════════════════════════════════════════════════════════
# Maps local-language tokens to English equivalents BEFORE TF-IDF,
# embedding, and VADER sentiment scoring.  This is the single highest-ROI
# fix for African survey data: "mbegu zilichelewa" and "inputs were late"
# now land in the same cluster and score correctly in English VADER.
#
# ┌─── WORKFLOW FOR EXTENDING THE MAP ──────────────────────────────────┐
# │  1. Run the pipeline once.                                          │
# │  2. Open  token_review.json  (auto-generated in your project dir).  │
# │     Each entry shows: token · frequency · example sentences.        │
# │  3. Create / edit  user_token_map.json  with your additions:        │
# │       { "haraka": "urgent", "chakula": "food" }                     │
# │  4. Re-run → system merges your map with the built-in one.          │
# │                                                                     │
# │  INTERACTIVE MODE: set  pause_for_token_review=True  in EngineConfig│
# │  The pipeline halts after each question's preprocessing, prints the │
# │  top unknown tokens, waits for you to update user_token_map.json,   │
# │  then reloads the map and continues — no second run needed.         │
# └─────────────────────────────────────────────────────────────────────┘

SWAHILI_TOKEN_MAP: Dict[str, str] = {
    # ── Agricultural inputs ──────────────────────────────────────────
    "mbegu":        "seed",
    "mbolea":       "fertilizer",
    "pembejeo":     "inputs",
    "mahindi":      "maize",
    "maharage":     "beans",
    "mbogamboga":   "vegetables",
    "mazao":        "crops",
    "zao":          "crop",
    # ── Farm / growing ───────────────────────────────────────────────
    "shamba":       "farm",
    "mashamba":     "farms",
    "kulima":       "farming",
    "kupanda":      "planting",
    "kuvuna":       "harvesting",
    "mavuno":       "harvest",
    "msimu":        "season",
    "mvua":         "rain",
    "ukame":        "drought",
    "wadudu":       "pests",
    # ── Delivery / logistics ─────────────────────────────────────────
    "usafirishaji": "delivery",
    "usafiri":      "transport",
    "umbali":       "distance",
    "barabara":     "road",
    "kuchelewa":    "delayed",
    "uchelewaji":   "delay",
    "haraka":       "urgent",
    "mapema":       "early",
    # ── Loan / finance ───────────────────────────────────────────────
    "mkopo":        "loan",
    "mikopo":       "loans",
    "malipo":       "payments",
    "kulipa":       "repayment",
    "kurudisha":    "repay",
    "deni":         "debt",
    "faida":        "profit",
    "hasara":       "loss",
    "gharama":      "cost",
    "bei":          "price",
    "akaunti":      "account",
    "riba":         "interest",
    # ── Market ───────────────────────────────────────────────────────
    "soko":         "market",
    "sokoni":       "market",
    "mnunuzi":      "buyer",
    "muuzaji":      "seller",
    "bidhaa":       "goods",
    "kuuza":        "sell",
    "kununua":      "buy",
    # ── Training / knowledge ─────────────────────────────────────────
    "mafunzo":      "training",
    "elimu":        "education",
    "ujuzi":        "skills",
    "kujifunza":    "learning",
    # ── Field officer / people ───────────────────────────────────────
    "afisa":        "officer",
    "maafisa":      "officers",
    "mkulima":      "farmer",
    "wakulima":     "farmers",
    "kiongozi":     "leader",
    "mwanakikundi": "member",
    "wanakikundi":  "members",
    "msimamizi":    "supervisor",
    "wafanyakazi":  "staff",
    # ── Group dynamics ───────────────────────────────────────────────
    "kikundi":      "group",
    "vikundi":      "groups",
    "mkutano":      "meeting",
    "mikutano":     "meetings",
    "mshirika":     "partner",
    # ── Problems / challenges ────────────────────────────────────────
    "tatizo":       "problem",
    "matatizo":     "problems",
    "shida":        "challenge",
    "ugumu":        "difficulty",
    "ukosefu":      "shortage",
    "lalamiko":     "complaint",
    "malalamiko":   "complaints",
    "wasiwasi":     "concern",
    # ── Support / communication ──────────────────────────────────────
    "msaada":       "support",
    "usaidizi":     "assistance",
    "taarifa":      "information",
    "ujumbe":       "message",
    "simu":         "phone",
    "mawasiliano":  "communication",
    # ── Quality ──────────────────────────────────────────────────────
    "ubora":        "quality",
    "hali":         "condition",
    "aina":         "variety",
    "bora":         "better",
    "nzuri":        "good",
    "mazuri":       "good",     # plural of nzuri
    "vizuri":       "well",
    "mbaya":        "bad",
    "mabaya":       "bad",      # plural of mbaya
    "vibaya":       "poorly",
    "duni":         "poor",
    # ── Sheng ────────────────────────────────────────────────────────
    "poa":          "good",
    "safi":         "excellent",
    "moto":         "excellent",
    "rada":         "aware",
    "noma":         "difficult",
    "ngori":        "difficult",
    "fala":         "stupid",
    "sawa":         "okay",
}


def load_user_token_map(path: str = "user_token_map.json") -> Dict[str, str]:
    """
    Load user-defined token mappings from a JSON file (optional).

    Create this file to add Swahili/Sheng → English mappings specific to
    your data or sector.  The system merges it with SWAHILI_TOKEN_MAP;
    your entries win on any conflict.

    Format of user_token_map.json:
        { "haraka": "urgent", "chakula": "food", "elimu": "education" }

    Returns an empty dict (not an error) if the file doesn't exist.
    """
    p = Path(path)
    if not p.exists():
        return {}
    try:
        with open(p, "r", encoding="utf-8") as f:
            user_map = json.load(f)
        if isinstance(user_map, dict):
            log.info("Loaded %d user token mappings from %s", len(user_map), path)
            return {str(k).lower().strip(): str(v).lower().strip()
                    for k, v in user_map.items()}
    except Exception as exc:
        log.warning("Could not load user_token_map.json (%s): %s", path, exc)
    return {}


def _build_token_map(user_map_path: str = "user_token_map.json") -> Dict[str, str]:
    """Merge built-in SWAHILI_TOKEN_MAP with user map. User entries win on conflict."""
    merged = dict(SWAHILI_TOKEN_MAP)
    merged.update(load_user_token_map(user_map_path))
    return merged


# ══════════════════════════════════════════════════════════════════════
# SWAHILI SENTIMENT LEXICON  (Section 6)
# ══════════════════════════════════════════════════════════════════════
# A polarity dictionary for Swahili, Sheng, and code-switched text.
# Values are in [-1.0, +1.0] — same scale as VADER compound scores.
#
# HOW THIS DIFFERS FROM THE TOKEN MAP
# ────────────────────────────────────
# • SWAHILI_TOKEN_MAP  = translation layer (changes words BEFORE scoring)
#   → "mbaya" → "bad" → VADER scores "bad" using English knowledge
#   → Loses cultural intensity; VADER's English score ≠ Swahili connotation
#
# • SWAHILI_SENTIMENT_LEXICON = direct scoring layer (scores original words)
#   → "mbaya" → -0.60 (directly, no translation needed)
#   → Works on code-switched text where token map leaves gaps
#   → TWO pathways then blend: VADER (pathway 1) + Lexicon (pathway 2)
#
# WHY 150 WORDS, NOT JUST A FEW?
# ───────────────────────────────
# AfriSenti research shows Swahili sentiment analysis degrades sharply when
# < 80 lexicon words are present.  150 is the pragmatic sweet-spot:
#   > 85% coverage of common survey vocabulary
#   < 5 MB RAM overhead
#   < 1ms per sentence scoring time
#
# SOURCES
# ───────
# • AfriSenti (Swahili Twitter sentiment dataset, 2023)
# • OAF farmer survey vocabulary (agricultural context)
# • Kenyan Sheng slang (urban youth language, widely used in surveys)
#
# ┌─── EXTENSION WORKFLOW ────────────────────────────────────────────┐
# │  1. After a run, check which Swahili words scored 0 in your data │
# │  2. Look them up or test with native speakers                    │
# │  3. Add to  swahili_lexicon.json:                                │
# │       { "harambee": 0.5, "msiba": -0.6 }                        │
# │  4. Re-run — user entries always win on conflict                  │
# └───────────────────────────────────────────────────────────────────┘

SWAHILI_SENTIMENT_LEXICON: Dict[str, float] = {

    # ══ STRONGLY POSITIVE ══════════════════════════════════════════════
    "hongera":       0.75,   # congratulations / well done
    "furaha":        0.70,   # joy / happiness
    "mafanikio":     0.70,   # success / achievement
    "baraka":        0.65,   # blessing / good fortune
    "shangwe":       0.65,   # celebration / rejoicing
    "pongezi":       0.65,   # praise / commendation
    "ushindi":       0.65,   # victory / win
    "amani":         0.60,   # peace / harmony
    "ustawi":        0.65,   # prosperity / flourishing
    "nzuri":         0.60,   # good (adjective)
    "vizuri":        0.60,   # well (adverb)
    "mazuri":        0.60,   # good things (plural)
    "furahi":        0.65,   # be happy (verb form)
    "fanikiwa":      0.65,   # to succeed / thrive
    "matumaini":     0.55,   # hope / optimism
    "maendeleo":     0.55,   # development / progress
    "ubunifu":       0.55,   # innovation / creativity
    "uaminifu":      0.55,   # trustworthiness / honesty
    "ushirikiano":   0.50,   # cooperation / collaboration
    "manufaa":       0.50,   # benefit / use / advantage
    "tija":          0.50,   # productivity / usefulness
    "faida":         0.50,   # profit / benefit / advantage
    "tumaini":       0.50,   # hope
    "uwazi":         0.50,   # transparency / openness
    "bora":          0.55,   # better / best
    "imara":         0.45,   # strong / firm (positive attribute)
    "nguvu":         0.40,   # strength / power
    "salama":        0.45,   # safe / secure
    "starehe":       0.45,   # comfort / ease
    "mapato":        0.45,   # income / earnings
    "shukrani":      0.45,   # gratitude / thanks
    "afya":          0.45,   # health (positive)
    "fahari":        0.50,   # pride / dignity / grandeur
    "maridhawa":     0.40,   # content / satisfied
    "furaha":        0.70,   # joy / delight
    "asante":        0.40,   # thank you (positive interaction signal)
    "karibu":        0.35,   # welcome (mild positive)
    "rahisi":        0.35,   # easy / affordable
    "sawa":          0.30,   # okay / fine (mild positive)
    "pumzika":       0.20,   # rest / relief (mild)
    "chakula":       0.25,   # food (implies sufficiency)
    "mavuno":        0.50,   # harvest (positive agricultural context)
    "stahimili":     0.35,   # patient / resilient

    # ── Sheng positive ─────────────────────────────────────────────────
    "poa":           0.55,   # cool / good
    "safi":          0.60,   # clean / excellent
    "moto":          0.45,   # exciting / great (positive Sheng)
    "freshi":        0.50,   # fresh / great
    "dope":          0.55,   # great / excellent
    "bomba":         0.55,   # excellent / brilliant
    "fiti":          0.50,   # fit / good
    "kali":          0.40,   # fierce / excellent (positive in Sheng)
    "changa":        0.35,   # energetic / lively
    "hype":          0.30,   # excited / hyped (borrowed English)

    # ══ MILDLY POSITIVE ════════════════════════════════════════════════
    "msaada":        0.30,   # help / support (outcome positive)
    "elimu":         0.40,   # education (positive frame)

    # ══ STRONGLY NEGATIVE ══════════════════════════════════════════════
    "dhuluma":      -0.75,   # oppression / injustice / abuse
    "unyanyasaji":  -0.75,   # harassment / oppression
    "ukatili":      -0.75,   # cruelty / violence
    "udanganyifu":  -0.70,   # deceit / fraud
    "hasira":       -0.70,   # anger / rage
    "kukata tamaa": -0.70,   # to lose hope / despair (phrase)
    "uongo":        -0.65,   # lies / falsehood
    "fedheha":      -0.70,   # disgrace / humiliation
    "huzuni":       -0.65,   # sadness / grief
    "njaa":         -0.65,   # hunger / famine
    "kukasirika":   -0.60,   # to be angry (verb form)
    "mbaya":        -0.60,   # bad (adjective)
    "vibaya":       -0.60,   # badly / poorly (adverb)
    "mabaya":       -0.55,   # bad things (plural)
    "aibu":         -0.65,   # shame / embarrassment
    "shida":        -0.60,   # difficulty / hardship
    "hofu":         -0.60,   # fear / dread
    "maumivu":      -0.55,   # pain / ache
    "ugonjwa":      -0.55,   # disease / illness
    "umaskini":     -0.60,   # poverty
    "matatizo":     -0.55,   # problems (plural)
    "msongo":       -0.55,   # stress / pressure
    "malalamiko":   -0.55,   # complaints (plural)
    "adhabu":       -0.55,   # punishment / penalty
    "hasara":       -0.55,   # loss / damage
    "kushindwa":    -0.55,   # to fail / be defeated
    "kuumia":       -0.55,   # to be hurt / injured
    "duni":         -0.50,   # inferior / substandard
    "uchafu":       -0.55,   # dirt / corruption
    "ugumu":        -0.50,   # hardness / difficulty
    "wasiwasi":     -0.55,   # worry / anxiety
    "tatizo":       -0.50,   # problem / issue
    "msumbufu":     -0.50,   # troublesome / annoying
    "lalamiko":     -0.50,   # complaint / grievance
    "kero":         -0.45,   # grievance / complaint (formal)
    "mkanganyiko":  -0.45,   # confusion / mix-up
    "tataniko":     -0.45,   # confusion / entanglement
    "sintofahamu":  -0.40,   # misunderstanding
    "onyo":         -0.40,   # warning / reprimand
    "uchovu":       -0.45,   # tiredness / fatigue
    "udhaifu":      -0.50,   # weakness / frailty
    "upungufu":     -0.45,   # shortage / deficiency
    "ukosefu":      -0.45,   # lack / absence
    "kikwazo":      -0.45,   # obstacle / barrier
    "gharama":      -0.35,   # high cost / expense (negative framing)
    "kupoteza":     -0.50,   # to lose
    "mzigo":        -0.40,   # burden / heavy load
    "chelewa":      -0.45,   # be late / delay (verb)
    "kuchelewa":    -0.45,   # to be late / delayed
    "uchelewaji":   -0.50,   # lateness / delay (noun)
    "dhambi":       -0.55,   # sin / wrongdoing
    # Agricultural negatives
    "ukame":        -0.55,   # drought / dry spell
    "wadudu":       -0.35,   # pests (agricultural)
    "uharibifu":    -0.55,   # destruction / damage
    # Sheng negative
    "stress":       -0.55,   # stressed out
    "noma":         -0.50,   # difficult / bad (Sheng)
    "ngori":        -0.50,   # harsh / difficult (Sheng)
    "fala":         -0.60,   # stupid / foolish (Sheng insult)
    "rada":         -0.30,   # problematic / aware of issue (Sheng)
    "chizi":        -0.25,   # crazy (negative context)

    # ══ INTENSIFIERS — mild positive tilt; flip sign on negated words ══
    "sana":          0.12,   # very / a lot (amplifier)
    "kabisa":        0.15,   # completely / absolutely (amplifier)
    "mno":           0.12,   # too much / excessively
    "zaidi":         0.10,   # more / further

    # ══ NEGATION MARKERS — slight negative tilt ════════════════════════
    "hapana":       -0.15,   # no / not at all
    "kamwe":        -0.20,   # never / not at all
    "bila":         -0.20,   # without / lacking
    "bado":         -0.05,   # still / yet (implies delay)
}


def load_swahili_lexicon(
    user_path: str = "swahili_lexicon.json",
) -> Dict[str, float]:
    """
    Load and merge the built-in Swahili sentiment lexicon with a user extension.

    HOW THE LEXICON IS USED — TWO COMPLEMENTARY PATHWAYS
    ─────────────────────────────────────────────────────
    Pathway 1 — VADER injection:
      All lexicon words are added to VADER's internal vocabulary so the
      VADER compound score benefits when it encounters them in normalized text.

    Pathway 2 — Direct lexicon scoring (_compute_lexicon_score):
      compute_sentiment() scores the *clean* (pre-token-map) text directly
      and blends the result with VADER using a language-weighted formula:
          Swahili  (lang="sw"):   50% VADER  +  50% Lexicon
          Code-mix (other lang):  65% VADER  +  35% Lexicon
          English  (lang="en"):   85% VADER  +  15% Lexicon

    This dual approach means:
      • Words IN the token map → translated → VADER scores them in English ✓
      • Words NOT in the token map → pass through untranslated → Lexicon
        scores them directly in Swahili ✓
      • Result: code-switched sentences are scored correctly in both halves ✓

    User extension (swahili_lexicon.json):
      Create this file to add sector-specific or regional terms.
      User entries always win on conflict with built-in entries.
      Format: { "harambee": 0.5, "msiba": -0.6, "shangwe": 0.7 }

    Returns
    ───────
    dict  { token: polarity_float }  (built-in + user merged)
    """
    merged = dict(SWAHILI_SENTIMENT_LEXICON)
    p = Path(user_path)
    if not p.exists():
        return merged
    try:
        with open(p, "r", encoding="utf-8") as f:
            user_lex = json.load(f)
        if isinstance(user_lex, dict):
            validated = {
                str(k).lower().strip(): float(v)
                for k, v in user_lex.items()
                if isinstance(v, (int, float)) and -1.0 <= float(v) <= 1.0
            }
            merged.update(validated)
            log.info(
                "Swahili lexicon: loaded %d user entries from %s "
                "(%d total entries after merge)",
                len(validated), user_path, len(merged),
            )
    except Exception as exc:
        log.warning("Could not load %s (%s): %s", user_path, type(exc).__name__, exc)
    return merged


# Cached English word set for unknown-token detection (lazy-loaded once)
_ENGLISH_WORD_SET: Optional[set] = None


def _get_english_word_set() -> set:
    """
    Lazily load the NLTK English words corpus.  Cached after first call.
    Falls back to a minimal set if the corpus is unavailable.
    """
    global _ENGLISH_WORD_SET
    if _ENGLISH_WORD_SET is not None:
        return _ENGLISH_WORD_SET
    try:
        from nltk.corpus import words as _nltk_words
        _ENGLISH_WORD_SET = {w.lower() for w in _nltk_words.words()}
        log.info("NLTK English word set loaded (%d words)", len(_ENGLISH_WORD_SET))
    except Exception:
        # Minimal safety-net covering the most common survey vocabulary
        _ENGLISH_WORD_SET = set(ENGLISH_STOP_WORDS) | {
            "late", "early", "good", "bad", "poor", "quality", "delivery",
            "training", "loan", "payment", "market", "officer", "group",
            "member", "seed", "fertilizer", "harvest", "farm", "support",
            "field", "problem", "issue", "cost", "price", "distance",
        }
    return _ENGLISH_WORD_SET

# ══════════════════════════════════════════════════════════════════════
# KOBO / ODK LOADER  (Section 3)
# ══════════════════════════════════════════════════════════════════════
# KoboToolbox and ODK both export data with a consistent set of system
# columns and structural conventions that break standard loaders.
# This function handles all of them in one place.
#
# WHAT IT FIXES
# ─────────────
# 1. Metadata columns  — _uuid, _submission_time, _id, _index,
#    _validation_status, formhub/uuid, meta/instanceID, _parent_index,
#    _xform_id_string, _geolocation, _tags, _notes, _status, _submitted_by
#    These are Kobo/ODK system fields and must not reach the NLP pipeline.
#
# 2. Group-prefixed column names  — KoboToolbox wraps questions inside
#    groups and exports them as "group_name/question_name" (XLS form) or
#    "group_name-question_name" (some CSV exports).  These are flattened
#    to just the question name so text_columns stays clean.
#    Example:  "household_info/q1_inputs"  →  "q1_inputs"
#
# 3. Repeat-group index columns  — "_index" and "_parent_index" appear
#    in repeat-group exports and are stripped.
#
# 4. Encoding  — Kobo CSV exports often arrive as Windows-1252 or
#    UTF-8-BOM from older form versions; chardet handles this.
#
# USAGE (standalone)
# ──────────────────
#   df = load_kobo_export("my_kobo_export.xlsx")
#   print(df.columns.tolist())   # shows clean question names
#
# USAGE (automatic)
# ─────────────────
#   Pass any Kobo/ODK file to FeedbackInsightEngine — _load_file()
#   detects the fingerprints and calls this automatically.

# System columns exported by KoboToolbox and ODK — always strip these
_KOBO_META_PREFIXES: tuple = (
    "_",              # _uuid, _id, _submission_time, _index, _geolocation …
    "formhub/",       # formhub/uuid
    "meta/",          # meta/instanceID, meta/deprecatedID
    "formhub-",       # alternative separator in some exports
    "meta-",
)

_KOBO_META_EXACT: set = {
    # Explicit names that don't start with _ but are still metadata
    "start", "end", "today", "deviceid", "simserial", "phonenumber",
    "username", "caseid", "xform_id_string", "version",
    "submitted_by", "validation_status", "status", "uuid",
}


def _is_kobo_meta(col: str) -> bool:
    """Return True if this column is a Kobo/ODK system/metadata column."""
    c = col.strip()
    for prefix in _KOBO_META_PREFIXES:
        if c.lower().startswith(prefix):
            return True
    return c.lower() in _KOBO_META_EXACT


def _flatten_kobo_column(col: str) -> str:
    """
    Flatten a group-prefixed Kobo/ODK column name to its final segment.

    Examples
    --------
    "household_info/q1_inputs"    →  "q1_inputs"
    "section_a-q2_training"       →  "q2_training"
    "repeat_group/q3_fieldofficer"→  "q3_fieldofficer"
    "plain_column"                →  "plain_column"   (unchanged)
    """
    # XLSForm uses "/" as group separator; some CSV exports use "-"
    for sep in ("/", "-"):
        if sep in col:
            return col.split(sep)[-1].strip()
    return col.strip()


def load_kobo_export(path: str) -> pd.DataFrame:
    """
    Load a KoboToolbox or ODK export and return a clean DataFrame.

    Handles
    -------
    • .xlsx and .csv exports (encoding auto-detected)
    • Metadata column stripping (_uuid, _submission_time, formhub/uuid, …)
    • Group-prefixed column flattening  (group/question → question)
    • Duplicate column names after flattening (suffixed _2, _3, …)
    • Empty rows (all-NaN) removed

    Parameters
    ----------
    path : str
        Path to the Kobo/ODK export file (.xlsx or .csv).

    Returns
    -------
    pd.DataFrame  with clean column names and no metadata columns.
    """
    p = Path(path)
    suffix = p.suffix.lower()

    # ── Load raw file ─────────────────────────────────────────────────
    if suffix in (".xlsx", ".xls"):
        df = pd.read_excel(p, dtype=str)
    else:
        # Auto-detect encoding (Kobo CSVs vary widely)
        raw = p.read_bytes()[:50_000]
        result = chardet.detect(raw)
        enc = result.get("encoding") or "utf-8"
        try:
            df = pd.read_csv(p, encoding=enc, dtype=str)
        except Exception:
            # encoding_errors='replace' keeps bad bytes instead of crashing
            df = pd.read_csv(p, encoding="utf-8", encoding_errors="replace",
                             dtype=str)

    original_cols = list(df.columns)
    log.info("Kobo loader: raw file has %d rows × %d columns", *df.shape)

    # ── Strip metadata columns ────────────────────────────────────────
    meta_cols  = [c for c in df.columns if _is_kobo_meta(c)]
    data_cols  = [c for c in df.columns if not _is_kobo_meta(c)]
    df = df[data_cols].copy()

    if meta_cols:
        log.info(
            "Kobo loader: stripped %d metadata column(s): %s",
            len(meta_cols),
            ", ".join(meta_cols[:8]) + (" …" if len(meta_cols) > 8 else ""),
        )

    # ── Flatten group-prefixed column names ───────────────────────────
    flat_names: List[str] = []
    seen: Dict[str, int] = {}
    for col in df.columns:
        flat = _flatten_kobo_column(col)
        if flat in seen:
            seen[flat] += 1
            flat = f"{flat}_{seen[flat]}"   # e.g. q1_inputs_2
        else:
            seen[flat] = 1
        flat_names.append(flat)

    renamed = {old: new for old, new in zip(df.columns, flat_names) if old != new}
    if renamed:
        df.rename(columns=renamed, inplace=True)
        log.info(
            "Kobo loader: flattened %d group-prefixed column(s): %s",
            len(renamed),
            ", ".join(f"{o} → {n}" for o, n in list(renamed.items())[:5])
            + (" …" if len(renamed) > 5 else ""),
        )

    # ── Drop fully-empty rows ─────────────────────────────────────────
    before = len(df)
    df = df.dropna(how="all").reset_index(drop=True)
    dropped = before - len(df)
    if dropped:
        log.info("Kobo loader: dropped %d empty row(s).", dropped)

    log.info(
        "Kobo loader: ready — %d rows × %d columns.  "
        "Set text_columns to any of: %s",
        len(df), len(df.columns),
        ", ".join(df.columns.tolist()),
    )
    return df


def _is_kobo_file(df: pd.DataFrame) -> bool:
    """
    Detect whether a loaded DataFrame looks like a Kobo/ODK export.

    Checks for the three most reliable fingerprints:
      1. A column starting with "_" (e.g. _uuid, _submission_time)
      2. A column containing "/" (group-path separator in XLSForm)
      3. The exact column name "formhub/uuid" or "meta/instanceID"
    """
    cols_lower = {c.lower() for c in df.columns}
    # Fingerprint 1: system columns with leading underscore
    if any(c.startswith("_") for c in df.columns):
        return True
    # Fingerprint 2: group-path separator
    if any("/" in c for c in df.columns):
        return True
    # Fingerprint 3: known Kobo metadata exact names
    if cols_lower & {"formhub/uuid", "meta/instanceid", "start", "end",
                     "deviceid", "simserial", "phonenumber"}:
        return True
    return False


# ══════════════════════════════════════════════════════════════════════
# RECOMMENDATION RULES  ─ Africa / Kenya context
# ══════════════════════════════════════════════════════════════════════

RECOMMENDATION_RULES: List[Tuple[List[str], str]] = [
    # ══════════════════════════════════════════════════════════════════
    # CRITICAL THEMES (Appear in all 5 questions)
    # ══════════════════════════════════════════════════════════════════

    # 1. INPUT DELIVERY — TIMING & LOGISTICS
    (["delivery", "delivered", "late", "delayed", "inputs delivered",
      "planting window", "rains", "early", "dispatch", "arrive",
      "collection point", "distribute", "distribution"],
     "Overhaul input supply timing: (1) Pre-position all input stocks at "
     "sub-location collection points BEFORE rains start (by November for "
     "December-January planting); (2) Use SMS alerts to notify farmers exactly "
     "when inputs arrive at collection point; (3) Coordinate timing with county "
     "extension for weather forecasting; (4) Set weekly collection schedule with "
     "transport buddies to reduce individual collection burden."),

    # 2. LOAN REPAYMENT — ALIGNMENT WITH HARVEST
    (["loan", "repayment", "repay", "harvest", "tight", "schedule",
      "boda", "payment", "mpesa", "default", "liability", "penalised",
      "penalty", "grace", "yield", "loss", "installment", "instalment"],
     "Redesign loan repayment structure: (1) Align repayment AFTER harvest "
     "(not during growing season) — grace period of 30-45 days post-harvest "
     "to allow selling crops; (2) Allow partial repayment (50%) after first "
     "harvest sale, balance due 2 weeks after final harvest; (3) Offer M-Pesa "
     "AND cash payment options (SMS system sometimes fails); (4) Remove group "
     "liability — instead: graduated individual penalties (mild for first late "
     "payment, strict only after 3+ months); (5) Provide harvest-tracking SMS: "
     "farmers report crop sale amounts, system auto-calculates due date; "
     "(6) If crop fails (drought/pests), offer 1-month extension (don't penalize)."),

    # 3. FIELD OFFICER — PRESENCE & SUPPORT
    (["officer", "field officer", "visits", "visit", "phone", "reach",
      "changed", "turnover", "replaced", "remote", "infrequent", "sub-location",
      "difficult", "shamba", "afisa", "spread", "coverage"],
     "Strengthen field officer coverage: (1) Reduce officer-to-farmer ratio "
     "(current is too high) to max 200 farmers per officer; (2) Set minimum visit "
     "frequency: 2× per month during growing season, 1× monthly off-season; "
     "(3) Provide officers motorbike/fuel allowance to reach remote areas; "
     "(4) Create WhatsApp group for farmers to ask questions between visits; "
     "(5) Implement structured check-in: officers send weekly SMS with seasonal tips; "
     "(6) Address turnover by raising salaries and offering relocation bonuses "
     "(reduce external transfers)."),

    # 4. GROUP DYNAMICS & MEMBER ACCOUNTABILITY
    (["group", "members", "member", "group members", "problems", "meetings",
      "leader", "communicate", "communication", "liability", "default",
      "conflict", "trust", "attendance", "rules", "chama"],
     "Rebuild group trust & governance: (1) Provide group leaders with 1-day "
     "facilitation training (communication, conflict resolution); (2) Create "
     "Clear Membership Rules document in Swahili: roles, meeting frequency (2× "
     "monthly minimum), contributions expected, consequences of default; "
     "(3) Move meeting attendance tracking from oral to digital (USSD: farmer "
     "texts 'HERE' to register, creates transparent record); (4) Implement "
     "Graduated Liability: new members (0%), year 2 (25%), year 3 (50%), "
     "founding members (100%) — prevents unfair punishment of new farmers; "
     "(5) Pay group leader small incentive fee (5% of repayment fee) for 100% "
     "on-time repayment; (6) Post meeting minutes on WhatsApp + physical notice board."),

    # 5. TRAINING — LANGUAGE & TIMING
    (["training", "training sessions", "sessions", "english", "language",
      "understand", "understood", "local", "swahili", "learned", "technique",
      "spacing", "practical", "suppose", "workshop", "demo"],
     "Redesign training delivery: (1) Conduct ALL training in Swahili FIRST "
     "(not English) — translate materials to local language; (2) Schedule "
     "training OFF-harvest (Feb-Apr, Sept-Oct) when farmers have free time; "
     "(3) Move from classroom to DEMONSTRATION PLOTS (hands-on field learning); "
     "(4) Record short 5-minute videos in Swahili showing: seed spacing, "
     "fertilizer application, pest management — farmers replay offline; "
     "(5) Offer women-only sessions (separate from mixed groups) for household schedules; "
     "(6) Use Swahili proverbs (e.g., 'Haraka haraka haina baraka') to reinforce patience."),

    # ══════════════════════════════════════════════════════════════════
    # HIGH PRIORITY THEMES (Appear in 3-4 questions)
    # ══════════════════════════════════════════════════════════════════

    # 6. INPUT QUALITY — SEEDS & FERTILIZER
    (["seeds", "seed", "fertilizer", "fertiliser", "mbolea", "mbegu",
      "rotten", "germinate", "germination", "variety", "quality", "poor",
      "performed", "perform", "spacing", "agronomic"],
     "Implement pre-distribution quality assurance: (1) Test all seed batches "
     "for germination rate (must exceed 85%) before release to farmers; (2) Have "
     "supplier contracts specify quality penalties for rotten seed/fertilizer; "
     "(3) Train farmer Quality Committees to spot-check inputs at collection; "
     "(4) Establish rapid replacement protocol (48 hours) for defective inputs; "
     "(5) Record batch numbers so failed inputs can be traced & credited; "
     "(6) For underperforming varieties, collect farmer feedback and adjust seed selection."),

    # 7. MARKET LINKAGE — PRICE & ACCESS
    (["market", "market linkage", "boda boda", "transport", "cost", "price",
      "buyer", "sell", "collection", "sokoni", "mavuno", "post-harvest",
      "storage", "nearest", "far", "profit", "income"],
     "Establish farmer-friendly market linkages: (1) Develop forward contracts "
     "with 2-3 certified buyers (maize traders, aggregators) — set FIXED PRICES "
     "before harvest; (2) Organize bulked sales (groups of 5-10 farmers) to improve "
     "bargaining power vs middlemen; (3) Negotiate flat boda-boda delivery rates "
     "(e.g., 500 KES per bag) rather than per-km (saves 30-40%); (4) Create sub-location "
     "collection hubs so farmers don't travel to county market; (5) Send SMS price alerts: "
     "'Maize trading at 3200/bag at Eldoret market today'; (6) If farmer chooses own buyer, "
     "OAF provides subsidized transport support."),

    # 8. COLLECTION POINT — ACCESSIBILITY
    (["collection point", "collection", "point", "far", "distance", "mbali",
      "nearest", "village", "location", "access", "transport", "delivery",
      "reach"],
     "Decentralize input distribution: (1) Establish collection points at "
     "sub-location level (not just ward/county level) — target radius max 3 km "
     "from any farmer; (2) Partner with existing community centers (schools, "
     "health centers, churches) as collection depots; (3) Set collection windows "
     "3-4 days per week so farmers can choose convenient times; (4) Shuttle "
     "service: farmers pool transport costs to reduce per-person burden; "
     "(5) For remotest farmers, offer mobile collection (OAF vehicle visits bi-weekly)."),

    # 9. M-PESA & DIGITAL RESILIENCE
    (["mpesa", "mpesa repayment", "system", "system failed", "down",
      "payment", "ussd", "code", "airtel", "safaricom", "network", "online"],
     "Implement multi-channel payment resilience: (1) Enable M-Pesa for "
     "repayment BUT maintain CASH fallback option at collection points (network "
     "fails 10-15% of the time); (2) Test M-Pesa on all providers (Airtel too, "
     "not just Safaricom); (3) Provide USSD code option for balance checks; "
     "(4) Train officers to accept physical payment books + cash receipts for "
     "network downtime; (5) Partner with mobile money agents at collection points "
     "for instant cash-in/out; (6) Send payment confirmation SMS to farmer "
     "(reduces disputes over timing)."),

    # ══════════════════════════════════════════════════════════════════
    # CULTURAL & MULTILINGUAL CONTEXTS
    # ══════════════════════════════════════════════════════════════════

    # 10. SWAHILI-LANGUAGE CONTEXTS (Mbolea, Mafunzo, Mavuno)
    (["ilikuwa", "mbolea", "mbegu", "mafunzo", "mavuno", "shama",
      "afisa", "sokoni", "usafiri", "nzuri", "sawa", "poa"],
     "Enhance cultural + linguistic integration: (1) Conduct ALL key messaging "
     "(loan terms, quality guarantees, market prices) in Swahili FIRST, not English; "
     "(2) Record training videos in Swahili with locally-relevant examples (neighbor's "
     "harvest success, local pests); (3) Conduct focus groups in Swahili (farmers express "
     "nuanced concerns better in mother language); (4) Train field officers in Swahili "
     "facilitation skills; (5) Use Swahili proverbs in training to reinforce key messages; "
     "(6) Celebrate harvest successes in Swahili (e.g., 'Mavuno yako ilikuwa moto!' = "
     "'Your harvest was excellent!')."),

    # 11. ACRE FUND PROGRAM — SUPPORT & SATISFACTION
    (["acre fund", "acre", "fund", "support", "program", "helped", "satisfied",
      "satisfied", "experience", "expectations", "deliver", "improvement",
      "works", "success"],
     "Amplify positive impacts while addressing gaps: (1) For satisfied farmers, "
     "record success stories (50-word testimonials + photos) for training of new groups; "
     "(2) Identify what's working: ask each happy farmer 'What was most helpful?' and "
     "double down; (3) For less-satisfied: run mini-focus groups (5-6 farmers) to understand "
     "specific pain points; (4) Create feedback loop: quarterly SMS survey asking 'Rate OAF "
     "support 1-10'; (5) Invite satisfied farmers to mentor new groups (incentivize with "
     "500 KES per group mentored); (6) Address systemic gaps (seed quality, timeliness); "
     "(7) Measure impact: track yield increases year-on-year by village/group."),

    # ══════════════════════════════════════════════════════════════════
    # FALLBACK (Generic)
    # ══════════════════════════════════════════════════════════════════

    # 12. GENERAL / OTHER FEEDBACK
    (["other", "feedback", "general", "misc"],
     "For mixed or unclear feedback: (1) Conduct targeted informal conversations "
     "(1-on-1 visits) with farmers who expressed concern; (2) Ask: 'What would improve this?' "
     "and listen carefully; (3) If pattern emerges (multiple farmers mention same issue), "
     "escalate to management with farmer quotes; (4) Co-design solution WITH farmers "
     "(not for them); (5) Test solution with 2-3 farmer groups before full rollout; "
     "(6) Track impact and adjust based on farmer feedback."),
]

FALLBACK_RECOMMENDATION = (
    "Conduct targeted focus-group discussions to identify the root "
    "cause before designing an intervention."
)

# ── TIP: HOW TO EXTEND RECOMMENDATION_RULES FOR YOUR SECTOR ──────────
# After running the pipeline, check the console for:
#   "KEYWORD TIPS" — these are the actual terms found in your data
#   that did NOT match any existing rule.
# Copy those terms into a new rule tuple above, then write a
# sector-specific recommendation. Example for a WASH project:
#
#   (["borehole", "pump", "handwashing", "latrine", "open defecation"],
#    "Prioritise borehole rehabilitation and community-led sanitation "
#    "campaigns; train local pump mechanics for sustained maintenance."),
#


# ══════════════════════════════════════════════════════════════════════
# MAIN ENGINE
# ══════════════════════════════════════════════════════════════════════

class FeedbackInsightEngine:
    """
    End-to-end survey text analysis pipeline optimised for
    Kenyan / African multilingual data on 8 GB RAM.
    """

    def __init__(
        self,
        file_path: str,
        text_columns: List[str],
        config: Optional[EngineConfig] = None,
    ):
        self.file_path = Path(file_path)
        self.text_columns = text_columns
        self.config = config or EngineConfig()

        # ── load data ────────────────────────────────────────────────
        self.df = self._load_file(self.file_path)
        log.info("Loaded %d rows × %d columns", *self.df.shape)

        # ── shared NLP objects (loaded once) ─────────────────────────
        log.info(
            "Loading sentence-transformer model: %s …",
            self.config.embedding_model,
        )
        self.model = SentenceTransformer(self.config.embedding_model)

        self.sia = SentimentIntensityAnalyzer()
        # ── Section 6: Load Swahili lexicon + inject into VADER ───────
        # Load BEFORE extra_sentiment_lexicon so user overrides win.
        self._swahili_lexicon: Dict[str, float] = load_swahili_lexicon(
            self.config.swahili_lexicon_path
        )
        self.sia.lexicon.update(self._swahili_lexicon)
        # extra_sentiment_lexicon (EngineConfig) applied last — wins on conflict.
        self.sia.lexicon.update(self.config.extra_sentiment_lexicon)
        log.info(
            "Sentiment lexicons loaded: %d Swahili/Sheng entries injected into VADER.",
            len(self._swahili_lexicon),
        )

        self.vectorizer = TfidfVectorizer(
            stop_words=list(COMBINED_STOP_WORDS),
            ngram_range=self.config.tfidf_ngram_range,
            min_df=2,
            max_df=0.9,
            max_features=self.config.tfidf_max_features,
        )

        # ── per-question state (reset each loop iteration) ───────────
        self.q_df: pd.DataFrame = pd.DataFrame()
        self.embeddings: np.ndarray = np.array([])
        self.tfidf_matrix = None
        self.feature_names: np.ndarray = np.array([])
        self.kmeans = None
        self.cluster_summary: Dict = {}
        self.summary_df: pd.DataFrame = pd.DataFrame()
        self.key_insights: List[str] = []
        self.dataset_summary: Dict = {}
        self.emerging_issues: pd.DataFrame = pd.DataFrame()
        self.cluster_to_theme_map: Dict = {}  # Cluster ID → Theme name mapping
        # Section 9: set by cluster_feedback(), consumed by generate_dataset_summary()
        self._last_silhouette_score: Optional[float] = None

        # ── cross-question artefacts ──────────────────────────────────
        self.question_results: Dict = {}
        self.cross_question_themes: List[str] = []

        # ── LLM client (optional) ─────────────────────────────────────
        # Cache maps frozenset-of-top-3-keywords → theme name so the same
        # cluster never makes a second API call within a run.
        self._theme_name_cache: Dict[str, str] = {}
        self._anthropic_client = None
        self._gemini_client = None   # Section 10

        want_llm = (
            self.config.use_llm_naming or self.config.use_llm_recommendations
        )

        if want_llm and self.config.use_gemini:
            # ── Section 10: Gemini path ───────────────────────────────
            if GEMINI_AVAILABLE:
                try:
                    # Configure is a no-op if genai.configure() was already
                    # called by the notebook with an API key.  If the env var
                    # GEMINI_API_KEY is set, the library picks it up automatically.
                    import os as _os
                    _api_key = _os.environ.get("GEMINI_API_KEY", "")
                    if _api_key:
                        _gemini_lib.configure(api_key=_api_key)
                    self._gemini_client = _gemini_lib.GenerativeModel(
                        self.config.gemini_model
                    )
                    log.info(
                        "Gemini client initialised (model: %s) — LLM features enabled.",
                        self.config.gemini_model,
                    )
                except Exception as exc:
                    log.warning(
                        "Could not init Gemini client (%s). "
                        "Falling back to rule-based methods.", exc
                    )
            else:
                log.warning(
                    "google-generativeai package not installed — Gemini disabled. "
                    "Run:  pip install google-generativeai"
                )
        elif want_llm and not self.config.use_gemini:
            # ── Existing Anthropic / Claude path ─────────────────────
            if ANTHROPIC_AVAILABLE:
                try:
                    self._anthropic_client = _anthropic_lib.Anthropic()
                    log.info("Anthropic client initialised — LLM features enabled.")
                except Exception as exc:
                    log.warning("Could not init Anthropic client (%s). "
                                "Falling back to rule-based methods.", exc)
            else:
                log.warning(
                    "anthropic package not installed — LLM features disabled. "
                    "Run: pip install anthropic"
                )

        # ── NEW: Load config and initialize engines (v3+ Upgrade) ────
        self.pipeline_config = load_config()
        
        # Initialize advanced engines if enabled
        self._init_advanced_engines()

        # ── Section 1: Token map (built-in + user-defined) ────────────
        self._token_map: Dict[str, str] = _build_token_map(
            self.config.token_map_path
        )
        log.info(
            "Token map ready: %d entries (built-in + user).  "
            "Unknown tokens → %s after each run.",
            len(self._token_map),
            self.config.token_review_path,
        )
        # Accumulated unknown-token data across questions (cleared per run)
        self._token_review_data: Dict[str, dict] = {}

        # ── Section 8: Theme Registry (persistent name stability) ─────
        # Loaded once at startup; new names written back at end of run().
        self._theme_registry: Dict[str, str] = self._load_theme_registry()

    # ══════════════════════════════════════════════════════════════════
    # THEME REGISTRY  (Section 8)
    # ══════════════════════════════════════════════════════════════════

    def _load_theme_registry(self) -> Dict[str, str]:
        """
        Load the persistent theme name registry from disk.

        The registry maps a sorted-keyword fingerprint (the same key used by
        _theme_name_cache) to a stable human-readable theme name.  Loading it
        once at startup means theme names generated in previous runs are reused
        automatically without any LLM call or rule-based computation.

        Returns an empty dict if the file does not exist yet — first run is
        normal and the registry is built up over time.

        To manually rename a theme: open theme_registry.json, edit the value
        for the relevant key, and re-run. The edited name will be used for all
        future runs until the underlying keywords change enough to produce a
        new fingerprint.
        """
        p = Path(self.config.theme_registry_path)
        if not p.exists():
            log.info(
                "Theme registry not found at '%s' — will create on first run.",
                p,
            )
            return {}
        try:
            with open(p, "r", encoding="utf-8") as f:
                registry = json.load(f)
            if isinstance(registry, dict):
                log.info(
                    "Theme registry loaded: %d stable names from '%s'.",
                    len(registry), p,
                )
                return {str(k): str(v) for k, v in registry.items()}
        except Exception as exc:
            log.warning("Could not load theme registry (%s): %s", p, exc)
        return {}

    def _save_theme_registry(self) -> None:
        """
        Persist the current in-memory theme name cache to theme_registry.json.

        Called at the end of every run() so that theme names produced during
        this run are available on the next run without repeating LLM calls.

        The registry is the UNION of the loaded registry and all names generated
        in the current run — it only ever grows, never shrinks (unless manually
        edited).
        """
        # Merge: registry entries win over cache on conflict (manual edits survive)
        merged = dict(self._theme_name_cache)   # new names from this run
        merged.update(self._theme_registry)      # registry overrides (manual edits)
        self._theme_registry = merged

        p = Path(self.config.theme_registry_path)
        try:
            with open(p, "w", encoding="utf-8") as f:
                json.dump(
                    dict(sorted(merged.items())),  # sorted keys for readability
                    f, indent=2, ensure_ascii=False,
                )
            log.info(
                "Theme registry saved: %d entries → '%s'.",
                len(merged), p,
            )
        except Exception as exc:
            log.warning("Could not save theme registry (%s): %s", p, exc)

    def _init_advanced_engines(self) -> None:
        """Initialize trigger/action engines from config."""
        try:
            # Try absolute imports first (when run as main script)
            try:
                from trigger_engine import TriggerEngine
                from action_engine import ActionEngine
            except ImportError:
                # Fall back to relative imports (when imported as module)
                from .trigger_engine import TriggerEngine
                from .action_engine import ActionEngine
            
            self.trigger_engine = TriggerEngine(
                self.pipeline_config.get("thresholds", {})
            )
            self.action_engine = ActionEngine(
                sector=self.pipeline_config.get("sector", "")
            )
            
            log.info("✓ Advanced engines initialized (triggers, actions)")
        except Exception as exc:
            log.warning("Could not initialize advanced engines (%s). "
                       "Ensure trigger_engine.py, action_engine.py "
                       "exist in project directory.", exc)
            self.trigger_engine = None
            self.action_engine = None

    def _map_sentiment_label(self, score: float) -> str:
        """
        Convert sentiment score to label for trigger logic.
        """
        if score <= -0.05:
            return "Problem"
        elif score >= 0.05:
            return "Positive"
        else:
            return "Neutral"


    def _compute_dynamic_thresholds(self, themes: list) -> dict:
        """
        Compute percentile-based thresholds from theme impacts.
        """
        impacts = [t.get("impact", 0) for t in themes if t.get("impact") is not None]

        if not impacts:
            return {"high_impact": 1, "medium_impact": 1}

        impacts_sorted = sorted(impacts)

        n = len(impacts_sorted)

        # Small dataset fallback
        if n < 5:
            high = max(impacts_sorted)
            medium = impacts_sorted[n // 2]
            return {"high_impact": high, "medium_impact": medium}

        def percentile(data, p):
            k = (len(data) - 1) * (p / 100)
            f = int(k)
            c = min(f + 1, len(data) - 1)
            if f == c:
                return data[int(k)]
            return data[f] + (data[c] - data[f]) * (k - f)

        high = percentile(impacts_sorted, 80)
        medium = percentile(impacts_sorted, 50)

        return {
            "high_impact": max(1, round(high)),
            "medium_impact": max(1, round(medium))
        }

    # ══════════════════════════════════════════════════════════════════
    # FILE LOADING
    # ══════════════════════════════════════════════════════════════════

    @staticmethod
    def _detect_encoding(path: Path) -> str:
        """Use chardet to detect CSV encoding (handles Windows-1252, UTF-8-BOM …)."""
        raw = path.read_bytes()[:50_000]
        result = chardet.detect(raw)
        enc = result.get("encoding") or "utf-8"
        log.info("Detected file encoding: %s (confidence %.0f%%)",
                 enc, (result.get("confidence") or 0) * 100)
        return enc

    def _load_file(self, path: Path) -> pd.DataFrame:
        """
        Load survey data from .xlsx, .xls, or .csv.

        Auto-detects KoboToolbox / ODK exports by checking for their
        structural fingerprints (_uuid column, group/question paths, etc.)
        and routes them through load_kobo_export() automatically.

        Standard files (no Kobo fingerprints) are loaded normally.
        """
        suffix = path.suffix.lower()

        # ── First pass: load raw to check for Kobo fingerprints ───────
        if suffix in (".xlsx", ".xls"):
            df_raw = pd.read_excel(path, dtype=str, nrows=2)
        else:
            enc = self._detect_encoding(path)
            try:
                df_raw = pd.read_csv(path, encoding=enc, dtype=str, nrows=2)
            except Exception:
                df_raw = pd.read_csv(
                    path, encoding="utf-8", errors="replace", dtype=str, nrows=2
                )

        # ── Route to Kobo loader if fingerprints found ─────────────────
        if _is_kobo_file(df_raw):
            log.info(
                "Kobo/ODK export detected in '%s' — using Kobo loader.", path.name
            )
            return load_kobo_export(str(path))

        # ── Standard load (no Kobo fingerprints) ──────────────────────
        if suffix in (".xlsx", ".xls"):
            return pd.read_excel(path)
        else:
            enc = self._detect_encoding(path)
            try:
                return pd.read_csv(path, encoding=enc)
            except Exception:
                return pd.read_csv(path, encoding="utf-8",
                                   encoding_errors="replace")

    # ══════════════════════════════════════════════════════════════════
    # LANGUAGE DETECTION
    # ══════════════════════════════════════════════════════════════════

    @staticmethod
    def detect_language(text: str) -> str:
        """Return ISO 639-1 code or 'en' if detection fails."""
        if not LANGDETECT_AVAILABLE or len(text.split()) < 4:
            return "en"
        try:
            return lang_detect(text)
        except Exception:
            return "en"

    # ══════════════════════════════════════════════════════════════════
    # TEXT CLEANING  ─ language-aware
    # ══════════════════════════════════════════════════════════════════

    def clean_text(self, text: str) -> str:
        text = str(text).strip()

        # Detect language BEFORE lowercasing / stripping
        lang = self.detect_language(text)
        self.q_df.at[getattr(self, "_current_idx", 0), "lang"] = lang

        text = text.lower()
        
        # ── NEW: Sector-aware text normalization (v3+ upgrade) ────
        text = self.normalize_text(text)

        if lang in ("sw", "en"):
            # Keep only ASCII letters + spaces for now
            # (Swahili uses standard Latin alphabet so this is safe)
            text = re.sub(r"[^a-z\s]", " ", text)
        else:
            # For unidentified / mixed scripts keep letters broadly
            text = re.sub(r"[^\w\s]", " ", text, flags=re.UNICODE)

        # Normalise whitespace
        text = re.sub(r"\s+", " ", text).strip()
        return text

    def normalize_text(self, text: str) -> str:
        """
        Normalize text in two passes:
          Pass 1 — Phrase-level: fix multi-word expressions (m-pesa → mpesa).
          Pass 2 — Token-level: map Swahili/Sheng tokens → English equivalents
                   using the merged SWAHILI_TOKEN_MAP + user_token_map.json.

        ⚠️  SECTOR-AWARE:
        Phrase-level rules below are specific to smallholder agriculture.
        Add your sector's rules in the same style.

        HEALTHCARE examples:
            text = text.replace("hiv aids", "hiv")
            text = text.replace("antenatal care", "antenatal")

        WASH examples:
            text = text.replace("pit latrine", "latrine")
            text = text.replace("bore hole", "borehole")
        """
        # ── Pass 1: phrase-level fixes (order matters — longer first) ─
        text = text.replace("extension officer", "officer")
        text = text.replace("field officer",     "officer")
        text = text.replace("boda boda",         "boda")
        text = text.replace("bodaboda",          "boda")
        text = text.replace("m-pesa",            "mpesa")
        text = text.replace("m pesa",            "mpesa")
        text = text.replace("matatu",            "transport")
        text = text.replace("sawa sawa",         "okay")   # Sheng phrase

        # ── Pass 2: token-level map ────────────────────────────────────
        text = self.apply_token_map(text)
        return text

    def apply_token_map(self, text: str) -> str:
        """
        Apply the loaded token map to every whitespace-separated token.

        Tokens not found in the map are returned unchanged.
        This is intentionally simple and fast (O(n) dict lookups).

        Called by normalize_text() so it runs on every response before
        TF-IDF, embedding, and VADER sentiment scoring.
        """
        if not self._token_map:
            return text
        return " ".join(self._token_map.get(tok, tok) for tok in text.split())

    # ══════════════════════════════════════════════════════════════════
    # TOKEN REVIEW  ─ unknown-token scanner + JSON report
    # ══════════════════════════════════════════════════════════════════

    def _save_token_review(self, question_col: str) -> None:
        """
        Scan compressed_text tokens that are (a) not English, (b) not in
        stop-words, and (c) not already in the token map.  Write results to
        token_review.json so the user knows what to add to user_token_map.json.

        WORKFLOW
        ────────
        1st run  → token_review.json created with unknown tokens + frequencies
        User     → reviews token_review.json, adds mappings to user_token_map.json
        2nd run  → better clustering, better sentiment, better keywords

        INTERACTIVE MODE (pause_for_token_review=True)
        ──────────────────────────────────────────────
        Pipeline pauses after this call, prints top tokens, and waits for the
        user to update user_token_map.json before clustering begins.
        """
        english_words = _get_english_word_set()

        # Build the "already known" set so we only surface genuinely new tokens
        known: set = (
            COMBINED_STOP_WORDS
            | set(self._token_map.keys())
            | set(self._token_map.values())
            | english_words
        )

        # ── Collect token frequencies + example sentences ─────────────
        token_counts: Dict[str, int] = {}
        token_examples: Dict[str, List[str]] = {}

        for _, row in self.q_df.iterrows():
            compressed = str(row.get("compressed_text", ""))
            clean_src  = str(row.get("clean_text", ""))[:120]
            for tok in compressed.split():
                tok = tok.lower().strip()
                if len(tok) < 3 or tok in known:
                    continue
                token_counts[tok] = token_counts.get(tok, 0) + 1
                if tok not in token_examples:
                    token_examples[tok] = []
                if len(token_examples[tok]) < 2:
                    token_examples[tok].append(clean_src)

        # ── Merge into cumulative store (across questions in this run) ─
        for tok, count in token_counts.items():
            entry = self._token_review_data.get(tok, {
                "count": 0, "questions": [], "examples": [],
            })
            entry["count"] += count
            if question_col not in entry["questions"]:
                entry["questions"].append(question_col)
            for ex in token_examples.get(tok, []):
                if ex not in entry["examples"] and len(entry["examples"]) < 3:
                    entry["examples"].append(ex)
            self._token_review_data[tok] = entry

        if not token_counts:
            log.info("  ✅ No unknown tokens found in %s", question_col)
            return

        log.info(
            "  📝 %d unknown tokens in '%s' — see %s",
            len(token_counts), question_col, self.config.token_review_path,
        )

        # ── Write JSON report ──────────────────────────────────────────
        sorted_data = dict(
            sorted(self._token_review_data.items(),
                   key=lambda x: -x[1].get("count", 0))
        )
        output = {
            "generated": datetime.now(EAT).strftime("%Y-%m-%d %H:%M EAT"),
            "instructions": (
                "These tokens were not recognized as English and are not in "
                "the built-in Swahili/Sheng map.  For each token you want to "
                "translate, add it to user_token_map.json and re-run.\n"
                "Format:  { \"token\": \"english_equivalent\", ... }"
            ),
            "user_token_map_path": self.config.token_map_path,
            "total_unknown_tokens": len(sorted_data),
            "tokens": sorted_data,
        }
        try:
            review_path = Path(self.config.token_review_path)
            with open(review_path, "w", encoding="utf-8") as f:
                json.dump(output, f, indent=2, ensure_ascii=False)
        except Exception as exc:
            log.warning("Could not write token_review.json: %s", exc)

        # ── Optional interactive pause ─────────────────────────────────
        if self.config.pause_for_token_review and token_counts:
            top = sorted(token_counts.items(), key=lambda x: -x[1])[:8]
            SEP = "─" * 62
            print(f"\n{'═' * 62}")
            print(f"  UNKNOWN TOKENS  —  {question_col}")
            print(f"{'═' * 62}")
            print(f"  {'Token':<20}  Count  Example")
            print(SEP)
            for tok, cnt in top:
                ex = (token_examples.get(tok) or [""])[0][:50]
                print(f"  {tok:<20}  {cnt:>5}  {ex}")
            print(f"\n  Full list  →  {self.config.token_review_path}")
            print(f"  Add mappings →  {self.config.token_map_path}")
            print(SEP)
            input("  Press Enter to continue with current mappings → ")
            # Reload in case the user updated the file while paused
            self._token_map = _build_token_map(self.config.token_map_path)
            log.info("Token map reloaded: %d entries total", len(self._token_map))

    # ══════════════════════════════════════════════════════════════════
    # SARCASM / HEDGED-POSITIVE FLAG
    # ══════════════════════════════════════════════════════════════════

    HEDGED_PATTERNS = re.compile(
        r"\b(not bad|could be better|not the best|sort of ok|i guess|"
        r"not great|nothing special|so so|meh|average|below average|"
        r"acceptable i suppose|okay i guess)\b",
        re.IGNORECASE,
    )

    # Section 6: Swahili / Sheng hedged language patterns
    # These detect indirect, non-committal, or mildly sarcastic phrasing
    # common in Kenyan rural survey responses.
    SWAHILI_HEDGED_PATTERNS = re.compile(
        r"\b(nadhani|labda|pengine|kidogo tu|si vibaya sana|si mbaya sana|"
        r"inaweza kuwa bora|haijalishi sana|iko sawa tu|nadhani sawa|"
        r"si baya kabisa|inafaa kidogo|sawa kidogo|lakini bado|"
        r"kwa kiasi fulani|si mbaya)\b",
        re.IGNORECASE,
    )

    def detect_hedged_sentiment(self, text: str) -> float:
        """
        Return a negative adjustment if hedged/sarcastic language is found.

        Section 6 upgrade: now checks BOTH English and Swahili hedged patterns.

        Rationale:
          Hedged phrases like "nadhani sawa" (I guess it's okay) and
          "si vibaya sana" (not too bad) carry implied dissatisfaction —
          they are polite deflections, not genuine positives.
          Each matched pattern applies a -0.2 penalty to the sentiment score.

        Returns:
            float: negative adjustment (0.0 if no hedged language found)
        """
        en_matches = self.HEDGED_PATTERNS.findall(text)
        sw_matches = self.SWAHILI_HEDGED_PATTERNS.findall(text)
        return -0.2 * (len(en_matches) + len(sw_matches))

    # ══════════════════════════════════════════════════════════════════
    # SENTENCE SPLITTING
    # ══════════════════════════════════════════════════════════════════

    def split_sentences(self, text: str) -> List[str]:
        parts = re.split(r"[.!?;]", text)
        return [
            p.strip()
            for p in parts
            if len(p.split()) >= self.config.min_sentence_words
        ]

    # ══════════════════════════════════════════════════════════════════
    # COMPRESS TEXT  ─ stop-word filtered, for clustering
    # ══════════════════════════════════════════════════════════════════

    def compress_text(self, text: str) -> str:
        tokens = re.findall(r"\b[a-zA-Z]+\b", text.lower())
        filtered = [
            t for t in tokens
            if len(t) > 2 and t not in COMBINED_STOP_WORDS
        ]
        return " ".join(filtered)

    # ══════════════════════════════════════════════════════════════════
    # NEAR-DUPLICATE DEDUPLICATION
    # ══════════════════════════════════════════════════════════════════

    def deduplicate_responses(self) -> pd.DataFrame:
        """
        Collapse near-duplicate compressed texts.
        Keeps the first occurrence; annotates duplicates with original_count.
        This prevents one respondent repeating the same complaint
        from inflating a cluster (a known weakness of survey tools).
        """
        seen: Dict[str, int] = {}
        keep_indices = []

        for idx, row in self.q_df.iterrows():
            key = row["compressed_text"][:80]  # fingerprint first 80 chars
            if key not in seen:
                seen[key] = 1
                keep_indices.append(idx)
            else:
                seen[key] += 1

        before = len(self.q_df)
        # reset_index is CRITICAL: TF-IDF matrix rows are 0..n-1 but
        # .loc[keep_indices] preserves original (gapped) index values,
        # causing IndexError when extract_keywords() slices the matrix.
        self.q_df = self.q_df.loc[keep_indices].reset_index(drop=True)
        after = len(self.q_df)

        if before - after > 0:
            log.info("Deduplication removed %d near-duplicate sentences "
                     "(%d → %d)", before - after, before, after)

        return self.q_df

    # ══════════════════════════════════════════════════════════════════
    # PREPROCESSING
    # ══════════════════════════════════════════════════════════════════

    def preprocess(self) -> None:
        rows = []
        langs = []
        orig_idxs = []  # Section 5: track source respondent for each sentence

        for _, resp_row in self.q_df.iterrows():
            text     = resp_row["response"]
            orig_idx = resp_row.get("orig_idx", -1)
            clean    = self.clean_text(str(text))
            lang     = self.detect_language(str(text))
            for s in self.split_sentences(clean):
                rows.append(s)
                langs.append(lang)
                orig_idxs.append(orig_idx)

        self.q_df = pd.DataFrame({
            "clean_text": rows,
            "lang":       langs,
            "orig_idx":   orig_idxs,   # Section 5: respondent link preserved
        })

        self.q_df["compressed_text"] = self.q_df["clean_text"].apply(
            self.compress_text
        )

        # Drop empty compressed texts
        self.q_df = self.q_df[
            self.q_df["compressed_text"].str.strip() != ""
        ].copy().reset_index(drop=True)

        log.info("After preprocessing: %d sentence units", len(self.q_df))

        # ── Section 1: scan for unmapped non-English tokens ───────────
        # Writes token_review.json and (optionally) pauses for user input.
        self._save_token_review(
            getattr(self, "_current_question_col", "unknown")
        )

    # ══════════════════════════════════════════════════════════════════
    # TF-IDF
    # ══════════════════════════════════════════════════════════════════

    def build_tfidf_matrix(self) -> None:
        texts = self.q_df["compressed_text"].tolist()
        self.tfidf_matrix = self.vectorizer.fit_transform(texts)
        self.feature_names = self.vectorizer.get_feature_names_out()

    # ══════════════════════════════════════════════════════════════════
    # SENTIMENT  ─ Blended VADER + Swahili Lexicon (Section 6)
    # ══════════════════════════════════════════════════════════════════

    def _compute_lexicon_score(self, text: str) -> float:
        """
        Score text directly using the Swahili/Sheng sentiment lexicon.

        This is PATHWAY 2 — the complement to VADER's Pathway 1.

        HOW IT WORKS
        ────────────
        • Operates on clean_text (after lowercasing, before token mapping).
        • For each token in the text:
            - If a negation marker preceded it (si, bila, hapana, kamwe, la):
              score is flipped and dampened to -0.7× (partial negation,
              not full reversal — natural language is rarely a perfect flip).
            - If an intensifier preceded it (sana, kabisa, mno):
              score is amplified ×1.3.
        • Mean of all recognized token scores, clamped to [-1, 1].
        • Returns 0.0 if NO lexicon tokens are found (no adjustment).

        NEGATION WINDOW
        ────────────────
        The negation flag resets after any non-stop-word token, not just
        after a recognized lexicon word.  This prevents long-range negation
        ("si ... three words later ... nzuri") from incorrectly flipping scores.

        Args:
            text: clean_text string (lowercased, regex-cleaned)

        Returns:
            float in [-1.0, 1.0]; 0.0 if no lexicon words recognized
        """
        if not text or not self._swahili_lexicon:
            return 0.0

        tokens = text.lower().split()
        if not tokens:
            return 0.0

        NEGATORS    = {"si", "bila", "hapana", "kamwe", "la", "not", "no", "never"}
        INTENSIFIERS = {"sana", "kabisa", "mno", "zaidi", "very", "so", "sana"}
        # Short connectors that should NOT reset the negation window
        CONNECTORS  = {"na", "ya", "wa", "kwa", "ni", "au", "pia"}

        scores: List[float] = []
        negate    = False
        intensify = False

        for tok in tokens:
            clean_tok = tok.strip(".,!?;:'\"")

            if clean_tok in NEGATORS:
                negate    = True
                intensify = False
                continue

            if clean_tok in INTENSIFIERS:
                intensify = True
                continue

            score = self._swahili_lexicon.get(clean_tok)
            if score is not None:
                if intensify:
                    score = score * 1.3          # amplify
                if negate:
                    score = score * -0.7         # partial negation
                scores.append(max(-1.0, min(1.0, score)))
                negate    = False
                intensify = False
            else:
                # Unknown token: reset negation window unless it's a connector
                if clean_tok not in CONNECTORS:
                    negate    = False
                    intensify = False

        if not scores:
            return 0.0

        return max(-1.0, min(1.0, sum(scores) / len(scores)))

    def compute_sentiment(self) -> None:
        """
        Compute sentiment scores using a THREE-PATHWAY blended approach.

        PATHWAY 1 — VADER (on normalized / token-mapped text):
          • Best for English sentences and Swahili words translated by the
            token map (e.g. "mbaya" → "bad" → VADER scores "bad").
          • Enhanced by: extra_sentiment_lexicon + SWAHILI_SENTIMENT_LEXICON
            both injected into VADER\'s vocabulary in __init__.

        PATHWAY 2 — Swahili Lexicon (_compute_lexicon_score):
          • Operates on clean_text BEFORE token mapping.
          • Catches Swahili/Sheng words the token map left untranslated.
          • Handles negation (si, bila, hapana) and intensifiers (sana, kabisa).

        PATHWAY 3 — Complaint Signal Adjustment (sentiment_config.py):
          • Detects structural complaint signals VADER cannot see:
              - Contrastive conjunctions ("but", "however", "lakini")
              - Unmet expectation phrases ("did not", "was supposed to")
              - Complaint verbs/nouns ("delayed", "failed", "tatizo")
              - Domain negative lexicon (agricultural & NGO programme words)
              - Extended 8-token negation window (vs VADER\'s 3-token window)
          • Returns a negative adjustment added BEFORE blending.
          • Hedged-language detection is applied alongside this pathway.

        BLENDING FORMULA (language-weighted, rebalanced for complaint data):
          ┌────────────────┬──────────────┬────────────────┐
          │ Detected lang  │  VADER wt.   │  Lexicon wt.   │
          ├────────────────┼──────────────┼────────────────┤
          │ sw (Swahili)   │    40%       │     60%        │
          │ other African  │    55%       │     45%        │
          │ en (English)   │    75%       │     25%        │
          └────────────────┴──────────────┴────────────────┘
          VADER weight reduced: VADER drifts positive on unknown tokens.
          The complaint layer + lexicon corrects for this bias.
        """
        lang_col = "lang" in self.q_df.columns
        texts = self.q_df["clean_text"].tolist()
        langs = self.q_df["lang"].tolist() if lang_col else ["en"] * len(texts)

        scores = []
        for text, lang in zip(texts, langs):
            lang = str(lang).strip() if lang else "en"

            # ── Pathway 1: VADER on normalized text ──────────────────
            base = self.sia.polarity_scores(text)["compound"]

            # ── Pathway 3: Complaint signals + hedged language ────────
            # compute_complaint_adjustment covers structural complaint
            # signals, domain lexicon, and extended negation window.
            # detect_hedged_sentiment covers polite deflections.
            complaint_adj = compute_complaint_adjustment(text, lang)
            hedge_adj     = self.detect_hedged_sentiment(text)
            vader_score   = max(-1.0, min(1.0, base + complaint_adj + hedge_adj))

            # ── Pathway 2: Swahili lexicon on clean text ──────────────
            lex_score = self._compute_lexicon_score(text)

            # ── Language-weighted blend ───────────────────────────────
            # VADER weight reduced vs previous version — VADER
            # systematically drifts positive on unknown tokens.
            if lang == "sw":
                final = 0.40 * vader_score + 0.60 * lex_score
            elif lang in ("und", "af", "so", "yo", "ha", "ig", "am"):
                final = 0.55 * vader_score + 0.45 * lex_score
            else:
                # English (or unknown → safe default)
                final = 0.75 * vader_score + 0.25 * lex_score

            scores.append(max(-1.0, min(1.0, final)))

        self.q_df["sentiment"] = scores

    # ══════════════════════════════════════════════════════════════════
    # EMBEDDINGS  ─ deduplicated + memory-managed
    # ══════════════════════════════════════════════════════════════════

    def generate_embeddings(self) -> None:
        texts = self.q_df["compressed_text"].tolist()
        unique_texts = list(dict.fromkeys(texts))  # order-preserving dedup

        log.info("Encoding %d unique sentence units …", len(unique_texts))

        # batch_size=32: safe for the multilingual L12 model on 8 GB RAM.
        # (L6 English-only model could use 64, but L12 is ~2× the parameters.)
        unique_embeddings = self.model.encode(
            unique_texts,
            batch_size=32,
            show_progress_bar=False,
            convert_to_numpy=True,
        )

        embedding_map = dict(zip(unique_texts, unique_embeddings))
        self.embeddings = np.array([embedding_map[t] for t in texts])

        # PCA: reduce 384 → config.pca_components to save RAM & speed up KMeans
        n_components = min(
            self.config.pca_components,
            self.embeddings.shape[0] - 1,
            self.embeddings.shape[1],
        )
        log.info("PCA: %d → %d dims", self.embeddings.shape[1], n_components)
        pca = PCA(n_components=n_components, random_state=42)
        self.embeddings = pca.fit_transform(self.embeddings).astype(np.float32)

        del unique_embeddings
        gc.collect()

    # ═══════════════════════════════════════════════════════════════════
    # EMBEDDING CACHE (v3+ Upgrade)
    # ═══════════════════════════════════════════════════════════════════
    # Dramatic speed boost on re-runs: cache embeddings to disk and reuse

    def get_embeddings_cached(
        self,
        texts: List[str],
        cache_path: str = "embeddings_cache.pkl",
        force_recompute: bool = False,
    ) -> np.ndarray:
        """
        Get embeddings with disk caching for huge speed boost on re-runs.
        
        PERFORMANCE: First run: 60s. Cached re-runs: <1s. ~60× speedup.
        
        Args:
            texts: list of strings to encode
            cache_path: where to save pickle cache
            force_recompute: if True, ignore cache and recompute
        
        Returns:
            np.ndarray of embeddings (shape: [len(texts), embedding_dim])
        """
        cache_file = Path(cache_path)
        
        # Try loading from cache
        if cache_file.exists() and not force_recompute:
            try:
                with open(cache_file, "rb") as f:
                    cached = pickle.load(f)
                if len(cached) == len(texts):
                    log.info("✓ Loaded embeddings from cache (%s)", cache_path)
                    return cached
            except Exception as exc:
                log.warning("Cache load failed (%s). Recomputing…", exc)
        
        # Compute fresh embeddings
        log.info("Computing embeddings (no valid cache) …")
        embeddings = self.model.encode(
            texts,
            batch_size=32,
            show_progress_bar=True,
            convert_to_numpy=True,
        )
        
        # Save to cache
        try:
            with open(cache_file, "wb") as f:
                pickle.dump(embeddings, f)
            log.info("✓ Saved embeddings cache to %s", cache_path)
        except Exception as exc:
            log.warning("Could not save cache (%s)", exc)
        
        return embeddings

    # ══════════════════════════════════════════════════════════════════
    # CLUSTER COUNT SELECTION
    # ══════════════════════════════════════════════════════════════════

    def choose_clusters(self) -> int:
        n = len(self.q_df)

        if n < 200:
            best_k, best_score = 2, -1.0
            for k in range(2, min(9, n)):
                km = KMeans(n_clusters=k, random_state=42, n_init=10)
                labels = km.fit_predict(self.embeddings)
                score = silhouette_score(self.embeddings, labels,
                                         sample_size=min(500, n))
                if score > best_score:
                    best_k, best_score = k, score
            # Section 9: store so generate_dataset_summary() can report it
            self._last_silhouette_score = round(float(best_score), 3)
            return best_k

        suggested = int(np.sqrt(n / 2))
        k = min(suggested, self.config.max_clusters)
        # Section 9: compute silhouette post-hoc for large datasets too
        # (sampled at 1 000 points so it stays fast on 8 GB RAM)
        self._last_silhouette_score = None   # computed after clustering in cluster_feedback()
        return k

    # ══════════════════════════════════════════════════════════════════
    # CLUSTERING  ─ MiniBatchKMeans for large datasets
    # ══════════════════════════════════════════════════════════════════

    def cluster_feedback(self) -> None:
        k = self.choose_clusters()
        log.info("Clustering into k=%d groups …", k)

        n = len(self.q_df)
        if n > self.config.large_dataset_threshold:
            self.kmeans = MiniBatchKMeans(
                n_clusters=k,
                random_state=42,
                n_init=5,
                batch_size=min(1024, n),
            )
        else:
            self.kmeans = KMeans(n_clusters=k, random_state=42, n_init=10)

        self.q_df["cluster"] = self.kmeans.fit_predict(self.embeddings)

        # Section 9: for large datasets compute silhouette AFTER fitting.
        # Sample at most 1 000 points so it stays fast on 8 GB RAM.
        if self._last_silhouette_score is None and k >= 2:
            try:
                sil = silhouette_score(
                    self.embeddings,
                    self.q_df["cluster"],
                    sample_size=min(1_000, n),
                    random_state=42,
                )
                self._last_silhouette_score = round(float(sil), 3)
            except Exception:
                self._last_silhouette_score = None

        quality = ""
        if self._last_silhouette_score is not None:
            s = self._last_silhouette_score
            quality = "good" if s > 0.35 else ("acceptable" if s > 0.20 else "weak")
            log.info(
                "Silhouette score: %.3f (%s clustering — "
                ">0.35 good, >0.20 acceptable, ≤0.20 weak)",
                s, quality,
            )

    # ══════════════════════════════════════════════════════════════════
    # FILTER CLUSTERS  ─ separate emerging vs discarded
    # ══════════════════════════════════════════════════════════════════

    def filter_clusters(self) -> None:
        counts = self.q_df["cluster"].value_counts()
        min_s = self.config.min_cluster_size
        emerging_min = self.config.emerging_min_size

        valid = counts[counts >= min_s].index
        emerging = counts[(counts >= emerging_min) & (counts < min_s)].index

        self.emerging_df = self.q_df[
            self.q_df["cluster"].isin(emerging)
        ].copy().reset_index(drop=True)

        self.q_df = self.q_df[
            self.q_df["cluster"].isin(valid)
        ].copy().reset_index(drop=True)

        log.info(
            "Clusters: %d main | %d emerging | discarded: %d",
            len(valid), len(emerging),
            len(counts) - len(valid) - len(emerging),
        )

        # Re-sync the TF-IDF matrix rows to match the filtered q_df.
        # extract_keywords() uses q_df.index values to slice tfidf_matrix,
        # so both must share the same positional indices after filtering.
        valid_positions = [
            i for i, c in enumerate(
                self.kmeans.labels_
            ) if c in valid
        ]
        self.tfidf_matrix = self.tfidf_matrix[valid_positions]

    # ══════════════════════════════════════════════════════════════════
    # KEYWORD EXTRACTION  ─ phrase-aware, generic-filtered
    # ══════════════════════════════════════════════════════════════════

    GENERIC_WORDS = {
        "service", "good", "bad", "thing", "things", "experience",
        "people", "place", "really", "very", "much", "feel", "make",
        "like", "know", "need", "want", "use", "get", "go", "come",
        "say", "tell", "ask", "give", "take", "see", "think",
    }

    def extract_keywords(
        self, cluster_indices: List[int], top_n: int = 6
    ) -> List[str]:
        cluster_matrix = self.tfidf_matrix[cluster_indices]
        cluster_scores = np.asarray(cluster_matrix.mean(axis=0)).ravel()
        global_scores = np.asarray(self.tfidf_matrix.mean(axis=0)).ravel()
        importance = cluster_scores - global_scores

        keywords: List[str] = []
        for idx in importance.argsort()[::-1]:
            term = self.feature_names[idx]
            if term in self.GENERIC_WORDS:
                continue
            keywords.append(term)
            if len(keywords) >= top_n:
                break
        return keywords

    # ══════════════════════════════════════════════════════════════════
    # THEME NAMING  ─ LLM-first with smart rule-based fallback
    # ══════════════════════════════════════════════════════════════════

    def generate_theme_name(
        self, keywords: List[str], samples: Optional[List[str]] = None
    ) -> str:
        """
        Public entry-point for theme naming.

        Priority order (Section 8 — Theme Registry):
          1. Theme Registry  — stable name from a previous run (disk-persisted).
             This ensures the same cluster of keywords always produces the
             same label across runs, preventing report confusion.
          2. In-memory cache — already computed this run (avoids duplicate calls).
          3. LLM path        — Gemini (Section 10) or Claude API call.
          4. Rule-based      — improved bigram/unigram fallback (always available).
        """
        if not keywords:
            return "Other Feedback"

        # Fingerprint: sorted top-3 keywords (stable across calls)
        cache_key = "|".join(sorted(keywords[:3]))

        # 1. Registry lookup (Section 8) — survives across runs
        if cache_key in self._theme_registry:
            # Warm the in-memory cache too so downstream code is consistent
            self._theme_name_cache[cache_key] = self._theme_registry[cache_key]
            return self._theme_registry[cache_key]

        # 2. In-memory cache — already resolved this run
        if cache_key in self._theme_name_cache:
            return self._theme_name_cache[cache_key]

        # 3+4. Generate name via LLM or rule-based
        if self._gemini_client and self.config.use_llm_naming:
            # Section 10: Gemini path
            name = self._gemini_theme_name(keywords, samples or [])
        elif self._anthropic_client and self.config.use_llm_naming:
            # Existing Claude path
            name = self._llm_theme_name(keywords, samples or [])
        else:
            name = self._rule_based_theme_name(keywords)

        # Store in both caches (registry is flushed to disk at end of run)
        self._theme_name_cache[cache_key] = name
        return name

    def _rule_based_theme_name(self, keywords: List[str]) -> str:
        """
        Improved rule-based naming — avoids gibberish like 'Training Seeds'
        or 'Seeds Seed'.

        Strategy:
          1. Bigram TF-IDF features (e.g. 'late delivery') are already
             meaningful phrases → prefer them as the theme.
          2. For unigrams, skip any keyword whose stem is already represented
             by another keyword in the pair (avoids 'Train Training').
          3. Build at most a 3-token theme; deduplicate within it.
        """
        bigrams  = [k for k in keywords if " " in k]
        unigrams = [k for k in keywords if " " not in k]

        # ── Case 1: use the best bigram as the primary label ──────────
        if bigrams:
            primary = bigrams[0].title()
            # Try to append a discriminating unigram if it adds new info
            for u in unigrams:
                if u.lower() not in bigrams[0].lower():
                    return f"{primary} — {u.title()}"
            return primary

        # ── Case 2: two unigrams — check for stem overlap ─────────────
        if len(unigrams) >= 2:
            p1, p2 = unigrams[0], unigrams[1]
            # Overlap check: one is a substring of the other?
            if p1.lower() in p2.lower() or p2.lower() in p1.lower():
                # Redundant pair → use longer one + third keyword
                primary = max(p1, p2, key=len).title()
                if len(unigrams) >= 3:
                    return f"{primary} {unigrams[2].title()}"
                return primary
            # Pair looks fine
            return f"{p1.title()} {p2.title()}"

        # ── Case 3: single keyword ─────────────────────────────────────
        return unigrams[0].title() if unigrams else "Other Feedback"

    def _llm_theme_name(
        self, keywords: List[str], samples: List[str]
    ) -> str:
        """
        Ask Claude (Haiku) for a meaningful 2-4 word theme name.
        Falls back to rule-based on any API error.
        """
        sector_ctx = (
            f"The data comes from a {self.config.sector} programme."
            if self.config.sector else ""
        )
        sample_text = " | ".join(samples[:3]) if samples else "(none)"
        prompt = (
            f"You are labelling themes extracted from survey feedback. "
            f"{sector_ctx}\n\n"
            f"Top keywords: {', '.join(keywords[:6])}\n"
            f"Example responses: {sample_text}\n\n"
            "Give a clear, meaningful 2–4 word theme label that a non-technical "
            "M&E officer would immediately understand. "
            "Return ONLY the theme name — no punctuation, no explanation."
        )
        try:
            response = self._anthropic_client.messages.create(
                model=self.config.anthropic_model,
                max_tokens=20,
                messages=[{"role": "user", "content": prompt}],
            )
            raw = response.content[0].text.strip().strip('"').strip("'")
            # Sanity: reject if response is too long or empty
            if 2 <= len(raw.split()) <= 6 and raw:
                return raw.title()
            return self._rule_based_theme_name(keywords)
        except Exception as exc:
            log.debug("LLM theme naming failed (%s); using rule-based.", exc)
            return self._rule_based_theme_name(keywords)

    def _gemini_theme_name(
        self, keywords: List[str], samples: List[str]
    ) -> str:
        """
        Section 10 — Ask Gemini for a meaningful 2-4 word theme name.
        Mirrors _llm_theme_name() exactly; falls back to rule-based on any error.

        Uses temperature=0.0 (via generation_config) for maximum consistency,
        which is the Gemini equivalent of the deterministic Claude setting.
        """
        sector_ctx = (
            f"The data comes from a {self.config.sector} programme."
            if self.config.sector else ""
        )
        sample_text = " | ".join(samples[:3]) if samples else "(none)"
        prompt = (
            f"You are labelling themes extracted from survey feedback. "
            f"{sector_ctx}\n\n"
            f"Top keywords: {', '.join(keywords[:6])}\n"
            f"Example responses: {sample_text}\n\n"
            "Give a clear, meaningful 2–4 word theme label that a non-technical "
            "M&E officer would immediately understand. "
            "Return ONLY the theme name — no punctuation, no explanation."
        )
        try:
            gen_cfg = _gemini_lib.types.GenerationConfig(
                temperature=0.0,
                max_output_tokens=20,
            )
            response = self._gemini_client.generate_content(
                prompt,
                generation_config=gen_cfg,
            )
            raw = response.text.strip().strip('"').strip("'")
            if 2 <= len(raw.split()) <= 6 and raw:
                return raw.title()
            return self._rule_based_theme_name(keywords)
        except Exception as exc:
            log.debug("Gemini theme naming failed (%s); using rule-based.", exc)
            return self._rule_based_theme_name(keywords)

    # ══════════════════════════════════════════════════════════════════
    # CLUSTER SUMMARY BUILD
    # ══════════════════════════════════════════════════════════════════

    def build_cluster_summary(self) -> None:
        self.cluster_summary = {}
        for cid in sorted(self.q_df["cluster"].unique()):
            subset  = self.q_df[self.q_df["cluster"] == cid]
            texts   = subset["clean_text"].tolist()
            indices = subset.index.tolist()
            kw      = self.extract_keywords(indices, self.config.top_keywords)

            # ── Section 7: Respondent Count Fix ──────────────────────
            # sentence_count  = number of sentence units in this cluster.
            #                   One long response can produce several units,
            #                   so this OVER-counts distinct people.
            # respondent_count = unique orig_idx values = ACTUAL number of
            #                   distinct respondents who contributed to this
            #                   theme.  This is the correct "how many people
            #                   mentioned this" metric.
            sentence_count    = len(texts)
            respondent_count  = (
                subset["orig_idx"].nunique()
                if "orig_idx" in subset.columns
                else sentence_count          # graceful fallback if orig_idx absent
            )

            self.cluster_summary[cid] = {
                "count":            sentence_count,    # raw sentence units (internal)
                "respondent_count": respondent_count,  # PRIMARY metric (Section 7)
                "keywords":         kw,
                "samples":          texts[:5],
            }
            # Pre-warm the theme name cache while cluster data is fresh
            self.generate_theme_name(kw, texts[:3])

    # ══════════════════════════════════════════════════════════════════
    # SUMMARY TABLE  ─ % based on per-question raw response count
    # ══════════════════════════════════════════════════════════════════

    def build_summary_table(self) -> None:
        """
        Build a standardized summary of themes.

        This revised implementation standardizes theme structure, computes
        percentile-based thresholds, reinitializes the TriggerEngine with
        data-driven thresholds, applies Trigger/Action engines to
        each theme, validates results, and produces `self.summary_df`.

        SECTION 7 — RESPONDENT COUNT FIX
        ──────────────────────────────────
        Two counts are now tracked separately:
          • Respondent Count — how many UNIQUE respondents mentioned this theme.
            This is the truthful "how widespread is this" metric and drives
            Share (%) and the Trigger Engine impact thresholds.
          • Sentence Count   — total sentence units in the cluster.
            Informational only; useful for debugging clustering quality.

        Key rule: Impact fed to TriggerEngine = Respondent Count.
        Share (%) denominator = total unique respondents for the question.
        """
        total = self._question_response_count  # set in run() per question

        # 1) Collect raw theme records from cluster summary
        themes = []
        for cid, info in self.cluster_summary.items():
            theme_name = self.generate_theme_name(info.get("keywords", []))
            subset = self.q_df[self.q_df["cluster"] == cid]
            avg_sent = subset["sentiment"].mean() if len(subset) > 0 else 0.0

            # ── Section 7: use respondent_count as canonical count ────
            respondent_count = int(info.get("respondent_count", info.get("count", 0)))
            sentence_count   = int(info.get("count", 0))

            # Share (%) is now respondent-based — no more sentence inflation
            pct = round((respondent_count / max(total, 1)) * 100, 2)

            # intensity: few very negative sentences vs many mildly negative
            negative_sentences = subset[subset["sentiment"] < -0.3]
            intensity = (
                negative_sentences["sentiment"].abs().mean()
                if len(negative_sentences) > 0 else 0.0
            )
            # Impact Score uses respondent_count for a fair severity signal
            severity = respondent_count * abs(avg_sent)

            themes.append({
                "cluster_id":      cid,
                "name":            theme_name,
                "impact":          respondent_count,    # PRIMARY metric (Section 7)
                "respondent_count": respondent_count,   # alias — explicit label
                "sentence_count":  sentence_count,      # secondary / informational
                "share_pct":       pct,
                "impact_score":    round(severity, 2),
                "intensity":       round(intensity, 3),
                "sentiment_score": round(avg_sent, 3),
                "keywords":        [k.strip() for k in info.get("keywords", []) if k.strip()],
                "representative_quote": info.get("samples", [""])[0] if info.get("samples") else "",
                "recommendation":  info.get("recommendation", "") or "",
            })

        # 2) Map sentiment scores -> labels (strict mapping per spec)
        for t in themes:
            sc = t.get("sentiment_score", 0.0)
            if sc <= -0.05:
                t["sentiment"] = "Problem"
            elif sc >= 0.05:
                t["sentiment"] = "Positive"
            else:
                t["sentiment"] = "Neutral"

        # 3) Compute percentile-based thresholds for impact
        impacts = sorted([t["impact"] for t in themes])
        high_thresh = None
        med_thresh = None
        if len(impacts) == 0:
            high_thresh = med_thresh = 0
        elif len(impacts) < 5:
            # small dataset: use max and median
            high_thresh = max(impacts)
            med_thresh = int(pd.Series(impacts).median())
        else:
            # 80th and 50th percentiles
            high_thresh = int(pd.Series(impacts).quantile(0.8))
            med_thresh = int(pd.Series(impacts).quantile(0.5))

        # Handle uniform values edge-case
        if high_thresh is None:
            high_thresh = med_thresh = 0
        if high_thresh == med_thresh and len(impacts) > 0:
            # keep them equal; TriggerEngine will handle equality
            pass

        log.info(f"  ▶ Computed thresholds — HIGH: {high_thresh}, MEDIUM: {med_thresh}")

        # 4) Reinitialize TriggerEngine with new thresholds (overrides previous)
        try:
            from trigger_engine import TriggerEngine
            self.trigger_engine = TriggerEngine({"high_impact": high_thresh, "medium_impact": med_thresh})
        except Exception:
            log.warning("TriggerEngine import failed — continuing with existing engine")

        # 5) Apply all engines to each theme and enrich
        enriched = []
        for t in themes:
            theme_input = {
                "name": t["name"],
                "impact": t["impact"],
                "sentiment": t.get("sentiment", "Neutral"),
                "sentiment_score": t.get("sentiment_score", 0.0),
                "keywords": t.get("keywords", []),
                "recommendation": t.get("recommendation", ""),
            }

            # Trigger
            triggers = []
            if getattr(self, "trigger_engine", None) and self.pipeline_config.get("features", {}).get("enable_trigger_engine", True):
                try:
                    triggers = self.trigger_engine.evaluate(theme_input) or []
                except Exception as exc:
                    log.debug("Trigger engine error for %s: %s", t["name"], exc)

            # Determine priority label from triggers (pick highest)
            priority_label = "Low"
            if isinstance(triggers, list) and triggers:
                levels = [str(x.get("level", "")).upper() for x in triggers]
                if any(l == "HIGH" for l in levels):
                    priority_label = "High"
                elif any(l == "MEDIUM" for l in levels):
                    priority_label = "Medium"
                elif any(l == "POSITIVE" for l in levels):
                    priority_label = "Positive"

            # ── CRITICAL: feed priority back into theme_input so that
            #    action_engine.timeline() uses the correct value.
            #    Without this line, timeline() always defaults to "MEDIUM" → 30 days.
            theme_input["priority"] = priority_label

            # Trend is REMOVED — TriggerEngine is the single source of truth for priority

            # Action plan
            action_plan = {}
            if getattr(self, "action_engine", None) and self.pipeline_config.get("features", {}).get("enable_action_engine", True):
                try:
                    action_plan = self.action_engine.generate(theme_input) or {}
                except Exception as exc:
                    log.debug("Action engine error for %s: %s", t["name"], exc)

            # Compose actions string
            actions_list = action_plan.get("actions") if isinstance(action_plan.get("actions"), list) else []
            actions_joined = "; ".join(str(a) for a in actions_list) if actions_list else "No action specified"

            # Extract Trigger Status from triggers list (ALWAYS has one element)
            trigger_status = "No Alert"
            if isinstance(triggers, list) and triggers:
                trigger_status = triggers[0].get("message", "No Alert")
            
            enriched.append({
                "Cluster ID":       t["cluster_id"],
                "Theme":            t["name"],
                # ── Section 7: two count columns ──────────────────────
                "Respondent Count": t.get("respondent_count", t["impact"]),  # unique people
                "Sentence Count":   t.get("sentence_count", t["impact"]),    # sentence units
                "Share (%)":        t["share_pct"],
                "Impact Score":     t["impact_score"],
                "Intensity Score":  t["intensity"],
                "Average Sentiment":t["sentiment_score"],
                "Theme Sentiment":  t.get("sentiment", "Neutral"),
                "Key Keywords":     ", ".join(t.get("keywords", [])),
                "Representative Quote": t.get("representative_quote", ""),
                "triggers":         triggers,
                "Priority":         priority_label,
                "Trigger Status":   trigger_status,
                "Action Owner":     action_plan.get("owner", "Program Manager"),
                "Timeline (Days)":  action_plan.get("timeline", 30),
                "Actions":          actions_joined,
                "Recommendation":   t.get("recommendation", ""),
                "action_plan":      action_plan,
            })

        # 6) Validation: ensure required fields
        for rec in enriched:
            if not rec.get("Priority"):
                rec["Priority"] = "Low"
            if not rec.get("Trigger Status"):
                rec["Trigger Status"] = "No Alert"
            if not rec.get("Action Owner"):
                rec["Action Owner"] = "Program Manager"
            if not rec.get("Actions"):
                rec["Actions"] = "No action specified"

        # 7) Build final DataFrame
        self.summary_df = pd.DataFrame(enriched)
        # Sort by Impact Score desc for presentation
        if "Impact Score" in self.summary_df.columns:
            self.summary_df = self.summary_df.sort_values("Impact Score", ascending=False).reset_index(drop=True)

    # ═══════════════════════════════════════════════════════════════════
    # APPLY ADVANCED ENGINES (DEPRECATED)
    # ═══════════════════════════════════════════════════════════════════
    # Logic now moved to build_summary_table for single source of truth

    def apply_advanced_engines(self) -> None:
        """
        All advanced engine logic (triggers, actions) is now handled in
        build_summary_table() to ensure single source of truth.
        This method is retained for backward compatibility but does nothing.
        """
        # All logic now in build_summary_table
        if all(col in self.summary_df.columns for col in ("triggers", "Priority", "action_plan")):
            log.info("Advanced engines already applied in build_summary_table — skipping.")
            return
        
        log.info("✓ Advanced engines already applied")

    # ══════════════════════════════════════════════════════════════════
    # SENTIMENT LABEL  ─ keyword-aware + sentiment threshold
    # ══════════════════════════════════════════════════════════════════

    def classify_theme_sentiment(self) -> None:
        """Classify theme sentiment using BOTH sentiment score AND keywords.
        
        This prevents false neutrals where a theme seems like a problem
        but sentiment is borderline.
        
        Thresholds (v3.0):
        - Problem: sentiment < -0.05 OR (keywords match problem terms)
        - Positive: sentiment > 0.1 AND no problem keywords
        - Neutral: everything else
        """
        problem_keywords = {
            "problem", "issue", "complaint", "poor", "bad", "terrible",
            "awful", "horrible", "delay", "slow", "difficult", "challenge",
            "struggle", "frustrate", "disappoint", "dissat", "fail",
            "broken", "error", "bug", "stuck", "unavailable", "shortage",
            "late", "waiting", "queue", "rude", "unfriendly", "expensive",
            "costly", "mbaya", "tatizo", "shida", "lalamiko"
        }
        
        labels = []
        for idx, row in self.summary_df.iterrows():
            sentiment = row["Average Sentiment"]
            keywords = str(row.get("Key Keywords", "")).lower()
            
            # Check if problem keywords present
            has_problem_keyword = any(pk in keywords for pk in problem_keywords)
            
            if sentiment < -0.05 or has_problem_keyword:
                labels.append("Problem")
            elif sentiment > 0.1 and not has_problem_keyword:
                labels.append("Positive")
            else:
                labels.append("Neutral")
        
        self.summary_df["Theme Sentiment"] = labels

    # ══════════════════════════════════════════════════════════════════
    # RECOMMENDATIONS  ─ LLM-first, rule-based fallback, keyword tips
    # ══════════════════════════════════════════════════════════════════

    def generate_recommendations(self) -> None:
        """
        Generates one recommendation per theme.

        Mode A — Gemini (Section 10, use_gemini=True):
          Batch call to Gemini — all themes in one prompt → list of recs.

        Mode B — Claude/Anthropic (use_gemini=False, anthropic installed):
          Existing single batch API call to Claude Haiku.

        Mode C — Rule-based (fallback):
          Matches keywords against RECOMMENDATION_RULES.
          Prints KEYWORD TIPS report after matching.
        """
        if self._gemini_client and self.config.use_llm_recommendations:
            # Section 10: Gemini path
            self._gemini_generate_recommendations()
        elif self._anthropic_client and self.config.use_llm_recommendations:
            self._llm_generate_recommendations()
        else:
            self._rule_based_recommendations()
            self.print_keyword_tips(quiet=False)

    def _llm_generate_recommendations(self) -> None:
        """Batch API call: all themes → one request → list of recommendations."""
        sector_ctx = (
            f"The organisation works in the {self.config.sector} sector in "
            f"Kenya/East Africa."
            if self.config.sector
            else "The organisation operates in Kenya/East Africa."
        )

        # Build a compact JSON payload — only what the LLM needs
        themes_payload = [
            {
                "id": i,
                "theme": row["Theme"],
                "sentiment": row.get("Theme Sentiment", ""),
                "keywords": row["Key Keywords"],
                "quote": row["Representative Quote"],
            }
            for i, (_, row) in enumerate(self.summary_df.iterrows())
        ]

        prompt = (
            f"You are an experienced M&E advisor. {sector_ctx}\n\n"
            "For each feedback theme below, write ONE practical, specific "
            "recommendation (1–2 sentences). Use local African context where "
            "relevant (M-Pesa, boda-boda, SMS/USSD, county structures, etc.).\n\n"
            f"Themes:\n{json.dumps(themes_payload, indent=2)}\n\n"
            "Return a JSON array of strings — one recommendation per theme, "
            "in the SAME order as the input. "
            "Return ONLY the JSON array. No explanation, no markdown."
        )

        try:
            response = self._anthropic_client.messages.create(
                model=self.config.anthropic_model,
                max_tokens=800,
                messages=[{"role": "user", "content": prompt}],
            )
            raw = response.content[0].text.strip()
            # Strip markdown code fences if present
            raw = re.sub(r"^```[a-z]*\n?", "", raw)
            raw = re.sub(r"\n?```$", "", raw).strip()
            recs = json.loads(raw)

            if isinstance(recs, list) and len(recs) == len(self.summary_df):
                if "Recommendation" in self.summary_df.columns:
                    self.summary_df["Recommendation"] = recs
                else:
                    try:
                        self.summary_df.insert(5, "Recommendation", recs)
                    except Exception:
                        self.summary_df["Recommendation"] = recs
                return
            log.warning("LLM returned %d recs for %d themes — falling back.",
                        len(recs), len(self.summary_df))
        except Exception as exc:
            log.warning("LLM recommendation generation failed (%s) — "
                        "falling back to rule-based.", exc)

        # Fallback
        self._rule_based_recommendations()

    def _gemini_generate_recommendations(self) -> None:
        """
        Section 10 — Batch Gemini call: all themes → one prompt → list of recs.

        Mirrors _llm_generate_recommendations() exactly:
        • Same prompt structure and JSON-array response format.
        • Same code-fence stripping and length validation.
        • Falls back to rule-based on any error.

        Temperature is set to 0.0 for reproducible output.
        """
        sector_ctx = (
            f"The organisation works in the {self.config.sector} sector in "
            f"Kenya/East Africa."
            if self.config.sector
            else "The organisation operates in Kenya/East Africa."
        )

        themes_payload = [
            {
                "id": i,
                "theme": row["Theme"],
                "sentiment": row.get("Theme Sentiment", ""),
                "keywords": row["Key Keywords"],
                "quote": row["Representative Quote"],
            }
            for i, (_, row) in enumerate(self.summary_df.iterrows())
        ]

        prompt = (
            f"You are an experienced M&E advisor. {sector_ctx}\n\n"
            "For each feedback theme below, write ONE practical, specific "
            "recommendation (1–2 sentences). Use local African context where "
            "relevant (M-Pesa, boda-boda, SMS/USSD, county structures, etc.).\n\n"
            f"Themes:\n{json.dumps(themes_payload, indent=2)}\n\n"
            "Return a JSON array of strings — one recommendation per theme, "
            "in the SAME order as the input. "
            "Return ONLY the JSON array. No explanation, no markdown."
        )

        try:
            gen_cfg = _gemini_lib.types.GenerationConfig(
                temperature=0.0,
                max_output_tokens=1200,
            )
            response = self._gemini_client.generate_content(
                prompt,
                generation_config=gen_cfg,
            )
            raw = response.text.strip()
            # Strip markdown code fences if present
            raw = re.sub(r"^```[a-z]*\n?", "", raw)
            raw = re.sub(r"\n?```$", "", raw).strip()
            recs = json.loads(raw)

            if isinstance(recs, list) and len(recs) == len(self.summary_df):
                if "Recommendation" in self.summary_df.columns:
                    self.summary_df["Recommendation"] = recs
                else:
                    try:
                        self.summary_df.insert(5, "Recommendation", recs)
                    except Exception:
                        self.summary_df["Recommendation"] = recs
                log.info(
                    "Gemini generated %d recommendations successfully.", len(recs)
                )
                return
            log.warning(
                "Gemini returned %d recs for %d themes — falling back.",
                len(recs), len(self.summary_df),
            )
        except Exception as exc:
            log.warning(
                "Gemini recommendation generation failed (%s) — "
                "falling back to rule-based.", exc
            )

        # Fallback
        self._rule_based_recommendations()

    def _rule_based_recommendations(self) -> None:
        """Match theme keywords against RECOMMENDATION_RULES."""
        recs = []
        for _, row in self.summary_df.iterrows():
            text = (row["Theme"] + " " + row["Key Keywords"]).lower()
            matched = FALLBACK_RECOMMENDATION
            for rule_keywords, rec in RECOMMENDATION_RULES:
                if any(k in text for k in rule_keywords):
                    matched = rec
                    break
            recs.append(matched)
        if "Recommendation" in self.summary_df.columns:
            self.summary_df["Recommendation"] = recs
        else:
            try:
                self.summary_df.insert(5, "Recommendation", recs)
            except Exception:
                self.summary_df["Recommendation"] = recs

    def print_keyword_tips(self, quiet: bool = False) -> None:
        """
        ╔══════════════════════════════════════════════════════════════╗
        ║  KEYWORD TIPS — Recommendation Rule Coverage Report         ║
        ╚══════════════════════════════════════════════════════════════╝
        Prints all keywords found in this question, whether each matched
        a RECOMMENDATION_RULE, and which terms fell through to the
        generic fallback.

        Use this to extend RECOMMENDATION_RULES for your sector:
          1. Look at the ⚠ UNMATCHED section below.
          2. Group related unmatched terms into a new rule tuple.
          3. Write a sector-specific recommendation string.
          4. Add it to RECOMMENDATION_RULES near the top of the file.
        """
        if quiet or self.summary_df.empty:
            return

        all_rule_kw: set = set()
        for rule_kw, _ in RECOMMENDATION_RULES:
            all_rule_kw.update(rule_kw)

        SEP = "─" * 62
        print(f"\n{'═' * 62}")
        print("  KEYWORD TIPS  —  Recommendation Rule Coverage")
        print(f"{'═' * 62}")

        for _, row in self.summary_df.iterrows():
            theme   = row["Theme"]
            kw_str  = row.get("Key Keywords", "")
            rec     = row.get("Recommendation", "")
            kw_list = [k.strip() for k in kw_str.split(",") if k.strip()]

            matched   = [k for k in kw_list if k in all_rule_kw]
            unmatched = [k for k in kw_list if k not in all_rule_kw]
            used_fallback = (rec == FALLBACK_RECOMMENDATION)

            icon = "⚠ " if used_fallback else "✅"
            print(f"\n  {icon} Theme : {theme}")
            print(f"     Keywords : {', '.join(kw_list) or '(none)'}")
            if matched:
                print(f"     Matched rules on : {', '.join(matched)}")
            if unmatched:
                print(f"     ⚠ Unmatched terms : {', '.join(unmatched)}")
                print(f"       → Add these to RECOMMENDATION_RULES for "
                      f"better coverage.")
            if used_fallback:
                print(f"     ⚠ Used generic fallback — no rule matched.")

        print(f"\n{SEP}")
        print("  TIP: To add a rule, insert this pattern into RECOMMENDATION_RULES:")
        print('    (["keyword1", "keyword2", ...],')
        print('     "Your sector-specific recommendation here."),')
        print(f"{SEP}\n")

    # ══════════════════════════════════════════════════════════════════
    # PRIORITY LABELS  ─ intensity-aware
    # ══════════════════════════════════════════════════════════════════

    # ══════════════════════════════════════════════════════════════════
    # MERGE SIMILAR THEMES
    # ══════════════════════════════════════════════════════════════════

    def merge_similar_themes(self) -> None:
        themes = self.summary_df["Theme"].tolist()
        
        # Initialize mapping for all clusters
        self.cluster_to_theme_map = {}
        if "Cluster ID" in self.summary_df.columns:
            for idx, row in self.summary_df.iterrows():
                cluster_id = row.get("Cluster ID", -1)
                self.cluster_to_theme_map[cluster_id] = row.get("Theme", "Unknown")
        
        if len(themes) <= 1:
            return

        embeddings = self.model.encode(
            themes, convert_to_numpy=True, normalize_embeddings=True
        )

        merged_map: Dict[str, List[int]] = {}
        used: set = set()

        for i, theme_i in enumerate(themes):
            if i in used:
                continue
            merged_map[theme_i] = [i]
            for j in range(i + 1, len(themes)):
                if j in used:
                    continue
                if np.dot(embeddings[i], embeddings[j]) >= self.config.merge_threshold:
                    merged_map[theme_i].append(j)
                    used.add(j)

        # Update cluster ID mapping with merged themes
        self.cluster_to_theme_map = {}
        
        new_rows = []
        for main_theme, idxs in merged_map.items():
            subset = self.summary_df.iloc[idxs]
            # Get all cluster IDs being merged
            cluster_ids = subset["Cluster ID"].tolist() if "Cluster ID" in subset.columns else []
            for cid in cluster_ids:
                self.cluster_to_theme_map[cid] = main_theme
            
            # When multiple themes are merged, inherit Priority/Owner/Timeline
            # from the highest-priority row (HIGH > MEDIUM > POSITIVE > LOW).
            # This prevents a HIGH-priority theme from being downgraded because
            # it merged with a LOW one.
            PRIORITY_RANK = {"HIGH": 4, "MEDIUM": 3, "POSITIVE": 2, "LOW": 1, "": 0}
            best_idx = subset["Priority"].map(
                lambda p: PRIORITY_RANK.get(str(p).upper(), 0)
            ).idxmax()
            best_row = subset.loc[best_idx]

            row = {
                "Cluster ID": cluster_ids[0] if cluster_ids else -1,
                "Theme": main_theme,
                # ── Section 7: sum both count columns ─────────────────
                # NOTE: summing respondent counts across merged themes may
                # slightly over-count respondents who contributed to both
                # clusters (rare at merge_threshold=0.75). The alternative
                # (union of orig_idx sets) requires storing orig_idx in
                # summary_df, which is done via q_df but not propagated here.
                # For client reporting the summed value is acceptable and
                # clearly labelled.
                "Respondent Count": subset["Respondent Count"].sum() if "Respondent Count" in subset.columns else subset.get("Response Count", pd.Series([0])).sum(),
                "Sentence Count":   subset["Sentence Count"].sum()   if "Sentence Count"   in subset.columns else 0,
                "Share (%)": round(subset["Share (%)"].sum(), 2),
                "Impact Score": round(subset["Impact Score"].sum(), 2),
                "Intensity Score": round(subset["Intensity Score"].mean(), 3),
                "Average Sentiment": round(subset["Average Sentiment"].mean(), 3),
                "Key Keywords": ", ".join(subset["Key Keywords"].tolist()),
                "Representative Quote": subset["Representative Quote"].iloc[0],
                "Recommendation": best_row.get("Recommendation", subset["Recommendation"].iloc[0]),
                "Priority": best_row.get("Priority", "Low"),
                "Trigger Status": (
                    best_row.get("Trigger Status", "No Alert")
                    if "Trigger Status" in subset.columns else "No Alert"
                ),
                # ─ Preserve Action Owner and Timeline from the best-priority row ─
                "Action Owner": (
                    best_row.get("Action Owner", "Program Manager")
                    if "Action Owner" in subset.columns else "Program Manager"
                ),
                "Timeline (Days)": (
                    best_row.get("Timeline (Days)", 60)
                    if "Timeline (Days)" in subset.columns else 60
                ),
                "Actions": (
                    best_row.get("Actions", "No action specified")
                    if "Actions" in subset.columns else "No action specified"
                ),
                # ─ Preserve advanced engine columns ──────────────────
                "action_plan": (
                    best_row.get("action_plan", {})
                    if "action_plan" in subset.columns else {}
                ),
                "triggers": (
                    best_row.get("triggers", [])
                    if "triggers" in subset.columns else []
                ),
            }
            new_rows.append(row)

        self.summary_df = pd.DataFrame(new_rows).sort_values(
            "Impact Score", ascending=False
        )
        # Recompute sentiment label after merge
        self.classify_theme_sentiment()

    # ══════════════════════════════════════════════════════════════════
    # KEY INSIGHTS
    # ══════════════════════════════════════════════════════════════════

    def generate_key_insights(self) -> None:
        insights: List[str] = []
        df = self.summary_df

        if df.empty:
            self.key_insights = ["Insufficient data for insights."]
            return

        top = df.iloc[0]
        insights.append(
            f"Most common theme: '{top['Theme']}' — "
            f"appearing in {top['Share (%)']}% of responses."
        )

        most_neg = df.sort_values("Average Sentiment").iloc[0]
        if most_neg["Theme"] != top["Theme"]:
            insights.append(
                f"Highest dissatisfaction: '{most_neg['Theme']}' "
                f"(avg sentiment {most_neg['Average Sentiment']:.2f})."
            )

        high_intensity = df[df["Intensity Score"] > 0.5]
        if not high_intensity.empty:
            themes_hi = ", ".join(f"'{t}'" for t in high_intensity["Theme"].head(3))
            insights.append(
                f"High emotional intensity detected in: {themes_hi}. "
                "A small number of respondents expressed strong dissatisfaction — "
                "these are early-warning signals."
            )

        # Loud vs widespread distinction (Section 7: now uses Respondent Count)
        resp_col = "Respondent Count" if "Respondent Count" in df.columns else "Response Count"
        loud = df[
            (df[resp_col] <= df[resp_col].quantile(0.25)) &
            (df["Average Sentiment"] < -0.3)
        ]
        if not loud.empty:
            insights.append(
                f"Note: '{loud.iloc[0]['Theme']}' has few respondents but high negative "
                "sentiment — this may be a loud minority, not a widespread issue."
            )

        positives = df[df["Theme Sentiment"] == "Positive"]
        if not positives.empty:
            pt = positives.iloc[0]["Theme"]
            insights.append(
                f"Top positive area: '{pt}' — reinforce and communicate "
                "this strength to stakeholders."
            )

        self.key_insights = insights

    # ══════════════════════════════════════════════════════════════════
    # DATASET SUMMARY
    # ══════════════════════════════════════════════════════════════════

    def generate_dataset_summary(self) -> None:
        avg_len = self.q_df["clean_text"].apply(
            lambda x: len(str(x).split())
        ).mean()

        lang_counts = self.q_df.get("lang", pd.Series(dtype=str)).value_counts()
        lang_note = ", ".join(
            f"{lang}:{cnt}" for lang, cnt in lang_counts.head(4).items()
        ) if not lang_counts.empty else "n/a"

        # ── Section 7: Respondent coverage stats ──────────────────────
        # Count how many unique respondents appear in at least one theme
        # vs the total respondents for the question.
        resp_col = "Respondent Count" if "Respondent Count" in self.summary_df.columns else "Response Count"
        total_respondents = self._question_response_count

        # Unique respondents captured across ALL themes (via q_df orig_idx)
        # — this is the ground-truth figure from the sentence-level data.
        respondents_in_themes = (
            self.q_df["orig_idx"].nunique()
            if "orig_idx" in self.q_df.columns
            else total_respondents
        )
        coverage_pct = round(
            (respondents_in_themes / max(total_respondents, 1)) * 100, 1
        )

        self.dataset_summary = {
            "Total Responses":                    total_respondents,
            "Respondents Captured in Themes":     respondents_in_themes,
            "Theme Coverage (%)":                 coverage_pct,
            "Sentence Units Analysed":            len(self.q_df),
            "Avg Sentences per Respondent":       round(
                len(self.q_df) / max(respondents_in_themes, 1), 1
            ),
            "Average Response Length (words)":    round(avg_len, 1),
            "Themes Identified":                  len(self.summary_df),
            "Emerging Issues Found":              len(
                getattr(self, "emerging_df", pd.DataFrame())
            ),
            "Languages Detected":                 lang_note,
            # Section 9 — Silhouette Score
            "Silhouette Score":                   self._last_silhouette_score,
            "Cluster Quality":                    (
                "Good (>0.35)" if (self._last_silhouette_score or 0) > 0.35
                else "Acceptable (>0.20)" if (self._last_silhouette_score or 0) > 0.20
                else "Weak (≤0.20)" if self._last_silhouette_score is not None
                else "N/A"
            ),
        }

    # ══════════════════════════════════════════════════════════════════
    # LABEL RESPONSES
    # ══════════════════════════════════════════════════════════════════

    def label_data(self) -> None:
        theme_map = {
            cid: self.generate_theme_name(info["keywords"])
            for cid, info in self.cluster_summary.items()
        }
        self.q_df["Theme"] = self.q_df["cluster"].map(theme_map)

    # ══════════════════════════════════════════════════════════════════
    # CROSS-QUESTION THEME DETECTION
    # ══════════════════════════════════════════════════════════════════

    def detect_cross_question_themes(self) -> None:
        """Find themes recurring in ≥2 questions — a key gap in most tools."""
        from collections import Counter
        all_themes: List[str] = []
        for qr in self.question_results.values():
            all_themes.extend(qr["summary"]["Theme"].tolist())

        counts = Counter(all_themes)
        self.cross_question_themes = [
            t for t, c in counts.most_common() if c >= 2
        ]
        if self.cross_question_themes:
            log.info(
                "Cross-question recurring themes: %s",
                self.cross_question_themes,
            )

    # ══════════════════════════════════════════════════════════════════
    # EXCEL STYLE HELPERS
    # ══════════════════════════════════════════════════════════════════

    HEADER_COLOUR  = "2F5597"   # navy blue (with white text)
    HEADING_COLOUR = "1F3864"   # darker navy for sheet title bar (with white text)
    STRIPE_COLOUR  = "EEF2FF"   # soft lavender stripe
    SUBHEAD_COLOUR = "D6E4F7"   # light blue for section sub-headers (with BLACK text)

    ACCENT_COLOURS = {
        "Problem":  "FFE0E0",   # soft red
        "Positive": "E0FFE8",   # soft green
        "Neutral":  "FFF9E0",   # soft yellow
    }

    # Preferred minimum column widths by header name (characters)
    COL_MIN_WIDTHS = {
        "Theme":                          32,
        "Question":                       20,
        # Section 7: both count columns
        "Respondent Count":               18,
        "Sentence Count":                 16,
        "Response Count":                 18,   # backward-compat fallback
        "Share (%)":                      14,
        "Impact Score":                   16,
        "Intensity Score":                16,
        "Average Sentiment":              20,
        "Priority":                       20,
        "Theme Sentiment":                18,
        "Key Keywords":                   38,
        "Representative Quote":           52,
        "Recommendation":                 48,
        "Quote":                          55,
        "Languages Detected":             28,
        "Respondents Captured in Themes": 28,
        "Theme Coverage (%)":             20,
        "Avg Sentences per Respondent":   26,
    }
    COL_DEFAULT_MIN = 14   # fallback minimum for any column not listed above

    def _thin_border(self):
        thin = Side(style="thin")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def _medium_border(self):
        med = Side(style="medium")
        return Border(left=med, right=med, top=med, bottom=med)

    # ------------------------------------------------------------------
    # write_sheet_heading
    #   Writes a professional, centered, merged title bar at row 1.
    #   Returns the next available row number (always row 3, leaving a
    #   blank spacer row beneath the heading).
    # ------------------------------------------------------------------
    def write_sheet_heading(
        self,
        ws,
        title: str,
        subtitle: str = "",
        merge_cols: int = 10,
    ) -> int:
        """
        Row 1 : merged navy title bar — large white bold text, centered.
        Row 2 : merged subtitle bar (smaller, italic) — OR blank spacer.
        Returns 3 so callers know where the table starts.
        """
        navy_fill = PatternFill(
            start_color=self.HEADING_COLOUR,
            end_color=self.HEADING_COLOUR,
            fill_type="solid",
        )
        sub_fill = PatternFill(
            start_color=self.HEADER_COLOUR,
            end_color=self.HEADER_COLOUR,
            fill_type="solid",
        )

        end_col = get_column_letter(merge_cols)

        # ── Row 1 : main title ────────────────────────────────────────
        ws.merge_cells(f"A1:{end_col}1")
        title_cell = ws["A1"]
        title_cell.value = title
        title_cell.font = Font(
            name="Calibri", size=16, bold=True, color="FFFFFF"
        )
        title_cell.fill = navy_fill
        title_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=False
        )
        ws.row_dimensions[1].height = 32

        # ── Row 2 : subtitle / generated date ────────────────────────
        ws.merge_cells(f"A2:{end_col}2")
        sub_cell = ws["A2"]
        sub_cell.value = subtitle
        sub_cell.font = Font(
            name="Calibri", size=10, italic=True, color="FFFFFF"
        )
        sub_cell.fill = sub_fill
        sub_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 18

        return 3   # next usable row

    # ------------------------------------------------------------------
    # style_table_header  —  styles a specific row (the column-header row)
    # ------------------------------------------------------------------
    def style_table_header(self, ws, header_row: int = 1) -> None:
        fill = PatternFill(
            start_color=self.HEADER_COLOUR,
            end_color=self.HEADER_COLOUR,
            fill_type="solid",
        )
        for cell in ws[header_row]:
            if cell.value is None:
                continue
            # Use white text for navy headers
            cell.font = Font(
                name="Calibri", size=10, bold=True, color="FFFFFF"
            )
            cell.fill = fill
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = self._thin_border()
        ws.row_dimensions[header_row].height = 28

    # ------------------------------------------------------------------
    # zebra_rows  —  light stripe on every even data row
    # ------------------------------------------------------------------
    def zebra_rows(self, ws, start_row: int = 2) -> None:
        fill = PatternFill(
            start_color=self.STRIPE_COLOUR,
            end_color=self.STRIPE_COLOUR,
            fill_type="solid",
        )
        for r in range(start_row, ws.max_row + 1):
            if r % 2 == 0:
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r, column=col)
                    # Don't overwrite sentiment-coloured rows
                    if cell.fill.fgColor.rgb in (
                        "00000000", "FF" + self.STRIPE_COLOUR
                    ):
                        cell.fill = fill

    # ------------------------------------------------------------------
    # add_table_borders
    # ------------------------------------------------------------------
    def add_table_borders(self, ws, start_row: int = 1) -> None:
        border = self._thin_border()
        for row in ws.iter_rows(
            min_row=start_row, max_row=ws.max_row,
            min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                cell.border = border

    # ------------------------------------------------------------------
    # smart_col_widths
    #   Uses COL_MIN_WIDTHS for known columns; measures content otherwise.
    #   Applies a comfortable padding and hard caps.
    # ------------------------------------------------------------------
    def smart_col_widths(self, ws) -> None:
        for i, col_cells in enumerate(ws.columns, 1):
            header_val = str(ws.cell(row=1, column=i).value or "")
            # Check both row-1 (heading) and row-3/4 for actual header name
            for check_row in (1, 2, 3, 4):
                hdr = str(ws.cell(row=check_row, column=i).value or "")
                if hdr in self.COL_MIN_WIDTHS:
                    header_val = hdr
                    break

            preferred_min = self.COL_MIN_WIDTHS.get(
                header_val, self.COL_DEFAULT_MIN
            )

            # Also measure actual content
            max_content = max(
                (len(str(c.value)) for c in col_cells if c.value), default=0
            )
            # Use whichever is bigger, but cap at 60
            width = min(max(preferred_min, max_content + 3), 60)
            ws.column_dimensions[get_column_letter(i)].width = width

    # ------------------------------------------------------------------
    # sentiment_colour_rows
    # ------------------------------------------------------------------
    def sentiment_colour_rows(self, ws, sentiment_col_idx: int,
                               start_row: int = 2) -> None:
        for r in range(start_row, ws.max_row + 1):
            cell = ws.cell(row=r, column=sentiment_col_idx)
            colour = self.ACCENT_COLOURS.get(str(cell.value))
            if colour:
                fill = PatternFill(
                    start_color=colour, end_color=colour, fill_type="solid"
                )
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=col).fill = fill

    # ------------------------------------------------------------------
    # _data_row_height  —  set comfortable height on all data rows
    # ------------------------------------------------------------------
    @staticmethod
    def _set_data_row_heights(ws, start_row: int, height: float = 18) -> None:
        for r in range(start_row, ws.max_row + 1):
            if ws.row_dimensions[r].height is None or \
               ws.row_dimensions[r].height < height:
                ws.row_dimensions[r].height = height

    # ═══════════════════════════════════════════════════════════════════
    # PREPARE SUMMARY FOR EXCEL (v3+ Upgrade)
    # ═══════════════════════════════════════════════════════════════════

    def prepare_summary_for_excel(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Convert complex trigger/action dicts to readable Excel-friendly strings.
        
        Args:
            df: summary_df from question_results
        
        Returns:
            Modified dataframe with trigger/action columns as strings
        """
        export_df = df.copy()
        
        # Trigger Status should already be in summary_df from build_summary_table
        if "Trigger Status" not in export_df.columns and "triggers" in export_df.columns:
            export_df["Trigger Status"] = export_df["triggers"].apply(
                lambda triggers: (
                    " | ".join([t.get("icon", "") + " " + t.get("level", "") 
                               for t in triggers])
                    if isinstance(triggers, list) and triggers else "No Alert"
                )
            )
        elif "Trigger Status" not in export_df.columns:
            export_df["Trigger Status"] = "No Alert"
        
        # Action Owner and Timeline should already be in summary_df from build_summary_table
        if "Action Owner" not in export_df.columns or "Timeline (Days)" not in export_df.columns:
            if "action_plan" in export_df.columns:
                if "Action Owner" not in export_df.columns:
                    export_df["Action Owner"] = export_df["action_plan"].apply(
                        lambda ap: (
                            ap.get("owner", "Program Manager") if isinstance(ap, dict) else "Program Manager"
                        )
                    )
                
                if "Timeline (Days)" not in export_df.columns:
                    export_df["Timeline (Days)"] = export_df["action_plan"].apply(
                        lambda ap: (
                            str(ap.get("timeline", 60)) if isinstance(ap, dict) else "60"
                        )
                    )
            else:
                if "Action Owner" not in export_df.columns:
                    export_df["Action Owner"] = "Program Manager"
                if "Timeline (Days)" not in export_df.columns:
                    export_df["Timeline (Days)"] = "60"
        
        return export_df

    # ══════════════════════════════════════════════════════════════════
    # PDF SUMMARY REPORT  (Section 4)
    # ══════════════════════════════════════════════════════════════════

    def build_pdf_report(self, output: Optional[str] = None) -> None:
        """
        Generate a 2-page printable PDF executive brief.

        Page 1
        ------
        * Title bar (report title + generated timestamp)
        * Survey snapshot box (total responses / questions / themes / alerts)
        * Cross-question recurring themes (highlighted, if any)
        * HIGH & MEDIUM priority alert table

        Page 2+
        -------
        * Per-question blocks: theme table (Theme | Share% | Priority |
          Owner | Recommendation excerpt)

        Requires fpdf2 (pip install fpdf2).
        If not installed the method logs a hint and returns immediately
        without raising an exception.
        """
        if not FPDF_AVAILABLE:
            log.warning(
                "PDF report skipped — fpdf2 not installed. "
                "Fix:  pip install fpdf2"
            )
            return

        # ── Resolve output path ───────────────────────────────────────
        if output:
            pdf_path = output
        elif self.config.pdf_filename:
            pdf_path = self.config.pdf_filename
        else:
            base = self.config.output_filename
            pdf_path = re.sub(r"\.xlsx?$", ".pdf", base, flags=re.IGNORECASE)
            if not pdf_path.endswith(".pdf"):
                pdf_path += ".pdf"

        now_eat = datetime.now(EAT).strftime("%d %b %Y  %H:%M EAT")

        # ── Latin-1 sanitizer ─────────────────────────────────────────
        # fpdf2 built-in fonts (Helvetica) only cover Latin-1 (0x00-0xFF).
        # Replace any out-of-range characters before passing to cell().
        _CHAR_MAP = {
            "\u2014": " - ", "\u2013": "-",
            "\u2019": "'",   "\u2018": "'",
            "\u201c": '"',   "\u201d": '"',
            "\u2022": "*",   "\u2026": "...",
            "\u00d7": "x",   "\u2192": "->",
            "\u25b6": ">",   "\u2714": "v",
            "\u274c": "x",   "\u26a0": "!",
            "\U0001f534": "[HIGH]", "\U0001f7e1": "[MED]",
            "\U0001f7e2": "[POS]",  "\u26aa": "[LOW]",
        }

        def _s(text: str) -> str:
            """Sanitize to Latin-1; replace known specials first."""
            for ch, rep in _CHAR_MAP.items():
                text = text.replace(ch, rep)
            return text.encode("latin-1", errors="replace").decode("latin-1")

        # ── Colour palette (matches Excel scheme) ─────────────────────
        NAVY       = (31,  56, 100)
        MID_BLUE   = (47,  85, 151)
        LIGHT_BLUE = (214, 228, 247)
        WHITE      = (255, 255, 255)
        BLACK      = (0,   0,   0)
        DARK_GREY  = (60,  60,  60)
        STRIPE     = (238, 242, 255)

        PRI_BG = {
            "HIGH":     (255, 179, 179),
            "MEDIUM":   (255, 224, 179),
            "POSITIVE": (198, 239, 206),
            "LOW":      (240, 240, 240),
        }
        PRI_FG = {
            "HIGH":     (180, 0,   0),
            "MEDIUM":   (150, 75,  0),
            "POSITIVE": (0,   97,  0),
            "LOW":      (100, 100, 100),
        }

        # ── FPDF subclass: running header/footer on page 2+ ───────────
        _title   = _s(self.config.report_title)
        _gen     = _s(now_eat)
        _navy    = NAVY
        _mid     = MID_BLUE
        _white   = WHITE
        _dg      = DARK_GREY
        _black   = BLACK

        class _PDF(FPDF):
            def header(self_):
                if self_.page_no() == 1:
                    return
                self_.set_fill_color(*_mid)
                self_.rect(0, 0, 210, 8, "F")
                self_.set_y(2)
                self_.set_font("Helvetica", "B", 7)
                self_.set_text_color(*_white)
                self_.cell(0, 4, _title, align="C")
                self_.set_text_color(*_black)
                self_.ln(6)

            def footer(self_):
                self_.set_y(-12)
                self_.set_font("Helvetica", "", 7)
                self_.set_text_color(140, 140, 140)
                self_.cell(
                    0, 5,
                    f"Page {self_.page_no()}  |  {_gen}  |  Confidential",
                    align="C",
                )
                self_.set_text_color(*_black)

        pdf = _PDF(orientation="P", unit="mm", format="A4")
        pdf.set_margins(left=14, top=14, right=14)
        pdf.set_auto_page_break(auto=True, margin=14)
        pdf.add_page()

        EPW = pdf.epw    # effective page width  (~182 mm)
        LH  = 5.5        # standard line height
        SH  = 7.0        # section header height

        # ── Helpers ───────────────────────────────────────────────────
        def sec_hdr(title: str) -> None:
            pdf.set_fill_color(*NAVY)
            pdf.set_text_color(*WHITE)
            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(EPW, SH, _s(f"  {title}"), border=0, align="L",
                     fill=True, new_x="LMARGIN", new_y="NEXT")
            pdf.ln(1)

        def tbl_hdr(widths: list, labels: list) -> None:
            pdf.set_fill_color(*MID_BLUE)
            pdf.set_text_color(*WHITE)
            pdf.set_font("Helvetica", "B", 7)
            for w, lbl in zip(widths, labels):
                pdf.cell(w, SH, _s(f" {lbl}"), border=1, align="L",
                         fill=True, new_x="RIGHT", new_y="TOP")
            pdf.ln(SH)

        def data_row(widths: list, values: list,
                     bg: tuple, fg: tuple,
                     bold_col: int = -1) -> None:
            pdf.set_fill_color(*bg)
            for ci, (w, v) in enumerate(zip(widths, values)):
                style = "B" if ci == bold_col else ""
                colour = fg if ci == bold_col else DARK_GREY
                pdf.set_text_color(*colour)
                pdf.set_font("Helvetica", style, 7)
                pdf.cell(w, LH, _s(f" {str(v)[:int(w*1.9)]}"),
                         border=1, align="L", fill=True,
                         new_x="RIGHT", new_y="TOP")
            pdf.ln(LH)

        # ══════════════════════════════════════════════════════════════
        # PAGE 1  —  TITLE BAR
        # ══════════════════════════════════════════════════════════════
        pdf.set_fill_color(*NAVY)
        pdf.rect(0, 0, 210, 22, "F")
        pdf.set_y(4)
        pdf.set_font("Helvetica", "B", 15)
        pdf.set_text_color(*WHITE)
        pdf.cell(0, 8, _s(self.config.report_title),
                 align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 8)
        pdf.cell(0, 5, _s(f"Generated: {now_eat}"),
                 align="C", new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(*BLACK)
        pdf.ln(6)

        # ══════════════════════════════════════════════════════════════
        # PAGE 1  —  SURVEY SNAPSHOT
        # ══════════════════════════════════════════════════════════════
        total_resp = sum(
            qd["dataset_summary"].get("Total Responses", 0)
            for qd in self.question_results.values()
        )
        total_themes = sum(
            len(qd["summary"]) for qd in self.question_results.values()
        )
        high_n = sum(
            (qd["summary"].get("Priority", pd.Series()).str.upper() == "HIGH").sum()
            for qd in self.question_results.values()
            if "Priority" in qd["summary"].columns
        )

        sec_hdr("SURVEY SNAPSHOT")
        box_w = EPW / 4
        labels = ["Total Responses", "Questions Analysed",
                  "Themes Identified", "HIGH Alerts"]
        values = [str(total_resp), str(len(self.question_results)),
                  str(total_themes), str(high_n)]

        # Label row
        pdf.set_fill_color(*LIGHT_BLUE)
        pdf.set_text_color(*DARK_GREY)
        pdf.set_font("Helvetica", "B", 7)
        for lbl in labels:
            pdf.cell(box_w, LH, _s(f"  {lbl}"), border=1, align="L",
                     fill=True, new_x="RIGHT", new_y="TOP")
        pdf.ln(LH)

        # Value row
        pdf.set_fill_color(*WHITE)
        pdf.set_text_color(*NAVY)
        pdf.set_font("Helvetica", "B", 13)
        for val in values:
            pdf.cell(box_w, 9, val, border=1, align="C",
                     fill=True, new_x="RIGHT", new_y="TOP")
        pdf.ln(9)
        pdf.ln(3)

        # ══════════════════════════════════════════════════════════════
        # PAGE 1  —  CROSS-QUESTION THEMES
        # ══════════════════════════════════════════════════════════════
        if self.cross_question_themes:
            sec_hdr("THEMES RECURRING ACROSS MULTIPLE QUESTIONS")
            pdf.set_fill_color(255, 235, 235)
            pdf.set_text_color(150, 0, 0)
            pdf.set_font("Helvetica", "I", 8)
            themes_str = "  *  ".join(self.cross_question_themes[:6])
            pdf.multi_cell(EPW, LH, _s(f"  {themes_str}"),
                           border=1, align="L", fill=True)
            pdf.ln(3)

        # ══════════════════════════════════════════════════════════════
        # PAGE 1  —  PRIORITY ALERTS TABLE
        # ══════════════════════════════════════════════════════════════
        sec_hdr("PRIORITY ALERTS  -  HIGH & MEDIUM Issues")

        cw = [18, 52, 30, 20, 32, 16]
        tbl_hdr(cw, ["Priority", "Theme", "Question",
                      "Responses", "Owner", "Days"])

        alerts = []
        for qname, qdata in self.question_results.items():
            for _, row in qdata["summary"].iterrows():
                pri = str(row.get("Priority", "")).upper()
                if pri in ("HIGH", "MEDIUM"):
                    # Section 7: show respondent count in PDF alert table
                    impact_val = row.get(
                        "Respondent Count",
                        row.get("Response Count", "")
                    )
                    alerts.append({
                        "pri":   pri,
                        "theme": row.get("Theme", ""),
                        "q":     qname,
                        "count": impact_val,
                        "owner": row.get("Action Owner", ""),
                        "days":  str(row.get("Timeline (Days)", "")),
                    })

        alerts.sort(key=lambda a: (0 if a["pri"] == "HIGH" else 1,
                                   -int(a["count"] or 0)))

        if not alerts:
            pdf.set_font("Helvetica", "I", 8)
            pdf.set_text_color(*DARK_GREY)
            pdf.set_fill_color(*WHITE)
            pdf.cell(EPW, LH,
                     "  No HIGH or MEDIUM priority issues detected.",
                     border=1, align="L", fill=True,
                     new_x="LMARGIN", new_y="NEXT")
        else:
            for i, a in enumerate(alerts[:16]):
                bg = PRI_BG.get(a["pri"], WHITE)
                fg = PRI_FG.get(a["pri"], DARK_GREY)
                # Alternate slight shade
                shade = tuple(min(255, v + 10) for v in bg)
                row_bg = bg if i % 2 == 0 else shade
                vals = [a["pri"], a["theme"], a["q"],
                        str(a["count"]), a["owner"], a["days"]]
                data_row(cw, vals, row_bg, fg, bold_col=0)

        pdf.ln(2)

        # ══════════════════════════════════════════════════════════════
        # PAGE 2+  —  PER-QUESTION THEME TABLES
        # ══════════════════════════════════════════════════════════════
        pdf.add_page()

        qcw  = [50, 14, 18, 30, 70]
        qhdr = ["Theme", "Resp.", "Priority", "Owner", "Recommendation"]
        # "Resp." = Respondent Count (abbreviated to fit PDF column)

        for qname, qdata in self.question_results.items():
            s_df  = qdata["summary"]
            ds    = qdata["dataset_summary"]
            insig = qdata["insights"]

            q_label  = qname.replace("_", " ").title()
            resp_n   = ds.get("Total Responses", "")
            themes_n = ds.get("Themes Identified", "")
            # Section 9: include silhouette score in PDF question header
            sil_val  = ds.get("Silhouette Score")
            sil_str  = f"  |  sil: {sil_val:.3f}" if isinstance(sil_val, float) else ""

            # Question block header
            pdf.set_fill_color(*NAVY)
            pdf.set_text_color(*WHITE)
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(
                EPW, 6,
                _s(f"  {q_label}  "
                   f"({resp_n} responses  {themes_n} themes{sil_str})"),
                border=0, align="L", fill=True,
                new_x="LMARGIN", new_y="NEXT",
            )
            pdf.ln(0.5)

            # First insight
            if insig:
                pdf.set_fill_color(*LIGHT_BLUE)
                pdf.set_text_color(*DARK_GREY)
                pdf.set_font("Helvetica", "I", 7)
                pdf.multi_cell(EPW, 4.5,
                               _s(f"  > {insig[0]}"),
                               border=0, align="L", fill=True)
            pdf.ln(0.5)

            # Column headers
            tbl_hdr(qcw, qhdr)

            # Data rows — top 6 themes
            for row_i, (_, row) in enumerate(s_df.head(6).iterrows()):
                pri  = str(row.get("Priority", "LOW")).upper()
                bg   = PRI_BG.get(pri, WHITE)
                fg   = PRI_FG.get(pri, DARK_GREY)
                base = bg if row_i % 2 == 0 else STRIPE
                rec  = str(row.get("Recommendation", ""))
                rec_short = rec.split(".")[0][:75]
                if len(rec) > 75:
                    rec_short += "..."

                # Section 7: show respondent count in PDF (more meaningful than share%)
                resp_count = row.get(
                    "Respondent Count",
                    row.get("Response Count", "—")
                )
                vals = [
                    str(row.get("Theme", ""))[:38],
                    str(resp_count),
                    pri,
                    str(row.get("Action Owner", ""))[:22],
                    rec_short,
                ]
                data_row(qcw, vals, base, fg, bold_col=2)

            pdf.ln(4)

        # ── Save ──────────────────────────────────────────────────────
        try:
            pdf.output(pdf_path)
            log.info("PDF report saved -> %s", pdf_path)
        except Exception as exc:
            log.error("PDF save failed: %s", exc)

    # ══════════════════════════════════════════════════════════════════
    # GEOGRAPHIC SEGMENTATION  (Section 5)
    # ══════════════════════════════════════════════════════════════════

    def build_geographic_breakdown(self) -> None:
        """
        Compute per-location theme distributions across all questions.

        Requires:
          - self.config.location_column set to a valid column in self.df
          - orig_idx propagated through preprocess() (done in Section 5)

        Populates self.geo_breakdown:
        {
            question_name: {
                location_value: {
                    "theme_counts":  {theme_name: int},
                    "avg_sentiment": float,
                    "top_issue":     str,
                    "total":         int,
                }
            }
        }
        """
        loc_col = self.config.location_column
        if not loc_col:
            self.geo_breakdown: Dict = {}
            return

        if loc_col not in self.df.columns:
            log.warning(
                "location_column '%s' not found in data — "
                "geographic breakdown skipped. Available: %s",
                loc_col,
                ", ".join(self.df.columns.tolist()),
            )
            self.geo_breakdown = {}
            return

        log.info("Building geographic breakdown on column '%s' …", loc_col)
        self.geo_breakdown = {}
        min_resp = self.config.location_min_responses

        for qname, qdata in self.question_results.items():
            q_df    = qdata.get("q_df", pd.DataFrame()).copy()
            s_df    = qdata.get("summary", pd.DataFrame())

            if q_df.empty or "orig_idx" not in q_df.columns:
                log.debug("No orig_idx in q_df for %s — skipping geo.", qname)
                continue

            # Build theme label lookup from cluster_to_theme_map
            cluster_map = qdata.get("cluster_to_theme_map", {})

            # Attach theme name to each sentence unit
            if "cluster" in q_df.columns:
                q_df["Theme"] = q_df["cluster"].map(cluster_map).fillna("Unclassified")
            elif "Theme" not in q_df.columns:
                log.debug("No cluster/Theme column in q_df for %s.", qname)
                continue

            # Join location from original df via orig_idx
            loc_series = self.df[loc_col].astype(str)
            q_df["location"] = q_df["orig_idx"].map(loc_series)

            # Drop rows with missing location
            q_df = q_df.dropna(subset=["location"])
            q_df = q_df[q_df["location"].str.strip() != ""]

            if q_df.empty:
                continue

            # Group by location
            loc_breakdown: Dict = {}
            for loc_val, grp in q_df.groupby("location"):
                if len(grp) < min_resp:
                    continue   # skip locations with too few data points

                theme_counts = (
                    grp["Theme"]
                    .value_counts()
                    .to_dict()
                )
                avg_sent = (
                    grp["sentiment"].mean()
                    if "sentiment" in grp.columns
                    else 0.0
                )
                top_issue = max(theme_counts, key=theme_counts.get) if theme_counts else "—"

                loc_breakdown[str(loc_val)] = {
                    "theme_counts":  theme_counts,
                    "avg_sentiment": round(float(avg_sent), 3),
                    "top_issue":     top_issue,
                    "total":         len(grp),
                }

            if loc_breakdown:
                self.geo_breakdown[qname] = loc_breakdown
                log.info(
                    "  %s → %d locations with ≥%d responses",
                    qname, len(loc_breakdown), min_resp,
                )

    # ══════════════════════════════════════════════════════════════════
    # EXCEL REPORT  ─ multi-question, fully rebuilt
    # ══════════════════════════════════════════════════════════════════

    def build_excel_report(self, output: Optional[str] = None) -> None:
        output = output or self.config.output_filename
        now_eat = datetime.now(EAT).strftime("%d %b %Y %H:%M EAT")

        wb = Workbook()
        wb.remove(wb.active)

        # ══════════════════════════════════════════════════════════════
        # 1. EXECUTIVE SUMMARY (Fully Table-Based - v4.0)
        #    All content now in proper tables with borders to prevent overlapping
        # ══════════════════════════════════════════════════════════════
        ws = wb.create_sheet("Executive Summary")
        EXEC_MERGE = 5   # columns for heading merge

        next_row = self.write_sheet_heading(
            ws,
            title=self.config.report_title,
            subtitle=f"Generated: {now_eat}   |   Questions analysed: "
                     f"{len(self.question_results)}",
            merge_cols=EXEC_MERGE,
        )

        # blank spacer
        next_row += 1   # row 4

        # ─────────────────────────────────────────────────────────────
        # TABLE 1: Questions Analysed
        # ─────────────────────────────────────────────────────────────
        header = ws.cell(row=next_row, column=1, value="QUESTIONS ANALYSED")
        header.font = Font(bold=True, size=11, color="000000")
        header.fill = PatternFill(
            start_color=self.SUBHEAD_COLOUR,
            end_color=self.SUBHEAD_COLOUR,
            fill_type="solid",
        )
        header.border = self._thin_border()
        ws.merge_cells(start_row=next_row, start_column=1, 
                       end_row=next_row, end_column=EXEC_MERGE)
        ws.row_dimensions[next_row].height = 20
        next_row += 1

        for q in self.question_results:
            c = ws.cell(row=next_row, column=1, value=q)
            c.alignment = Alignment(wrap_text=True, vertical="center")
            c.border = self._thin_border()
            ws.merge_cells(start_row=next_row, start_column=1,
                          end_row=next_row, end_column=EXEC_MERGE)
            ws.row_dimensions[next_row].height = 18
            next_row += 1
        next_row += 1

        # ─────────────────────────────────────────────────────────────
        # TABLE 2: Cross-Question Themes (if any)
        # ─────────────────────────────────────────────────────────────
        if self.cross_question_themes:
            header = ws.cell(row=next_row, column=1, 
                           value="⚠  THEMES RECURRING ACROSS MULTIPLE QUESTIONS")
            header.font = Font(bold=True, size=11, color="FFFFFF")
            header.fill = PatternFill(
                start_color="C00000",
                end_color="C00000",
                fill_type="solid",
            )
            header.border = self._thin_border()
            ws.merge_cells(start_row=next_row, start_column=1,
                          end_row=next_row, end_column=EXEC_MERGE)
            ws.row_dimensions[next_row].height = 20
            next_row += 1

            for t in self.cross_question_themes:
                c = ws.cell(row=next_row, column=1, value=t)
                c.alignment = Alignment(wrap_text=True, vertical="center")
                c.border = self._thin_border()
                c.fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0",
                                    fill_type="solid")
                ws.merge_cells(start_row=next_row, start_column=1,
                              end_row=next_row, end_column=EXEC_MERGE)
                ws.row_dimensions[next_row].height = 18
                next_row += 1
            next_row += 1

        # ─────────────────────────────────────────────────────────────
        # Per-Question Summary Blocks (each in a table)
        # ─────────────────────────────────────────────────────────────
        for qname, qdata in self.question_results.items():
            s_df: pd.DataFrame = qdata["summary"]
            insights: List[str]  = qdata["insights"]
            ds: Dict             = qdata["dataset_summary"]

            # Section divider (full width header)
            divider = ws.cell(row=next_row, column=1, value=f"{qname.upper()}")
            divider.font = Font(bold=True, size=12, color="FFFFFF")
            divider.fill = PatternFill(
                start_color=self.HEADER_COLOUR,
                end_color=self.HEADER_COLOUR,
                fill_type="solid",
            )
            divider.border = self._thin_border()
            ws.merge_cells(
                start_row=next_row, start_column=1,
                end_row=next_row, end_column=EXEC_MERGE
            )
            ws.row_dimensions[next_row].height = 22
            next_row += 1

            # Stats table — Section 7: show both respondent & sentence counts
            stats_headers = [
                ("Total Respondents",         ds.get("Total Responses", "")),
                ("Respondents in Themes",     ds.get("Respondents Captured in Themes", "")),
                ("Theme Coverage (%)",         ds.get("Theme Coverage (%)", "")),
                ("Sentence Units Analysed",   ds.get("Sentence Units Analysed", "")),
                ("Avg Sentences / Respondent",ds.get("Avg Sentences per Respondent", "")),
                ("Avg Response Length (words)",ds.get("Average Response Length (words)", "")),
                ("Themes Identified",         ds.get("Themes Identified", "")),
                ("Languages",                 ds.get("Languages Detected", "")),
                # Section 9 — Silhouette Score
                ("Silhouette Score",          ds.get("Silhouette Score", "N/A")),
                ("Cluster Quality",           ds.get("Cluster Quality", "N/A")),
            ]
            # Section 9 — colour map for Cluster Quality value cell
            _QUALITY_COLOUR = {
                "Good":       "C6EFCE",   # green
                "Acceptable": "FFEB9C",   # amber
                "Weak":       "FFC7CE",   # red
            }

            for stat_label, stat_value in stats_headers:
                lc = ws.cell(row=next_row, column=1, value=stat_label)
                lc.font = Font(bold=True, size=10, color="000000")
                lc.fill = PatternFill(
                    start_color=self.SUBHEAD_COLOUR,
                    end_color=self.SUBHEAD_COLOUR,
                    fill_type="solid",
                )
                lc.border = self._thin_border()
                lc.alignment = Alignment(horizontal="left", vertical="center")

                vc = ws.cell(row=next_row, column=2, value=stat_value)
                vc.font = Font(bold=True, size=10)
                vc.border = self._thin_border()
                vc.alignment = Alignment(horizontal="center", vertical="center")

                # Section 9: colour-code the Cluster Quality value cell
                if stat_label == "Cluster Quality" and isinstance(stat_value, str):
                    for kw, hex_col in _QUALITY_COLOUR.items():
                        if kw in stat_value:
                            vc.fill = PatternFill(
                                start_color=hex_col,
                                end_color=hex_col,
                                fill_type="solid",
                            )
                            break

                ws.merge_cells(start_row=next_row, start_column=2,
                              end_row=next_row, end_column=EXEC_MERGE)
                ws.row_dimensions[next_row].height = 18
                next_row += 1
            next_row += 1

            # Insights table (single column with wrapped text)
            if insights:
                h = ws.cell(row=next_row, column=1, value="KEY INSIGHTS")
                h.font = Font(bold=True, size=10, color="000000")
                h.fill = PatternFill(
                    start_color=self.SUBHEAD_COLOUR,
                    end_color=self.SUBHEAD_COLOUR,
                    fill_type="solid",
                )
                h.border = self._thin_border()
                ws.merge_cells(start_row=next_row, start_column=1,
                              end_row=next_row, end_column=EXEC_MERGE)
                ws.row_dimensions[next_row].height = 18
                next_row += 1

                for ins in insights:
                    ic = ws.cell(row=next_row, column=1, value=f"•  {ins}")
                    ic.alignment = Alignment(wrap_text=True, vertical="top")
                    ic.border = self._thin_border()
                    ws.merge_cells(start_row=next_row, start_column=1,
                                  end_row=next_row, end_column=EXEC_MERGE)
                    ws.row_dimensions[next_row].height = 22
                    next_row += 1
                next_row += 1

            # Mini table header (Top themes for this question)
            mini_cols = ["Theme", "Responses", "Share (%)", "Priority",
                         "Recommendation"]
            for ci, h in enumerate(mini_cols, 1):
                c = ws.cell(row=next_row, column=ci, value=h)
                c.font = Font(bold=True, color="FFFFFF", size=9)
                c.fill = PatternFill(
                    start_color=self.HEADER_COLOUR,
                    end_color=self.HEADER_COLOUR,
                    fill_type="solid",
                )
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = self._thin_border()
            ws.row_dimensions[next_row].height = 20
            next_row += 1

            # Mini table data
            stripe_fill = PatternFill(
                start_color=self.STRIPE_COLOUR,
                end_color=self.STRIPE_COLOUR,
                fill_type="solid",
            )
            # Section 7: use Respondent Count (primary) with graceful fallback
            resp_col_name = (
                "Respondent Count" if "Respondent Count" in s_df.columns
                else "Response Count"
            )
            for row_i, (_, r) in enumerate(s_df.head(5).iterrows()):
                fill = stripe_fill if row_i % 2 == 1 else None
                vals = [
                    r["Theme"],
                    r[resp_col_name],
                    r["Share (%)"],
                    r.get("Priority", ""),
                    r.get("Recommendation", ""),
                ]
                for ci, val in enumerate(vals, 1):
                    c = ws.cell(row=next_row, column=ci, value=val)
                    c.border = self._thin_border()
                    c.alignment = Alignment(wrap_text=True, vertical="top")
                    if fill:
                        c.fill = fill
                ws.row_dimensions[next_row].height = 20
                next_row += 1
            next_row += 2

        self.smart_col_widths(ws)
        # Exec summary: give the Recommendation column extra room
        if ws.max_column >= 5:
            ws.column_dimensions["E"].width = 52

        # ══════════════════════════════════════════════════════════════
        # 2. PER-QUESTION THEME SHEETS (with v3+ Engines)
        # ══════════════════════════════════════════════════════════════
        # Section 7: "Respondent Count" is now the primary people metric.
        # "Sentence Count" is shown as a secondary column so analysts can
        # see how many sentence units a theme generated — useful for
        # evaluating clustering density.
        THEME_COLS = [
            "Theme", "Respondent Count", "Sentence Count", "Share (%)",
            "Impact Score", "Intensity Score", "Priority", "Theme Sentiment",
            "Average Sentiment", "Trigger Status", "Action Owner",
            "Timeline (Days)", "Key Keywords", "Representative Quote",
            "Recommendation",
        ]

        for qname, qdata in self.question_results.items():
            safe_name = re.sub(r"[^\w ]", "", qname)[:28].strip()
            ws = wb.create_sheet(safe_name)

            ds = qdata["dataset_summary"]
            subtitle = (
                f"Responses: {ds.get('Total Responses', '')}   |   "
                f"Themes: {ds.get('Themes Identified', '')}   |   "
                f"Languages: {ds.get('Languages Detected', '')}"
            )

            s_df: pd.DataFrame = qdata["summary"]
            # Prepare summary with trigger/action info
            s_df = self.prepare_summary_for_excel(s_df)
            available_cols = [c for c in THEME_COLS if c in s_df.columns]
            n_cols = max(len(available_cols), 8)

            next_row = self.write_sheet_heading(
                ws,
                title=f"Theme Analysis  —  {qname.replace('_', ' ').title()}",
                subtitle=subtitle,
                merge_cols=n_cols,
            )
            # next_row == 3 → write column headers here
            for ci, col_name in enumerate(available_cols, 1):
                ws.cell(row=next_row, column=ci, value=col_name)
            self.style_table_header(ws, header_row=next_row)
            data_start = next_row + 1

            for r in s_df[available_cols].itertuples(index=False):
                ws.append(list(r))

            # Wrap long-text columns
            for long_col in ("Representative Quote", "Recommendation",
                              "Key Keywords", "Action Owner", "Trigger Status"):
                if long_col in available_cols:
                    lcol = available_cols.index(long_col) + 1
                    for row in ws.iter_rows(
                        min_row=data_start, min_col=lcol, max_col=lcol
                    ):
                        for cell in row:
                            cell.alignment = Alignment(
                                wrap_text=True, vertical="top"
                            )

            # Colour trigger status cells (RED for HIGH, ORANGE for MEDIUM)
            if "Trigger Status" in available_cols:
                trigger_col = available_cols.index("Trigger Status") + 1
                for row in ws.iter_rows(
                    min_row=data_start, max_row=ws.max_row,
                    min_col=trigger_col, max_col=trigger_col
                ):
                    for cell in row:
                        cell_value = str(cell.value or "").upper()
                        if "HIGH" in cell_value:
                            cell.fill = PatternFill(
                                start_color="FF0000", end_color="FF0000",
                                fill_type="solid"
                            )
                            cell.font = Font(color="FFFFFF", bold=True)
                        elif "MEDIUM" in cell_value:
                            cell.fill = PatternFill(
                                start_color="FFA500", end_color="FFA500",
                                fill_type="solid"
                            )
                            cell.font = Font(color="FFFFFF", bold=True)
                        elif "POSITIVE" in cell_value:
                            cell.fill = PatternFill(
                                start_color="00B050", end_color="00B050",
                                fill_type="solid"
                            )
                            cell.font = Font(color="FFFFFF", bold=True)
                        else:
                            cell.fill = PatternFill(
                                start_color="F0F0F0", end_color="F0F0F0",
                                fill_type="solid"
                            )

            # Colour by sentiment, then stripe
            if "Theme Sentiment" in available_cols:
                sent_col = available_cols.index("Theme Sentiment") + 1
                self.sentiment_colour_rows(ws, sent_col, start_row=data_start)

            self.add_table_borders(ws, start_row=next_row)
            ws.auto_filter.ref = (
                f"A{next_row}:{get_column_letter(len(available_cols))}"
                f"{ws.max_row}"
            )
            self._set_data_row_heights(ws, start_row=data_start, height=20)
            self.smart_col_widths(ws)

        # ══════════════════════════════════════════════════════════════
        # 3. CHARTS SHEET (Professional Styling - v3.0)
        #    Features:
        #    - Clean professional styling with modern chart look
        #    - No gridlines for clarity
        #    - Clear titles and axis labels
        #    - Legend positioned right (not overlapping pie)
        # ══════════════════════════════════════════════════════════════
        ws = wb.create_sheet("Charts")

        CHART_BLOCK_ROWS = 38   # rows allocated per question block
        BAR_W, BAR_H   = 20, 13  # bar chart dimensions (optimized spacing)
        PIE_W, PIE_H   = 15, 13  # pie chart dimensions (optimized spacing)
        BAR_ANCHOR_COL = "D"     # bar chart left edge
        PIE_ANCHOR_COL = "U"     # pie chart left edge (well spaced from bar)

        # Page heading
        self.write_sheet_heading(
            ws,
            title="Charts  —  Feedback Distribution by Theme",
            subtitle=f"Generated: {now_eat}",
            merge_cols=26,
        )

        # Data area starts at row 4 (rows 1-2 = heading, row 3 = spacer)
        data_cursor = 4

        for q_idx, (qname, qdata) in enumerate(self.question_results.items()):
            s_df = qdata["summary"]
            safe_q = re.sub(r"[^\w ]", "", qname).strip().replace("_", " ").title()

            block_start = data_cursor  # anchor row for both charts

            # Section label above data
            lbl = ws.cell(row=block_start, column=1, value=safe_q)
            lbl.font = Font(bold=True, size=11, color="FFFFFF")
            lbl.fill = PatternFill(
                start_color=self.HEADER_COLOUR,
                end_color=self.HEADER_COLOUR,
                fill_type="solid",
            )
            ws.row_dimensions[block_start].height = 20

            # Column headers for data table
            ws.cell(row=block_start + 1, column=1, value="Theme")
            ws.cell(row=block_start + 1, column=2, value="Respondents")
            self.style_table_header(ws, header_row=block_start + 1)

            # Section 7: use Respondent Count; fall back to Response Count
            count_col = (
                "Respondent Count" if "Respondent Count" in s_df.columns
                else "Response Count"
            )

            # Data rows
            for offset, (_, r) in enumerate(
                s_df[["Theme", count_col]].iterrows(), start=2
            ):
                ws.cell(row=block_start + offset, column=1, value=r["Theme"])
                ws.cell(row=block_start + offset, column=2,
                        value=r[count_col])

            data_end = block_start + 1 + len(s_df)

            # References (header row included for title)
            data_ref = Reference(ws, min_col=2,
                                  min_row=block_start + 1, max_row=data_end)
            cats_ref = Reference(ws, min_col=1,
                                  min_row=block_start + 2, max_row=data_end)

            # ── Bar chart (Professional v3.0) ──────────────────────────
            bar = BarChart()
            bar.type    = "col"
            bar.style   = 11  # Modern professional style
            bar.title   = f"{safe_q}\nTheme Distribution"
            bar.y_axis.title = "Number of Responses"
            bar.x_axis.title = "Feedback Theme"
            bar.legend  = None          # No legend for cleaner look
            
            # Remove gridlines for professional appearance
            bar.y_axis.majorGridlines = None
            
            # Set axis scaling to start from 0 and be visible
            bar.y_axis.scaling.minVal = 0
            bar.y_axis.delete = False  # Ensure axis is visible
            bar.x_axis.delete = False  # Ensure axis is visible
            
            bar.add_data(data_ref, titles_from_data=True)
            bar.set_categories(cats_ref)
            bar.width  = BAR_W
            bar.height = BAR_H
            
            # Add data labels (values on top of bars) - openpyxl limitation
            # requires setting dataLbls XML attribute manually if using openpyxl
            # For now, chart shows cleanly with professional styling
            ws.add_chart(bar, f"{BAR_ANCHOR_COL}{block_start}")

            # ── Pie chart (Professional v3.0) ─────────────────────────
            pie = PieChart()
            pie.style   = 11  # Modern professional style
            pie.title   = f"{safe_q}\nPercentage Share"
            
            # Position legend to RIGHT to prevent overlap
            pie.legend.position = "r"  # 'r' = right
            
            pie.add_data(data_ref, titles_from_data=True)
            pie.set_categories(cats_ref)
            pie.width  = PIE_W
            pie.height = PIE_H
            
            # Anchor pie well to the right of the bar chart
            ws.add_chart(pie, f"{PIE_ANCHOR_COL}{block_start}")

            # Advance cursor by the fixed block height
            data_cursor += CHART_BLOCK_ROWS

        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 14

        # ══════════════════════════════════════════════════════════════
        # 3.5 ALERTS & ACTIONS (v3+ Decision Support Sheet)
        # ══════════════════════════════════════════════════════════════
        # This sheet surfaces HIGH/MEDIUM priority issues with action owners
        # and deadlines for immediate decision-making

        ws = wb.create_sheet("Alerts & Actions", 3)  # Insert after Charts

        next_row = self.write_sheet_heading(
            ws,
            title="🚨 PRIORITY ALERTS & ACTION PLANS",
            subtitle=f"Critical Issues Requiring Immediate Attention | {now_eat}",
            merge_cols=8,
        )
        next_row += 1

        # Collect all HIGH/MEDIUM priority issues across all questions
        priority_issues = []
        for qname, qdata in self.question_results.items():
            s_df = self.prepare_summary_for_excel(qdata["summary"])
            for idx, row in s_df.iterrows():
                triggers = row.get("triggers", [])
                if isinstance(triggers, list):
                    for trigger in triggers:
                        if trigger.get("level") in ("HIGH", "MEDIUM"):
                            # Section 7: use Respondent Count as the impact figure
                            impact_val = row.get(
                                "Respondent Count",
                                row.get("Response Count", 0)
                            )
                            priority_issues.append({
                                "question": qname,
                                "theme": row["Theme"],
                                "priority": trigger.get("icon", ""),
                                "level": trigger.get("level", ""),
                                "message": trigger.get("message", ""),
                                "impact": impact_val,
                                "owner": row.get("Action Owner", "TBD"),
                                "deadline_days": trigger.get("deadline_days", 30),
                                "recommendation": row.get("Recommendation", "N/A"),
                            })

        # Sort by level (HIGH first) then by impact
        priority_issues.sort(
            key=lambda x: (
                0 if x["level"] == "HIGH" else 1,
                -x["impact"]
            )
        )

        if not priority_issues:
            # No alerts
            msg_cell = ws.cell(row=next_row, column=1, 
                              value="✓ No HIGH or MEDIUM priority alerts detected.")
            msg_cell.font = Font(size=12, color="00B050", bold=True)
            ws.merge_cells(start_row=next_row, start_column=1,
                          end_row=next_row, end_column=8)
            next_row += 1
        else:
            # Build alerts table
            headers = ["Priority", "Theme", "Question", "Impact", "Message",
                      "Action Owner", "Deadline (Days)", "Recommendation"]
            for ci, h in enumerate(headers, 1):
                hc = ws.cell(row=next_row, column=ci, value=h)
                hc.font = Font(bold=True, color="FFFFFF", size=11)
                hc.fill = PatternFill(
                    start_color="2F5597", end_color="2F5597",
                    fill_type="solid"
                )
                hc.alignment = Alignment(horizontal="center", vertical="center")
                hc.border = self._thin_border()
            
            ws.row_dimensions[next_row].height = 20
            next_row += 1

            # Data rows
            for issue in priority_issues:
                ws.cell(row=next_row, column=1, value=issue["priority"])
                ws.cell(row=next_row, column=2, value=issue["theme"])
                ws.cell(row=next_row, column=3, value=issue["question"])
                ws.cell(row=next_row, column=4, value=issue["impact"])
                ws.cell(row=next_row, column=5, value=issue["message"])
                ws.cell(row=next_row, column=6, value=issue["owner"])
                ws.cell(row=next_row, column=7, value=issue["deadline_days"])
                ws.cell(row=next_row, column=8, value=issue["recommendation"])

                # Colour the priority column (dimmed colors, no icons)
                priority_cell = ws.cell(row=next_row, column=1)
                priority_cell.value = issue["level"]  # Show text level, not emoji
                if issue["level"] == "HIGH":
                    priority_cell.fill = PatternFill(
                        start_color="FFB3B3", end_color="FFB3B3",
                        fill_type="solid"
                    )
                    priority_cell.font = Font(color="404040", bold=True)
                elif issue["level"] == "MEDIUM":
                    priority_cell.fill = PatternFill(
                        start_color="FFE0B3", end_color="FFE0B3",
                        fill_type="solid"
                    )
                    priority_cell.font = Font(color="404040", bold=True)

                # Wrap long text columns
                for col_idx in [2, 5, 8]:  # Theme, Message, Recommendation
                    cell = ws.cell(row=next_row, column=col_idx)
                    cell.alignment = Alignment(
                        wrap_text=True, vertical="top"
                    )

                # Add borders
                for col in range(1, 9):
                    ws.cell(row=next_row, column=col).border = self._thin_border()

                ws.row_dimensions[next_row].height = 25
                next_row += 1

            # Set column widths
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 22
            ws.column_dimensions["C"].width = 18
            ws.column_dimensions["D"].width = 8
            ws.column_dimensions["E"].width = 35
            ws.column_dimensions["F"].width = 20
            ws.column_dimensions["G"].width = 14
            ws.column_dimensions["H"].width = 40

            # Add autofilter
            ws.auto_filter.ref = f"A{next_row - len(priority_issues)}:H{next_row}"

        # ══════════════════════════════════════════════════════════════
        # 4. EXAMPLE QUOTES
        # ══════════════════════════════════════════════════════════════
        ws = wb.create_sheet("Example Quotes")

        next_row = self.write_sheet_heading(
            ws,
            title="Example Quotes  —  Representative Responses by Theme",
            subtitle=f"Generated: {now_eat}",
            merge_cols=3,
        )

        # Column headers at row 3
        for ci, h in enumerate(["Question", "Theme", "Quote"], 1):
            ws.cell(row=next_row, column=ci, value=h)
        self.style_table_header(ws, header_row=next_row)
        data_start = next_row + 1

        for qname, qdata in self.question_results.items():
            cs = qdata.get("cluster_summary", {})
            for cid, info in cs.items():
                theme = self.generate_theme_name(info["keywords"])
                for q in info["samples"]:
                    ws.append([qname, theme, q])

        # Wrap quote column
        for row in ws.iter_rows(min_row=data_start, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        self.add_table_borders(ws, start_row=next_row)
        self.zebra_rows(ws, start_row=data_start)
        self._set_data_row_heights(ws, start_row=data_start, height=22)
        self.smart_col_widths(ws)

        # ══════════════════════════════════════════════════════════════
        # 5. EMERGING ISSUES
        # ══════════════════════════════════════════════════════════════
        emerging_rows = []
        for qname, qdata in self.question_results.items():
            em_df: pd.DataFrame = qdata.get("emerging_df", pd.DataFrame())
            if not em_df.empty:
                em_df = em_df.copy()
                em_df.insert(0, "Question", qname)
                emerging_rows.append(em_df)

        if emerging_rows:
            ws = wb.create_sheet("⚠ Emerging Issues")
            all_em = pd.concat(emerging_rows, ignore_index=True)

            next_row = self.write_sheet_heading(
                ws,
                title="⚠  Emerging Issues  —  Weak Signals Worth Watching",
                subtitle="Low-volume clusters that may be early-warning signs. "
                         "Investigate before they become widespread.",
                merge_cols=max(len(all_em.columns), 6),
            )

            for ci, col_name in enumerate(all_em.columns, 1):
                ws.cell(row=next_row, column=ci, value=col_name)
            self.style_table_header(ws, header_row=next_row)
            data_start = next_row + 1

            for r in all_em.itertuples(index=False):
                ws.append(list(r))

            self.add_table_borders(ws, start_row=next_row)
            self.zebra_rows(ws, start_row=data_start)
            self._set_data_row_heights(ws, start_row=data_start, height=18)
            self.smart_col_widths(ws)

        # ══════════════════════════════════════════════════════════════
        # 6. CLUSTERED RESPONSES (raw data)
        # ══════════════════════════════════════════════════════════════
        ws = wb.create_sheet("Clustered Responses")

        # Build comprehensive dataset with themes + metadata
        all_q_dfs = []
        for qname, qdata in self.question_results.items():
            q_df = qdata.get("q_df", pd.DataFrame()).copy()
            s_df = qdata.get("summary", pd.DataFrame()).copy()
            cluster_map = qdata.get("cluster_to_theme_map", {})
            
            if q_df.empty:
                continue
            
            # Add Question column
            q_df.insert(0, "Question", qname)
            
            # Rename columns for clarity
            if "clean_text" in q_df.columns:
                q_df.rename(columns={"clean_text": "Clean Text"}, inplace=True)
            elif "response" in q_df.columns:
                q_df.rename(columns={"response": "Clean Text"}, inplace=True)
            
            if "lang" in q_df.columns:
                q_df.rename(columns={"lang": "Language"}, inplace=True)
            
            if "cluster" in q_df.columns:
                q_df.rename(columns={"cluster": "Cluster"}, inplace=True)
            
            if "sentiment" in q_df.columns:
                q_df.rename(columns={"sentiment": "Sentiment"}, inplace=True)
            
            # Build lookup: theme_name → theme metadata from summary_df
            theme_lookup = {}
            if not s_df.empty:
                for _, row in s_df.iterrows():
                    theme_name = row.get("Theme", "Unknown")
                    theme_lookup[theme_name] = {
                        "Theme":            theme_name,
                        "Theme Sentiment":  row.get("Theme Sentiment", "Neutral"),
                        "Priority":         row.get("Priority", "Low"),
                        "Trigger Status":   row.get("Trigger Status", "No Alert"),
                        "Action Owner":     row.get("Action Owner", "Program Manager"),
                        "Timeline (Days)":  str(row.get("Timeline (Days)", "30")),
                        # Section 7: include respondent count in clustered sheet
                        "Respondent Count": row.get(
                            "Respondent Count",
                            row.get("Response Count", "")
                        ),
                    }
            
            # Map cluster IDs to themes using cluster_to_theme_map, then lookup metadata
            cluster_col = "Cluster" if "Cluster" in q_df.columns else "cluster"
            if cluster_col in q_df.columns:
                # Create a function to lookup theme by cluster ID then get metadata
                def get_theme_metadata(cluster_id):
                    theme_name = cluster_map.get(cluster_id, "Unclassified")
                    return theme_lookup.get(theme_name, {
                        "Theme": "Unclassified",
                        "Theme Sentiment": "Neutral",
                        "Priority": "Low",
                        "Trigger Status": "No Alert",
                        "Action Owner": "Program Manager",
                        "Timeline (Days)": "30",
                    })
                
                # Add theme columns to q_df
                for col in ["Theme", "Theme Sentiment", "Priority",
                            "Trigger Status", "Action Owner", "Timeline (Days)"]:
                    q_df[col] = q_df[cluster_col].apply(lambda cid: get_theme_metadata(cid).get(col, ""))
            
            all_q_dfs.append(q_df)

        if all_q_dfs:
            combined = pd.concat(all_q_dfs, ignore_index=True)
            
            # Reorder columns for professional layout
            desired_order = [
                "Question",
                "Clean Text",
                "Language",
                "Sentiment",
                "Cluster",
                "Theme",
                "Respondent Count",   # Section 7: theme-level count for context
                "Theme Sentiment",
                "Priority",
                "Trigger Status",
                "Action Owner",
                "Timeline (Days)",
            ]
            
            # Keep only columns that exist and exclude internal columns
            internal_cols = {"compressed_text"}  # Internal columns to hide
            cols_to_use = [col for col in desired_order if col in combined.columns]
            # Add any remaining columns not in desired_order (excluding internal)
            for col in combined.columns:
                if col not in cols_to_use and col not in internal_cols:
                    cols_to_use.append(col)
            
            combined = combined[cols_to_use]
            
            n_cols = len(combined.columns)

            next_row = self.write_sheet_heading(
                ws,
                title="Clustered Responses  —  Complete Analysis Dataset",
                subtitle=f"Total responses: {len(combined)}   |   "
                         f"Questions: {len(self.question_results)}",
                merge_cols=min(n_cols, 12),
            )

            # Write headers
            for ci, col_name in enumerate(combined.columns, 1):
                ws.cell(row=next_row, column=ci, value=col_name)
            self.style_table_header(ws, header_row=next_row)
            data_start = next_row + 1

            # Write data rows
            for r in combined.itertuples(index=False):
                ws.append(list(r))

            ws.auto_filter.ref = (
                f"A{next_row}:{get_column_letter(len(combined.columns))}"
                f"{ws.max_row}"
            )
            self.add_table_borders(ws, start_row=next_row)
            self.zebra_rows(ws, start_row=data_start)
            self._set_data_row_heights(ws, start_row=data_start, height=16)
            self.smart_col_widths(ws)

        # ══════════════════════════════════════════════════════════════
        # GEOGRAPHIC BREAKDOWN SHEET  (Section 5)
        # ══════════════════════════════════════════════════════════════
        geo = getattr(self, "geo_breakdown", {})
        if geo:
            ws = wb.create_sheet("Geographic Breakdown")

            next_row = self.write_sheet_heading(
                ws,
                title=f"Geographic Breakdown  —  by {self.config.location_column}",
                subtitle=f"Theme distribution per location | {now_eat}",
                merge_cols=10,
            )
            next_row += 1   # blank spacer row

            # Palette for heatmap intensity: white → navy
            def _heat_fill(count: int, max_count: int) -> PatternFill:
                """Return a fill shading from very light to mid navy."""
                if max_count == 0:
                    return PatternFill()
                ratio = min(count / max_count, 1.0)
                # Interpolate from white (255) to #2F5597 (47,85,151)
                r = int(255 - ratio * (255 - 47))
                g = int(255 - ratio * (255 - 85))
                b = int(255 - ratio * (255 - 151))
                hex_col = f"{r:02X}{g:02X}{b:02X}"
                return PatternFill(
                    start_color=hex_col, end_color=hex_col, fill_type="solid"
                )

            for qname, loc_data in geo.items():
                if not loc_data:
                    continue

                # ── Section header ────────────────────────────────────
                sec = ws.cell(
                    row=next_row, column=1,
                    value=qname.upper().replace("_", " "),
                )
                sec.font  = Font(bold=True, size=11, color="FFFFFF")
                sec.fill  = PatternFill(
                    start_color=self.HEADER_COLOUR,
                    end_color=self.HEADER_COLOUR,
                    fill_type="solid",
                )
                sec.border = self._thin_border()
                ws.merge_cells(
                    start_row=next_row, start_column=1,
                    end_row=next_row, end_column=10,
                )
                ws.row_dimensions[next_row].height = 22
                next_row += 1

                # ── Collect all theme names across locations ───────────
                all_themes_here: List[str] = []
                for ld in loc_data.values():
                    all_themes_here.extend(ld["theme_counts"].keys())
                # Rank themes by total occurrences, show top 6
                from collections import Counter
                theme_order = [
                    t for t, _ in
                    Counter(all_themes_here).most_common(6)
                ]

                # ── Column headers: Location | Top Issue | Avg Sent | themes… ─
                headers = (
                    ["Location", "Responses", "Top Issue", "Avg Sentiment"]
                    + theme_order
                )
                for ci, h in enumerate(headers, 1):
                    hc = ws.cell(row=next_row, column=ci, value=h)
                    hc.font      = Font(bold=True, color="FFFFFF", size=9)
                    hc.fill      = PatternFill(
                        start_color=self.SUBHEAD_COLOUR,
                        end_color=self.SUBHEAD_COLOUR,
                        fill_type="solid",
                    )
                    hc.font      = Font(bold=True, color="000000", size=9)
                    hc.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                    hc.border = self._thin_border()
                ws.row_dimensions[next_row].height = 24
                next_row += 1
                data_start = next_row

                # Max count for heatmap scaling (per question)
                max_count = max(
                    (
                        tc
                        for ld in loc_data.values()
                        for tc in ld["theme_counts"].values()
                    ),
                    default=1,
                )

                # ── Data rows ─────────────────────────────────────────
                for ri, (loc_val, ld) in enumerate(
                    sorted(loc_data.items(), key=lambda x: -x[1]["total"])
                ):
                    sent  = ld["avg_sentiment"]
                    # Sentiment colour for the avg sentiment cell
                    if sent <= -0.05:
                        sent_colour = "FFE0E0"   # soft red
                    elif sent >= 0.05:
                        sent_colour = "E0FFE8"   # soft green
                    else:
                        sent_colour = "FFF9E0"   # soft yellow

                    row_vals = (
                        [loc_val, ld["total"], ld["top_issue"], round(sent, 3)]
                        + [ld["theme_counts"].get(t, 0) for t in theme_order]
                    )
                    for ci, val in enumerate(row_vals, 1):
                        c = ws.cell(row=next_row, column=ci, value=val)
                        c.border    = self._thin_border()
                        c.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                        # Heatmap shading for theme count columns
                        if ci > 4 and isinstance(val, int) and val > 0:
                            c.fill = _heat_fill(val, max_count)
                            # White text when background is dark enough
                            if val / max_count > 0.55:
                                c.font = Font(color="FFFFFF", size=9)
                        elif ci == 4:   # Avg Sentiment column
                            c.fill = PatternFill(
                                start_color=sent_colour,
                                end_color=sent_colour,
                                fill_type="solid",
                            )
                        elif ci == 3:   # Top Issue column — bold
                            c.font = Font(bold=True, size=9)

                    ws.row_dimensions[next_row].height = 18
                    next_row += 1

                # Auto-filter on the header row
                ws.auto_filter.ref = (
                    f"A{data_start - 1}:"
                    f"{get_column_letter(len(headers))}{next_row - 1}"
                )
                next_row += 2   # two blank rows before next question

            self.smart_col_widths(ws)
            ws.column_dimensions["A"].width = 20   # Location
            ws.column_dimensions["C"].width = 28   # Top Issue

        wb.save(output)
        log.info("✅ Report saved → %s", output)

    # ═══════════════════════════════════════════════════════════════════
    # JSON & ALERTS EXPORT (v3+ Feature)
    # ═══════════════════════════════════════════════════════════════════

    def _export_json_reports(self) -> None:
        """Export insights in machine-readable JSON format."""
        try:
            from output.json_writer import export_json
            
            # Collect all themes from all questions
            all_themes = []
            for qname, qdata in self.question_results.items():
                s_df = qdata["summary"]
                for idx, row in s_df.iterrows():
                    # Section 7: use Respondent Count as impact
                    impact = int(row.get(
                        "Respondent Count",
                        row.get("Response Count", 0)
                    ))
                    theme_dict = {
                        "question":        qname,
                        "name":            row.get("Theme", ""),
                        "impact":          impact,
                        "respondent_count":impact,
                        "sentence_count":  int(row.get("Sentence Count", 0)),
                        "sentiment":       row.get("Theme Sentiment", "Neutral"),
                        "keywords":        str(row.get("Key Keywords", "")).split(","),
                        "triggers":        row.get("triggers", []),
                        "action_plan": (
                            row.get("action_plan", {})
                            if isinstance(row.get("action_plan"), dict)
                            else {}
                        ),
                        "recommendation":  row.get("Recommendation", ""),
                    }
                    all_themes.append(theme_dict)
            
            export_json(
                all_themes,
                filename="output/insights.json",
                sector=self.pipeline_config.get("sector", ""),
            )
        except ImportError:
            log.warning("json_writer not available — skipping JSON export")
        except Exception as exc:
            log.error("JSON export failed: %s", exc)

    def _export_alerts_report(self) -> None:
        """Export priority alerts in machine-readable format."""
        try:
            from output.alert_formatter import generate_alerts, export_alerts_json, print_alerts
            
            # Collect all HIGH/MEDIUM priority alerts
            all_alerts = []
            for qname, qdata in self.question_results.items():
                s_df = qdata["summary"]
                for idx, row in s_df.iterrows():
                    triggers = row.get("triggers", [])
                    if isinstance(triggers, list):
                        for trigger in triggers:
                            if trigger.get("level") in ("HIGH", "MEDIUM"):
                                # Section 7: use Respondent Count
                                impact = row.get(
                                    "Respondent Count",
                                    row.get("Response Count", 0)
                                )
                                alert = {
                                    "theme_name":    row.get("Theme", ""),
                                    "question":      qname,
                                    "priority_icon": trigger.get("icon", ""),
                                    "priority_level":trigger.get("level", ""),
                                    "message":       trigger.get("message", ""),
                                    "deadline_days": trigger.get("deadline_days", 30),
                                    "impact_count":  impact,
                                    "recommendation":row.get("Recommendation", ""),
                                    "action_owner":  row.get("Action Owner", "TBD"),
                                }
                                all_alerts.append(alert)
            
            if all_alerts:
                # Print to console for immediate visibility
                print_alerts(all_alerts, max_display=5)
                
                # Export to JSON
                export_alerts_json(all_alerts, filename="output/alerts.json")
            else:
                log.info("✓ No HIGH/MEDIUM alerts to export")
        except ImportError:
            log.warning("alert_formatter not available — skipping alerts export")
        except Exception as exc:
            log.error("Alerts export failed: %s", exc)

    # ══════════════════════════════════════════════════════════════════
    # FULL PIPELINE
    # ══════════════════════════════════════════════════════════════════

    def run(self) -> None:
        # Reset token review accumulator at the start of each full run
        self._token_review_data = {}

        for column in self.text_columns:
            log.info("━━━ Analysing: %s ━━━", column)

            # --- Validate column ---
            if column not in self.df.columns:
                log.warning("Column '%s' not found — skipping.", column)
                continue

            raw_series = (
                self.df[column]
                .dropna()
                .astype(str)
                .str.strip()
            )
            raw_series = raw_series[raw_series.str.len() > 3]
            self._question_response_count = len(raw_series)

            if self._question_response_count < self.config.min_responses:
                log.warning(
                    "Column '%s' has only %d usable responses (min=%d) — skipping.",
                    column, self._question_response_count, self.config.min_responses,
                )
                continue

            # --- Build per-question df ---
            # ── Section 5: keep orig_idx so location can be joined back later.
            # raw_series.index = original row positions in self.df — critical
            # for the geographic breakdown join.  reset_index(drop=True) would
            # destroy this link, so we capture it as a column first.
            raw_df = raw_series.to_frame(name="response")
            raw_df["orig_idx"] = raw_df.index
            self.q_df = raw_df.reset_index(drop=True)

            # ── Section 1: track which column is being processed ──────
            self._current_question_col = column

            log.info("Preprocessing …")
            self.preprocess()

            log.info("Deduplicating near-identical sentences …")
            self.deduplicate_responses()

            log.info("Building TF-IDF …")
            self.build_tfidf_matrix()

            log.info("Sentiment analysis …")
            self.compute_sentiment()

            log.info("Generating embeddings + PCA …")
            self.generate_embeddings()

            log.info("Clustering …")
            self.cluster_feedback()

            log.info("Filtering clusters …")
            self.filter_clusters()

            log.info("Building cluster summary …")
            self.build_cluster_summary()

            log.info("Building summary table …")
            self.build_summary_table()

            log.info("Classifying theme sentiment …")
            self.classify_theme_sentiment()

            log.info("Applying advanced engines (triggers/actions) …")
            self.apply_advanced_engines()

            log.info("Generating recommendations …")
            self.generate_recommendations()

            log.info("Merging similar themes …")
            self.merge_similar_themes()

            log.info("Generating insights …")
            self.generate_key_insights()

            log.info("Dataset summary …")
            self.generate_dataset_summary()

            log.info("Labelling responses …")
            self.label_data()

            # --- Store per-question results ---
            self.question_results[column] = {
                "summary": self.summary_df.copy(),
                "insights": self.key_insights.copy(),
                "dataset_summary": self.dataset_summary.copy(),
                "cluster_summary": self.cluster_summary.copy(),
                "emerging_df": getattr(self, "emerging_df", pd.DataFrame()).copy(),
                "q_df": self.q_df.copy(),
                "cluster_to_theme_map": self.cluster_to_theme_map.copy(),
            }

            # Free per-question RAM
            del self.embeddings
            gc.collect()

        # --- Cross-question analysis ---
        if len(self.question_results) > 1:
            log.info("Detecting cross-question themes …")
            self.detect_cross_question_themes()

        # ── Section 5: geographic breakdown ───────────────────────────
        if self.config.location_column:
            log.info("Building geographic breakdown …")
            self.build_geographic_breakdown()

        # --- Keyword coverage tips (always printed when rule-based recs used) ---
        # If LLM recs were used, print a condensed keyword discovery summary.
        if (self._anthropic_client or self._gemini_client) and self.config.use_llm_recommendations:
            self._print_llm_keyword_summary()

        # ── Section 8: Persist theme registry to disk ─────────────────
        self._save_theme_registry()

        # ── NEW: Generate JSON + Alerts output (v3+ Upgrade) ──────────
        self._export_outputs()

        log.info("Building Excel report …")
        self.build_excel_report()

        # ── Section 4: PDF executive brief ────────────────────────────
        if self.config.generate_pdf_report:
            log.info("Building PDF summary …")
            self.build_pdf_report()

    def _export_outputs(self) -> None:
        """
        Export insights as JSON and generate priority alerts.
        Controlled by config/config.yaml output settings.
        """
        config = self.pipeline_config
        output_config = config.get("output", {})
        sector = config.get("sector", "")
        
        # Collect all themes from all questions
        all_themes = []
        for qname, qdata in self.question_results.items():
            s_df = qdata.get("summary", pd.DataFrame())
            
            # DEBUG: Check what columns exist in summary_df
            if all_themes == []:  # Only log once for first question
                log.info(f"  📊 Summary columns: {list(s_df.columns)}")
            
            for idx, row in s_df.iterrows():
                # Extract trigger status from raw triggers list
                triggers = row.get("triggers", [])
                trigger_status = "Low"
                if isinstance(triggers, list) and triggers:
                    # Get the highest priority trigger level
                    levels = {t.get("level", "Low") for t in triggers}
                    if "High" in levels:
                        trigger_status = "High"
                    elif "Medium" in levels:
                        trigger_status = "Medium"
                    elif "Positive" in levels:
                        trigger_status = "Positive"
                
                # Extract action owner and timeline from action_plan dict
                action_plan = row.get("action_plan", {})
                
                # Ensure action_plan is converted from pandas object to dict
                if isinstance(action_plan, dict):
                    assigned_owner = action_plan.get("owner", "Unknown")
                    timeline_days = action_plan.get("timeline", 30)
                else:
                    assigned_owner = "Unknown"
                    timeline_days = 30
                
                # Convert triggers to list of dicts (in case it's a pandas object)
                triggers_list = []
                if isinstance(triggers, list):
                    triggers_list = triggers
                elif pd.notna(triggers):
                    triggers_list = [] if triggers is None else [triggers]
                
                # DEBUG: Log first theme details
                if idx == 0 and qname == list(self.question_results.keys())[0]:
                    log.info(f"  🎯 First theme export check:")
                    log.info(f"     Theme: {row.get('Theme', 'N/A')}")
                    log.info(f"     action_plan type: {type(action_plan)}")
                    log.info(f"     action_plan value: {action_plan}")
                    log.info(f"     owner: {assigned_owner}")
                
                theme_dict = {
                    "question": qname,
                    "name": row.get("Theme", ""),
                    # Section 7: use respondent count as canonical impact figure
                    "impact": row.get(
                        "Respondent Count",
                        row.get("Response Count", 0)
                    ),
                    "respondent_count": row.get(
                        "Respondent Count",
                        row.get("Response Count", 0)
                    ),
                    "sentence_count": row.get("Sentence Count", 0),
                    "sentiment": row.get("Theme Sentiment", "Neutral"),
                    "keywords": (
                        [k.strip() for k in str(row.get("Key Keywords", "")).split(",")]
                        if pd.notna(row.get("Key Keywords"))
                        else []
                    ),
                    "average_sentiment_score": row.get("Average Sentiment", 0),
                    "recommendation": row.get("Recommendation", ""),
                    "trigger_status": trigger_status,
                    "assigned_owner": assigned_owner,
                    "timeline_days": timeline_days,
                    "triggers": triggers_list,
                    "action_plan": action_plan if isinstance(action_plan, dict) else {},
                }
                all_themes.append(theme_dict)
        
        # Export JSON if enabled
        if output_config.get("json", True):
            try:
                from output.json_writer import export_json
                export_json(
                    all_themes,
                    filename="output/insights.json",
                    sector=sector,
                )
            except ImportError:
                log.warning("JSON writer not available. Skipping JSON export.")
        
        # Generate and display alerts if enabled
        if output_config.get("alerts", True):
            try:
                from output.alert_formatter import generate_alerts, print_alerts, export_alerts_json
                alerts = generate_alerts(all_themes)
                print_alerts(alerts, max_display=10)
                export_alerts_json(alerts, filename="output/alerts.json")
            except ImportError:
                log.warning("Alert formatter not available. Skipping alerts.")

    def _print_llm_keyword_summary(self) -> None:
        """Print a brief keyword discovery log when LLM recommendations are active."""
        print(f"\n{'═' * 62}")
        print("  KEYWORD DISCOVERY SUMMARY  (LLM recommendations active)")
        print(f"{'═' * 62}")
        for qname, qdata in self.question_results.items():
            s_df = qdata["summary"]
            all_kw = set()
            for kw_str in s_df.get("Key Keywords", []):
                all_kw.update(k.strip() for k in str(kw_str).split(","))
            print(f"\n  {qname}:")
            print(f"    Themes     : {', '.join(s_df['Theme'].tolist())}")
            print(f"    All keywords found : {', '.join(sorted(all_kw))}")
        print(f"\n  TIP: If recommendations look off for any theme,")
        print(f"  set use_llm_recommendations=False and check the")
        print(f"  KEYWORD TIPS report to tune RECOMMENDATION_RULES.")
        print(f"{'═' * 62}\n")


# ══════════════════════════════════════════════════════════════════════
# ENTRY POINT  ─ configure and run
# ══════════════════════════════════════════════════════════════════════

if __name__ == "__main__":

    config = EngineConfig(
        pca_components=64,
        min_cluster_size=5,
        emerging_min_size=2,
        max_clusters=20,
        merge_threshold=0.75,
        tfidf_max_features=5_000,
        large_dataset_threshold=500,
        min_responses=10,

        # ── Embedding model (Section 2) ───────────────────────────────
        # Multilingual model — handles Swahili, Sheng, English natively.
        # Change to "all-MiniLM-L6-v2" for English-only data (slightly faster).
        embedding_model="paraphrase-multilingual-MiniLM-L12-v2",

        # ── Sector context ────────────────────────────────────────────
        # Set this to your programme area for better LLM outputs:
        #   "smallholder agriculture"  |  "primary healthcare"
        #   "microfinance / SACCO"     |  "WASH services"
        #   "secondary education"      |  "urban housing"
        sector="smallholder agriculture",  # ← change for your sector

        # ── LLM features (Section 10 — Gemini for Google Colab) ──────
        #
        # OPTION A: Use Gemini (free, Google Colab)
        #   1. Run this in a Colab cell BEFORE starting the engine:
        #        import google.generativeai as genai
        #        genai.configure(api_key="YOUR_GEMINI_API_KEY")
        #      OR set the GEMINI_API_KEY environment variable.
        #   2. pip install google-generativeai
        #   3. Set use_gemini=True below.
        #
        # OPTION B: Use Claude/Anthropic (paid, local machine)
        #   1. pip install anthropic
        #   2. Set ANTHROPIC_API_KEY environment variable.
        #   3. Keep use_gemini=False (default).
        #
        # OPTION C: No LLM (fully offline, rule-based)
        #   Set use_llm_naming=False, use_llm_recommendations=False.
        #
        use_gemini=True,               # ← True for Google Colab / Gemini
        gemini_model="gemini-2.0-flash",
        use_llm_naming=True,           # meaningful theme names via LLM
        use_llm_recommendations=True,  # sector-aware recommendations via LLM

        report_title="One Acre Fund — Farmer Feedback Insights 2026",
        output_filename="oaf_farmer_survey_insight_report.xlsx",

        # ── PDF summary (Section 4) ───────────────────────────────────
        # Requires: pip install fpdf2
        # Set False to skip PDF and produce only the Excel report.
        generate_pdf_report=True,

        # ── Geographic Segmentation (Section 5) ──────────────────────
        # Column in your data that holds location labels.
        # Set to "" to disable geographic breakdown.
        location_column="county",
        location_min_responses=10,

        # ── Theme Registry (Section 8) ────────────────────────────────
        # theme_registry.json is auto-created on first run.
        # Edit it manually to rename any theme permanently.
        theme_registry_path="theme_registry.json",
    )

    engine = FeedbackInsightEngine(
        file_path="oaf_farmer_survey_demo.xlsx",
        text_columns=[
            "q1_inputs",
            "q2_training",
            "q3_fieldofficer",
            "q4_loan",
            "q5_suggestions",
        ],
        config=config,
    )

    engine.run()